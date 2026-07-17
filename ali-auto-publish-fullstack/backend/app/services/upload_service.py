# -*- coding: utf-8 -*-
"""
产品上传服务层
重构自: 产品上传/main_属性融合.py

【架构说明】
本模块将原始脚本的 AlibabaPoster 类拆分为可被 API 调用的服务函数。
核心 Selenium 自动化逻辑保留在 automation/ 子模块中，本层只负责：
1. 任务编排（启动、暂停、停止）
2. 数据准备（标题加载、产品扫描、已发布记录管理）
3. 状态上报（通过 TaskInfo 更新进度）

【如何修改】
- 修改属性填写逻辑 → 编辑 automation/attribute_filler.py
- 修改图片上传逻辑 → 编辑 automation/image_uploader.py
- 修改价格设置逻辑 → 编辑 automation/price_setter.py
- 修改标题/关键词逻辑 → 编辑 automation/title_manager.py
- 增减属性配置 → 通过前端配置页面或直接编辑 data/config.json
"""
import os
import re
import time
import gc
import logging
from datetime import datetime, timedelta
from typing import Any, Dict, List, Optional, Set, Tuple

import pandas as pd

from app.core.settings import get_config, config_manager
from app.core.task_manager import TaskInfo
from app.core.logger import setup_logger

logger = setup_logger("upload_service")


# ===================== 数据管理函数 =====================

def load_published_products() -> Set[str]:
    """加载已发布产品ID集合"""
    cfg = get_config()
    published = set()
    filepath = cfg.paths.published_products_file
    if os.path.exists(filepath):
        with open(filepath, "r", encoding="utf-8") as f:
            for line in f:
                line = line.strip()
                if line:
                    published.add(line)
    return published


def save_published_product(pid: str):
    """记录已发布产品"""
    cfg = get_config()
    with open(cfg.paths.published_products_file, "a", encoding="utf-8") as f:
        f.write(f"{pid}\n")


def load_used_titles() -> Set[str]:
    """加载已使用的主标题"""
    cfg = get_config()
    used = set()
    filepath = cfg.paths.used_titles_file
    if os.path.exists(filepath):
        with open(filepath, "r", encoding="utf-8") as f:
            for line in f:
                line = line.strip()
                if line:
                    used.add(line)
    return used


def save_new_link_product(primary_id: str, house_type: str = ""):
    """将最新发品插入到首条数据下方，同时保持原有内容不被修改。文件不存在时自动新建。"""
    cfg = get_config()
    file_path = (cfg.data_analysis.new_links_file_path or "").strip()
    sheet_name = (cfg.data_analysis.new_links_sheet_name or "新链接").strip()
    col_name = (cfg.data_analysis.new_links_column_name or "新发链接").strip()

    if not file_path:
        logger.warning("未配置新发链接文件路径，跳过写入")
        return

    parent_dir = os.path.dirname(file_path)
    if parent_dir:
        os.makedirs(parent_dir, exist_ok=True)

    date_col = "发品日期"
    type_col = "类型"
    today = datetime.now().strftime("%y%m%d")
    pid = str(primary_id).strip()
    htype = str(house_type or "").strip()
    required_headers = [date_col, col_name, type_col]

    try:
        from openpyxl import Workbook, load_workbook

        is_new_file = not os.path.exists(file_path)
        if is_new_file:
            wb = Workbook()
            ws = wb.active
            ws.title = sheet_name
            ws.append(required_headers)
            header_map = {h: i + 1 for i, h in enumerate(required_headers)}
        else:
            wb = load_workbook(file_path)
            if sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
            else:
                ws = wb.create_sheet(title=sheet_name)

            header_map = {}
            if ws.max_row >= 1:
                for idx, cell in enumerate(ws[1], start=1):
                    name = str(cell.value).strip() if cell.value is not None else ""
                    if name and name not in header_map:
                        header_map[name] = idx

            if not header_map:
                ws.append(required_headers)
                header_map = {h: i + 1 for i, h in enumerate(required_headers)}
            else:
                next_col = max(header_map.values())
                for h in required_headers:
                    if h not in header_map:
                        next_col += 1
                        ws.cell(row=1, column=next_col, value=h)
                        header_map[h] = next_col

        pid_col_idx = header_map[col_name]
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=pid_col_idx, max_col=pid_col_idx):
            if str(row[0].value or "").strip() == pid:
                logger.info(f"新发链接已存在，跳过重复写入: {pid}")
                return

        insert_row = 2 if ws.max_row >= 1 else 1
        if ws.max_row >= 1:
            ws.insert_rows(insert_row, 1)

        ncol = max(header_map.values())
        row_values = [""] * ncol
        row_values[header_map[date_col] - 1] = today
        row_values[header_map[col_name] - 1] = pid
        row_values[header_map[type_col] - 1] = htype
        for col_idx, value in enumerate(row_values, start=1):
            ws.cell(row=insert_row, column=col_idx, value=value)

        if "Sheet" in wb.sheetnames and len(wb.sheetnames) > 1 and wb["Sheet"].max_row == 0:
            del wb["Sheet"]

        wb.save(file_path)
        if is_new_file:
            logger.info(f"已新建新发链接文件并写入: {file_path} | {pid} ({today})")
        else:
            logger.info(f"已将新发链接插入顶部: {pid} ({today})")
    except Exception as e:
        logger.error(f"写入新发链接监控失败: {e}")


def save_used_title(title: str):
    """记录已使用的标题"""
    cfg = get_config()
    with open(cfg.paths.used_titles_file, "a", encoding="utf-8") as f:
        f.write(f"{title}\n")


def natural_key(s: str):
    """自然排序键"""
    return [int(c) if c.isdigit() else c.lower() for c in re.split(r'(\d+)', s)]


def parse_primary_filename(filename: str) -> Tuple[Optional[str], str, str]:
    """解析首图文件名：日期-编号-分组-标题场景.jpg"""
    base = os.path.splitext(filename)[0]
    parts = base.split('-', 3)
    pid = f"{parts[0]}-{parts[1]}" if len(parts) >= 2 else None
    category = parts[2].strip() if len(parts) >= 3 else ""
    title_scene = parts[3].strip() if len(parts) >= 4 else ""
    return pid, category, title_scene


def parse_house_type_from_main2_filename(pid: str, cfg=None) -> str:
    """从主图目录的“第一张主图”文件名提取类型。

    规则：
    - 取主图目录内自然排序后的第一张图片文件名（通常是 2-*.jpg）
    - 以 '-' 分割后，取“最后一个字段”作为类型

    示例：
    - 2-玻璃幕墙.jpg -> 玻璃幕墙
    - 2-X-玻璃幕墙.jpg -> 玻璃幕墙
    """
    if not pid:
        return ""
    if cfg is None:
        cfg = get_config()

    main_dir = os.path.join(cfg.paths.main_image_dir, pid)
    if not os.path.isdir(main_dir):
        return ""

    candidates = sorted(
        [f for f in os.listdir(main_dir) if f.lower().endswith((".jpg", ".jpeg", ".png", ".webp"))],
        key=natural_key,
    )
    if not candidates:
        return ""

    first_main = candidates[0]
    name_no_ext = os.path.splitext(first_main)[0]
    tokens = [t.strip() for t in name_no_ext.split("-") if t.strip()]

    # 至少需要“序号 + 类型”两段，避免像 "2.jpg" 这种无类型命名
    if len(tokens) < 2:
        return ""

    return tokens[-1]


def _detect_title_excel_header_row_from_xf(xf: pd.ExcelFile, sheet_name: str = "产品标题", max_scan_rows: int = 10) -> int:
    """在已打开的 ExcelFile 对象上自动识别表头行（0-based）。"""
    try:
        preview = xf.parse(sheet_name=sheet_name, header=None, nrows=max_scan_rows)
    except Exception:
        return 0
    if preview is None or preview.empty:
        return 0
    scene_names = {"场景", "标题场景"}
    title_names = {"标题", "标题打乱重组"}
    for idx in range(len(preview)):
        raw_vals = [
            str(v).strip()
            for v in preview.iloc[idx].tolist()
            if pd.notna(v) and str(v).strip()
        ]
        if not raw_vals:
            continue
        has_scene = any(v in scene_names for v in raw_vals)
        has_title = any(v in title_names for v in raw_vals)
        if has_scene and has_title:
            return idx
    return 0


def _detect_title_excel_header_row(path: str, sheet_name: str = "产品标题", max_scan_rows: int = 10) -> int:
    """自动识别标题 Excel 表头行（0-based），兼容首行为空或说明行的情况。"""
    try:
        with pd.ExcelFile(path) as xf:
            return _detect_title_excel_header_row_from_xf(xf, sheet_name, max_scan_rows)
    except Exception:
        return 0


def _read_title_excel(path: str, nrows: Optional[int] = None) -> pd.DataFrame:
    """
    使用 with pd.ExcelFile(...) 上下文管理器读取标题 Excel。
    ExcelFile 在 with 块退出时会立即关闭底层文件句柄，
    不会产生任何临时文件，也不会残留句柄。
    """
    if not os.path.exists(path):
        raise FileNotFoundError(f"标题Excel不存在: {path}")
    with pd.ExcelFile(path) as xf:
        header_row = _detect_title_excel_header_row_from_xf(xf)
        kwargs: Dict[str, Any] = {
            "sheet_name": "产品标题",
            "dtype": str,
            "header": header_row,
        }
        if nrows is not None:
            kwargs["nrows"] = nrows
        return xf.parse(**kwargs)


def get_title_excel_health() -> Dict[str, Any]:
    """检查标题 Excel 是否可读（供状态接口提示用户）。"""
    import pandas as pd

    cfg = get_config()
    path = str(getattr(cfg.paths, "title_excel_path", "") or "").strip()
    if not path:
        return {"ok": False, "message": "未配置 paths.title_excel_path"}
    if not os.path.exists(path):
        return {"ok": False, "message": f"标题 Excel 不存在: {path}"}
    try:
        df = _read_title_excel(path, nrows=3)
        scene_col = "场景" if "场景" in df.columns else ("标题场景" if "标题场景" in df.columns else "")
        title_col = "标题" if "标题" in df.columns else ("标题打乱重组" if "标题打乱重组" in df.columns else "")
        if not scene_col or not title_col:
            return {
                "ok": False,
                "message": "Excel 缺少标题列，需包含：场景/标题（或旧版：标题场景/标题打乱重组）",
                "path": path,
                "columns": list(df.columns),
            }
        return {"ok": True, "message": "", "path": path, "scene_col": scene_col, "title_col": title_col}
    except Exception as e:
        return {"ok": False, "message": f"标题 Excel 读取失败: {e}", "path": path}


def load_titles_from_excel() -> Dict[str, List[Dict[str, str]]]:
    """
    从 Excel 加载所有标题（仅主标题）
    返回: {标题场景: [{"main": "主标题", "sub": "副标题"}, ...]}
    """
    import pandas as pd
    cfg = get_config()
    titles_by_scene: Dict[str, List[Dict[str, str]]] = {}

    try:
        df = _read_title_excel(cfg.paths.title_excel_path)
        # 新格式：场景 / 标题
        # 兼容旧格式：标题场景 / 标题打乱重组
        scene_col = "场景" if "场景" in df.columns else ("标题场景" if "标题场景" in df.columns else "")
        title_col = "标题" if "标题" in df.columns else ("标题打乱重组" if "标题打乱重组" in df.columns else "")
        if not scene_col or not title_col:
            logger.error("Excel 缺少标题列，需包含：场景/标题（或旧版：标题场景/标题打乱重组）")
            return {}
        df = df.dropna(subset=[scene_col, title_col])
        for _, row in df.iterrows():
            scene = str(row[scene_col]).strip()
            content = str(row[title_col]).strip()
            if not scene or not content:
                continue
            titles_by_scene.setdefault(scene, []).append({"main": content, "sub": ""})

        logger.info(f"成功加载标题，共 {sum(len(v) for v in titles_by_scene.values())} 组")
        return titles_by_scene
    except Exception as e:
        logger.error(f"标题加载失败: {e}")
        return {}


def load_keywords_from_excel() -> List[str]:
    """从 Excel 加载关键词，使用 with pd.ExcelFile 确保句柄及时释放。"""
    cfg = get_config()
    path = str(getattr(cfg.paths, "title_excel_path", "") or "").strip()
    if not path or not os.path.exists(path):
        return []
    try:
        with pd.ExcelFile(path) as xf:
            if "关键词" not in xf.sheet_names:
                return []
            df = xf.parse(sheet_name="关键词", dtype=str)
        if "关键词" not in df.columns:
            return []
        keywords = df["关键词"].dropna().astype(str).str.strip().tolist()
        return [k for k in keywords if k and k.lower() != "nan"]
    except Exception:
        return []


# ===================== 产品扫描 =====================

def scan_available_products() -> List[Dict]:
    """扫描可发布的产品"""
    cfg = get_config()
    published = load_published_products()
    used_titles = load_used_titles()
    titles_by_scene = load_titles_from_excel()

    products = []
    primary_dir = cfg.paths.primary_image_dir
    if not os.path.isdir(primary_dir):
        return products

    primary_files = sorted(
        [f for f in os.listdir(primary_dir) if f.lower().endswith(('.jpg', '.png'))],
        key=natural_key
    )

    for f in primary_files:
        pid, category, title_scene = parse_primary_filename(f)
        if not pid:
            continue

        is_published = pid in published
        available_titles = [
            pair for pair in titles_by_scene.get(title_scene, [])
            if pair["main"] not in used_titles
        ]

        # 检查主图目录
        main_dir = os.path.join(cfg.paths.main_image_dir, pid)
        main_image_count = len([
            img for img in os.listdir(main_dir) if img.lower().endswith(('.jpg', '.png'))
        ]) if os.path.isdir(main_dir) else 0

        products.append({
            "pid": pid,
            "filename": f,
            "category": category,
            "title_scene": title_scene,
            "is_published": is_published,
            "available_titles": len(available_titles),
            "main_image_count": main_image_count,
            "can_publish": not is_published and len(available_titles) > 0 and main_image_count > 0,
        })

    return products


def get_available_products_list() -> List[Dict]:
    """获取可发布产品列表（供 API 调用）"""
    all_products = scan_available_products()
    return [p for p in all_products if p["can_publish"]]


def get_published_products_list() -> List[Dict]:
    """获取已发布产品列表（供 API 调用）"""
    all_products = scan_available_products()
    return [p for p in all_products if p["is_published"]]


# ===================== 核心发布任务 =====================

def run_upload_task(task: TaskInfo, mode: str = "batch", max_products: Optional[int] = None, scheduled_time: Optional[str] = None):
    """
    自动发品主任务（由 TaskManager 在后台线程中调用）

    【重要】此函数是任务编排层，实际的浏览器自动化操作
    委托给 automation/ 子模块中的各个专门函数。

    Args:
        task: TaskInfo 对象，用于上报进度和检查停止信号
        mode: 发布模式 - batch(批量) / single(单个) / scheduled(定时)
        max_products: 最大发布数量（None 表示不限制）
    """
    cfg = get_config()
    logger.info(f"自动发品任务启动 - 模式: {mode}, 最大数量: {max_products}, 定时时间: {scheduled_time}")

    task.current_step = "正在初始化..."
    task.progress = 0

    try:
        if mode == "daily_scheduled":
            task.current_step = "每日定时循环模式已启动"
            while not task.should_stop():
                # 每轮按定时时间等待并执行一次发布
                run_upload_task(task, mode="scheduled", max_products=max_products, scheduled_time=scheduled_time)
                if task.should_stop():
                    break
                task.current_step = "本轮发布完成，持续定时模式等待下一轮"
                time.sleep(2)
            return

        if mode == "scheduled":
            cfg_sched = cfg.schedule
            target_text = (scheduled_time or "").strip()
            if not target_text:
                # 兼容旧配置窗口 "22:00-1:00"，取起始时间
                target_text = str(cfg_sched.publish_time_window or "").split("-")[0].strip()
            if not re.match(r"^\d{1,2}:\d{2}$", target_text):
                raise Exception("定时发品时间格式错误，请使用 HH:MM")

            hh, mm = target_text.split(":", 1)
            h = int(hh)
            m = int(mm)
            if h < 0 or h > 23 or m < 0 or m > 59:
                raise Exception("定时发品时间超出范围，请使用 00:00~23:59")

            now = datetime.now()
            target_dt = now.replace(hour=h, minute=m, second=0, microsecond=0)
            if target_dt <= now:
                target_dt = target_dt + timedelta(days=1)

            while True:
                if task.should_stop():
                    logger.info("收到停止信号，取消定时发品")
                    return
                remain = int((target_dt - datetime.now()).total_seconds())
                if remain <= 0:
                    break
                mins, secs = divmod(remain, 60)
                task.current_step = f"定时发品等待中（目标 {target_text}），剩余 {mins:02d}:{secs:02d}"
                time.sleep(1)
        # 1. 加载标题数据
        task.current_step = "加载标题数据..."
        titles_by_scene = load_titles_from_excel()
        if not titles_by_scene:
            task.current_step = "标题数据加载失败"
            raise Exception("无法加载标题数据，请检查 Excel 文件")

        # 2. 加载关键词
        keywords = load_keywords_from_excel()
        logger.info(f"已加载 {len(keywords)} 个关键词")

        # 3. 扫描可发布产品
        task.current_step = "扫描可发布产品..."
        available = get_available_products_list()
        if not available:
            task.current_step = "没有可发布的产品"
            logger.info("没有可发布的产品")
            return

        # 限制数量
        if max_products and max_products > 0:
            available = available[:max_products]
        elif cfg.upload.max_products_per_run > 0:
            available = available[:cfg.upload.max_products_per_run]

        task.total = len(available)
        logger.info(f"共 {len(available)} 个产品待发布")

        # 4. 初始化浏览器（延迟导入，避免无 selenium 环境报错）
        task.current_step = "启动浏览器..."
        from app.services.automation.browser_manager import BrowserManager
        browser = BrowserManager()
        if not browser.setup():
            raise Exception("浏览器启动失败")

        # 5. 登录
        task.current_step = "正在登录..."
        if not browser.login():
            raise Exception("登录失败，请检查 Cookie 或手动登录")

        # 6. 逐个发布
        published = load_published_products()
        used_titles = load_used_titles()
        success_count = 0
        upload_interval_seconds = max(0, int(getattr(cfg.upload, "upload_interval_seconds", 15) or 0))

        for i, product in enumerate(available):
            # 检查停止信号
            if task.should_stop():
                logger.info("收到停止信号，终止发布")
                break

            # 等待暂停恢复
            task.wait_if_paused()

            pid = product["pid"]
            filename = product["filename"]
            title_scene = product["title_scene"]
            category = product["category"]

            task.current_step = f"正在发布产品 [{i+1}/{len(available)}]: {pid}"
            task.progress = i + 1

            # 获取可用标题
            scene_titles = [
                pair for pair in titles_by_scene.get(title_scene, [])
                if pair["main"] not in used_titles
            ]
            if not scene_titles:
                logger.warning(f"产品 {pid} 没有可用标题，跳过")
                continue

            main_title = scene_titles[0]["main"]
            sub_title = scene_titles[0]["sub"]

            logger.info(f"=== 发布产品 {i+1}/{len(available)}: {pid} ===")
            logger.info(f"  标题: {main_title}")
            logger.info(f"  分组: {category}")

            try:
                # 委托给自动化模块执行发布
                from app.services.automation.product_publisher import publish_single_product
                success, primary_id = publish_single_product(
                    browser=browser,
                    product=product,
                    main_title=main_title,
                    sub_title=sub_title,
                    keywords=keywords,
                    cfg=cfg,
                    task=task,
                )

                if success:
                    success_count += 1
                    save_published_product(pid)
                    save_used_title(main_title)
                    published.add(pid)
                    used_titles.add(main_title)
                    logger.info(f"产品 {pid} 发布成功！")

                    if primary_id:
                        house_type = parse_house_type_from_main2_filename(pid, cfg)
                        save_new_link_product(primary_id, house_type)
                    else:
                        logger.warning(f"产品 {pid} 发布成功但未获取到 primaryId，跳过写入新发链接监控")

                    # 按配置休眠（发品间隔）
                    if i < len(available) - 1:
                        sleep_sec = upload_interval_seconds
                        task.current_step = f"按发品间隔休眠 {sleep_sec} 秒..."
                        for _ in range(sleep_sec):
                            if task.should_stop():
                                logger.info("休眠阶段收到停止信号，立即退出")
                                break
                            task.wait_if_paused()
                            time.sleep(1)
                else:
                    logger.error(f"产品 {pid} 发布失败")

            except Exception as e:
                logger.error(f"产品 {pid} 发布异常: {e}")

        task.progress = task.total
        task.current_step = f"完成！成功发布 {success_count}/{len(available)} 个产品"
        task.status = task.status  # 由 TaskManager 在函数真正返回后统一标记完成
        logger.info(f"自动发品任务完成，成功 {success_count}/{len(available)}")

    except Exception as e:
        logger.error(f"自动发品任务异常: {e}")
        task.error = str(e)
        raise
