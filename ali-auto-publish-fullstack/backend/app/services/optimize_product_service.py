# -*- coding: utf-8 -*-
"""
产品优化服务（执行逻辑对齐老脚本，配置来源对齐系统）

要求对齐：
1) 标题/属性/卖点值来自系统产出的优化结果文件（optimize_output_dir）
2) Cookie 路径来自系统 Cookie 管理配置（paths.cookie_file）
3) 提交方式复用自动发品 submit_and_verify
"""
import json
import os
import re
import time
import pickle
from types import SimpleNamespace
from typing import Dict, List, Optional, Any, Tuple

from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.keys import Keys
from selenium.common.exceptions import TimeoutException, NoSuchElementException, StaleElementReferenceException
from webdriver_manager.chrome import ChromeDriverManager

from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill
import pandas as pd

from app.core.logger import setup_logger
from app.core.settings import get_config
from app.core.task_manager import TaskInfo
from app.services.automation.submit_handler import submit_and_verify

logger = setup_logger("optimize_product_service")


def _build_compat_config():
    """将当前系统配置转换成老脚本风格字段（保证执行逻辑不变）"""
    cfg = get_config()
    attrs = cfg.attributes

    all_attrs: Dict[str, Dict] = {}
    for name, item in (attrs.all_attributes or {}).items():
        all_attrs[name] = {
            "container_id": item.container_id,
            "values": list(item.values or []),
            "input_id": item.input_id,
            "type": item.type,
            "select_type": item.select_type,
        }

    return SimpleNamespace(
        SHORT_SLEEP=attrs.short_sleep,
        NORMAL_SLEEP=attrs.normal_sleep,
        ATTR_WAIT_TIME=attrs.attr_wait_time,
        RETRY_TIMES=attrs.retry_times,
        TARGET_ATTRS=list(attrs.target_attrs or []),
        ALL_ATTRIBUTES_CONFIG=all_attrs,
        PAGE_LOAD_WAIT=getattr(cfg.upload, "page_load_wait", 20),
        DEFAULT_POSTING_URL=cfg.group_urls.default_posting_url,
        COOKIE_FILE=cfg.paths.cookie_file,
        OPTIMIZE_OUTPUT_DIR=cfg.upload.optimize_output_dir,
        OPTIMIZE_RESULT_EXCEL=getattr(cfg.upload, "optimize_result_excel", os.path.join(cfg.upload.optimize_output_dir, "产品优化结果记录.xlsx")),
    )


config = _build_compat_config()


def close_all_popups(driver):
    wait = WebDriverWait(driver, 1.5)
    popup_selectors = [
        ".next-dialog-close", ".product-sub-title-guide-balloon .next-balloon-close",
        ".float-popup .close-btn", ".modal-close", ".popup-close-icon",
        ".product-sub-title-guide-balloon i.next-icon-close", ".icbu-photobank .next-dialog-close"
    ]
    for selector in popup_selectors:
        try:
            close_btn = wait.until(EC.element_to_be_clickable((By.CSS_SELECTOR, selector)))
            driver.execute_script("arguments[0].click();", close_btn)
            time.sleep(getattr(config, 'SHORT_SLEEP', 0.1))
        except (TimeoutException, NoSuchElementException):
            pass
    try:
        driver.execute_script("""
        document.querySelectorAll('.product-sub-title-guide-balloon').forEach(el => {
            el.style.display = 'none'; if (el._tippy) el._tippy.hide();
        });
        document.querySelectorAll('.icbu-photobank').forEach(el => { el.style.display = 'none'; });
        """)
    except Exception:
        pass


def scroll_to_element(driver, element):
    driver.execute_script("arguments[0].scrollIntoView({block: 'center', behavior: 'smooth'});", element)
    time.sleep(0.5)


def fill_text_field(driver, selector, text, field_name="字段"):
    if not text:
        return False
    max_retries = 3
    for retry in range(max_retries):
        try:
            el = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, selector)))
            driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", el)
            driver.execute_script("document.activeElement.blur();", el)
            el.click()
            driver.execute_script("arguments[0].value = '';", el)
            el.send_keys(Keys.CONTROL + "a")
            el.send_keys(Keys.BACKSPACE)
            el.send_keys(text)
            filled_text = el.get_attribute("value")
            if filled_text == text:
                logger.info(f"已填写{field_name}: {text}")
                return True
            time.sleep(0.5)
        except StaleElementReferenceException:
            time.sleep(0.5)
        except Exception as e:
            logger.warning(f"填写{field_name}失败: {e}")
            if retry == max_retries - 1:
                return False
            time.sleep(0.5)
    return False


def clear_attribute_input(driver, wait, attr_name, container_id, input_selector, select_type):
    try:
        container = driver.find_element(By.ID, container_id)
        driver.execute_script("arguments[0].scrollIntoView({block: 'center', behavior: 'instant'});", container)
        time.sleep(0.3)
        driver.execute_script("""
        document.querySelectorAll('.product-sub-title-guide-balloon, .next-balloon, .compose-notice, .guide-layer').forEach(el => {
            el.style.display = 'none !important'; el.style.visibility = 'hidden';
            if (el._tippy) el._tippy.hide();
        });
        """)
        if select_type == "tag":
            for _ in range(10):
                try:
                    delete_btns = driver.find_elements(By.CSS_SELECTOR, f"#{container_id} .next-tag-close-btn")
                    if not delete_btns:
                        break
                    driver.execute_script("arguments[0].click();", delete_btns[0])
                    time.sleep(0.2)
                except Exception:
                    break
        else:
            attribute_input = wait.until(EC.element_to_be_clickable((By.CSS_SELECTOR, input_selector)))
            attribute_input.click()
            time.sleep(0.1)
            attribute_input.send_keys(Keys.CONTROL + "a")
            attribute_input.send_keys(Keys.BACKSPACE)
    except Exception:
        logger.warning(f"{attr_name}清空失败（兜底）")


def select_dropdown_option(driver, wait, attr_name, value, select_type):
    try:
        if select_type == "tag":
            return True
        if select_type == "single_search":
            dropdown_selector = ".next-overlay-wrapper.opened .sell-o-select-options, .next-overlay-wrapper.opened .next-select-menu"
            target_xpath = f"//li[contains(text(), '{value}') and (contains(@class, 'next-menu-item') or contains(@class, 'options-item'))]"
        elif select_type == "auto_complete":
            target_xpath = f"//div[contains(@class, 'next-auto-complete-option') and contains(text(), '{value}')]"
            wait.until(EC.element_to_be_clickable((By.XPATH, target_xpath)))
            target_item = driver.find_element(By.XPATH, target_xpath)
            driver.execute_script("arguments[0].click();", target_item)
            return True
        wait.until(EC.visibility_of_element_located((By.CSS_SELECTOR, dropdown_selector)))
        try:
            target_item = wait.until(EC.element_to_be_clickable((By.XPATH, target_xpath)))
        except Exception:
            target_item = wait.until(EC.element_to_be_clickable((By.CSS_SELECTOR, f"{dropdown_selector} li:first-child")))
        driver.execute_script("arguments[0].click();", target_item)
        return True
    except TimeoutException:
        logger.warning(f"{attr_name}下拉未出现，直接确认: {value}")
        return False
    except Exception as e:
        logger.warning(f"{attr_name}处理{value}失败: {e}")
        return False


def fill_single_attribute(driver, attr_name, attr_config, actual_values=None):
    SHORT_SLEEP = getattr(config, 'SHORT_SLEEP', 0.1)
    NORMAL_SLEEP = getattr(config, 'NORMAL_SLEEP', 0.3)
    WAIT_TIME = getattr(config, 'ATTR_WAIT_TIME', 5)
    RETRY_TIMES = getattr(config, 'RETRY_TIMES', 2)
    TARGET_ATTRS = getattr(config, "TARGET_ATTRS", ["型号", "品种", "形状", "品牌", "使用", "设计风格"])

    container_id = attr_config["container_id"]
    input_id = attr_config.get("input_id", "")
    select_type = attr_config["select_type"]
    wait = WebDriverWait(driver, WAIT_TIME)

    if actual_values is not None:
        valid_values = [v for v in actual_values if v and str(v).strip()]
        if not valid_values:
            logger.warning(f"{attr_name}指定值为空，跳过")
            return
    else:
        return

    input_selector = f"#{container_id} input#{input_id}" if input_id else f"#{container_id} .sell-catProp-struct .next-input input"
    attribute_input = None
    for _ in range(RETRY_TIMES):
        try:
            attribute_input = wait.until(EC.element_to_be_clickable((By.CSS_SELECTOR, input_selector)))
            break
        except TimeoutException:
            time.sleep(1)

    if not attribute_input:
        logger.warning(f"无法定位{attr_name}输入框，跳过")
        return

    clear_attribute_input(driver, wait, attr_name, container_id, input_selector, select_type)

    if attr_name in TARGET_ATTRS:
        attribute_input.click()
        time.sleep(SHORT_SLEEP)
        if select_type == "input":
            full_text = ", ".join(str(v) for v in valid_values)
            attribute_input.send_keys(Keys.CONTROL + "a")
            attribute_input.send_keys(Keys.BACKSPACE)
            for char in full_text:
                attribute_input.send_keys(char)
                time.sleep(0.05)
            driver.execute_script("arguments[0].blur();", attribute_input)
            return
        elif select_type == "tag":
            for val in valid_values:
                attribute_input.send_keys(str(val))
                time.sleep(SHORT_SLEEP)
                attribute_input.send_keys(Keys.ENTER)
            driver.execute_script("arguments[0].blur();", attribute_input)
            return
        elif select_type in ["auto_complete", "single_search"]:
            full_text = ", ".join(str(v) for v in valid_values)
            attribute_input.send_keys(Keys.CONTROL + "a")
            attribute_input.send_keys(Keys.BACKSPACE)
            for char in full_text:
                attribute_input.send_keys(char)
                time.sleep(0.05)
            driver.execute_script("arguments[0].blur();", attribute_input)
            return

    for value in valid_values:
        try:
            try:
                attribute_input.is_displayed()
            except Exception:
                attribute_input = wait.until(EC.element_to_be_clickable((By.CSS_SELECTOR, input_selector)))

            attribute_input.click()
            attribute_input.clear()
            time.sleep(SHORT_SLEEP)
            if attr_name in ["原产地", "产品类型"]:
                attribute_input.send_keys(Keys.CONTROL + "a", Keys.BACKSPACE)
                for char in str(value):
                    attribute_input.send_keys(char)
                    time.sleep(0.1)
            else:
                attribute_input.send_keys(str(value))
            time.sleep(NORMAL_SLEEP)

            if select_type == "single_search":
                if not select_dropdown_option(driver, wait, attr_name, str(value), select_type):
                    attribute_input.send_keys(Keys.ENTER)
                if attr_name == "产品类型":
                    driver.execute_script("document.body.click();")
            elif select_type == "tag":
                attribute_input.send_keys(Keys.ENTER)
                driver.execute_script("arguments[0].blur();", attribute_input)
            elif select_type == "auto_complete":
                attribute_input.send_keys(Keys.ENTER)
            elif select_type == "input":
                driver.execute_script("arguments[0].blur();", attribute_input)
                time.sleep(SHORT_SLEEP)

        except Exception as e:
            logger.warning(f"{attr_name}填写值{value}失败: {e}")


def fill_selling_points(driver, points_list):
    if not points_list:
        return False

    try:
        try:
            section = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, "struct-textDesc")))
            driver.execute_script("arguments[0].scrollIntoView({block: 'center', behavior: 'instant'});", section)
            time.sleep(1.5)
        except Exception:
            pass

        selectors = [
            "//div[@id='struct-textDesc']//div[@contenteditable='true']",
            "//div[@id='struct-textDesc']//textarea[@contenteditable='true']",
            "//div[contains(@id, 'textDesc')]//div[@contenteditable]",
            "//div[@data-struct='textDesc']//div[@contenteditable='true']",
        ]

        editor = None
        for selector in selectors:
            try:
                editor = WebDriverWait(driver, 5).until(EC.presence_of_element_located((By.XPATH, selector)))
                if editor.is_displayed() and editor.is_enabled():
                    break
            except Exception:
                continue

        if not editor or not editor.is_displayed():
            try:
                title = driver.find_element(By.XPATH, "//div[@id='struct-textDesc']//span[contains(text(), '卖点') or contains(text(), '商品卖点') or contains(text(), 'Selling')]")
                driver.execute_script("arguments[0].click();", title)
                time.sleep(1)
                editor = WebDriverWait(driver, 5).until(EC.presence_of_element_located((By.XPATH, selectors[0])))
            except Exception:
                return False

        scroll_to_element(driver, editor)
        editor.click()
        time.sleep(0.2)
        editor.send_keys(Keys.CONTROL + "a")
        editor.send_keys(Keys.BACKSPACE)
        time.sleep(0.2)

        for point in points_list:
            editor.send_keys(str(point))
            editor.send_keys(Keys.ENTER)
            time.sleep(0.1)

        driver.execute_script("arguments[0].blur();", editor)
        return True

    except Exception as e:
        logger.warning(f"填写卖点异常: {e}")
        return False


def _safe_pid(v: Any) -> str:
    s = str(v or "").strip()
    m = re.search(r"(\d{10,20})", s)
    return m.group(1) if m else s


def _contains_chinese(text: str) -> bool:
    return bool(re.search(r"[\u4e00-\u9fff]", str(text or "")))


def _sanitize_english_text(s: str) -> str:
    t = str(s or "").strip().strip("`“”\"' ")
    # 去掉类似“（主打结构+场景）：”前缀
    t = re.sub(r"^[\(（][^\)）]*[\)）]\s*[:：]?\s*", "", t)
    t = re.sub(r"^[A-Za-z\u4e00-\u9fff0-9_+\-\/ ]{1,40}[:：]\s*", "", t)
    # 只保留英文常见字符
    t = re.sub(r"[^A-Za-z0-9\-\.,/&()' ]+", " ", t)
    t = re.sub(r"\s+", " ", t).strip()
    return t


def _extract_title_v1(detail_text: str) -> str:
    text = str(detail_text or "")
    lines = [ln.strip() for ln in text.splitlines() if ln.strip()]

    in_block = False
    for i, ln in enumerate(lines):
        norm = re.sub(r"^[>#\-•*\d\.、\s]+", "", ln).strip()
        if ("优化后的标题" in norm) or ("标题版本" in norm):
            in_block = True
            continue
        if not in_block:
            continue

        m1 = re.search(r"^版本\s*1\s*[:：]?\s*(.*)$", norm, re.I)
        if m1:
            tail = _sanitize_english_text(m1.group(1))
            if tail and len(tail) > 8 and not _contains_chinese(tail):
                return tail
            for j in range(1, 5):
                if i + j < len(lines):
                    cand = _sanitize_english_text(lines[i + j])
                    if cand and len(cand) > 8 and not _contains_chinese(cand):
                        return cand
            break

    return ""


def _parse_attr_keywords(detail_text: str) -> List[str]:
    text = str(detail_text or "")
    m = re.search(r"属性埋词推荐\s*[：:]\s*【([^】]+)】", text)
    if not m:
        return []
    raw = m.group(1)
    kws = [x.strip() for x in raw.split(",") if x.strip()]
    # 只保留英文部分
    kws = [k for k in kws if not _contains_chinese(k)]
    return kws


def _parse_selling_lines(detail_text: str) -> List[str]:
    text = str(detail_text or "")
    out: List[str] = []

    # 1) 先截取“详情页卖点埋词推荐”区块（到“禁止埋入关键词”为止）
    block = ""
    m_block = re.search(r"详情页卖点埋词推荐[\s\S]*?(?:禁止埋入关键词|$)", text)
    if m_block:
        block = m_block.group(0)

    if not block:
        return out

    # 2) 优先解析 【keyword: sentence】 结构（支持跨行）
    bracket_items = re.findall(r"【([\s\S]*?)】", block)
    for item in bracket_items:
        s = re.sub(r"\s+", " ", str(item or "")).strip()
        if not s:
            continue
        if "：" in s:
            parts = s.split("：", 1)
        elif ":" in s:
            parts = s.split(":", 1)
        else:
            parts = []

        if len(parts) != 2:
            continue

        keyword = parts[0].strip()
        sentence = _sanitize_english_text(parts[1])
        if not sentence:
            continue
        if _contains_chinese(keyword) or _contains_chinese(sentence):
            continue
        out.append(sentence)

    # 3) 兜底：解析非【】结构的“keyword: sentence”行
    if not out:
        for ln in block.splitlines():
            s = str(ln or "").strip()
            if not s or "详情页卖点埋词推荐" in s or "禁止埋入关键词" in s:
                continue
            if "：" in s:
                parts = s.split("：", 1)
            elif ":" in s:
                parts = s.split(":", 1)
            else:
                continue
            if len(parts) != 2:
                continue
            keyword = parts[0].strip().strip("【】")
            sentence = _sanitize_english_text(parts[1].strip().strip("【】"))
            if not sentence:
                continue
            if _contains_chinese(keyword) or _contains_chinese(sentence):
                continue
            out.append(sentence)

    # 去重保序
    unique: List[str] = []
    seen = set()
    for x in out:
        if x not in seen:
            seen.add(x)
            unique.append(x)
    return unique


def _split_evenly(items: List[str], group_count: int) -> List[List[str]]:
    if group_count <= 0:
        return []
    groups: List[List[str]] = [[] for _ in range(group_count)]
    for idx, item in enumerate(items):
        groups[idx % group_count].append(item)
    return groups


def _detect_non_editable_reason(driver) -> str:
    """检测产品是否处于冻结/审核/不可编辑状态"""
    try:
        text = str(driver.find_element(By.TAG_NAME, "body").text or "").strip()
    except Exception:
        return ""

    rules: List[Tuple[str, str]] = [
        ("正在审核", "产品正在审核中，不可编辑"),
        ("请不要重复提交", "产品正在审核中，请勿重复提交"),
        ("冻结", "产品处于冻结状态，不可编辑"),
        ("不可编辑", "产品当前不可编辑"),
        ("无法编辑", "产品当前无法编辑"),
    ]

    for kw, msg in rules:
        if kw in text:
            return msg
    return ""


def _load_detail_txt_by_pid(detail_dir: str, pid: str) -> str:
    if not detail_dir or not os.path.isdir(detail_dir):
        return ""

    latest_day = ""
    day_dirs = [d for d in os.listdir(detail_dir) if re.fullmatch(r"\d{6}", str(d or "")) and os.path.isdir(os.path.join(detail_dir, d))]
    if day_dirs:
        day_dirs.sort(reverse=True)
        latest_day = os.path.join(detail_dir, day_dirs[0])
    else:
        latest_day = detail_dir

    candidates = []
    for n in os.listdir(latest_day):
        if not str(n).lower().endswith(".txt"):
            continue
        fp = os.path.join(latest_day, n)
        if not os.path.isfile(fp):
            continue
        candidates.append(fp)

    candidates.sort(key=lambda p: os.path.getmtime(p), reverse=True)

    target_pid = _safe_pid(pid)
    for fp in candidates:
        try:
            with open(fp, "r", encoding="utf-8") as f:
                content = f.read()
            m = re.search(r"产品ID\s*:\s*([^\n\r]+)", content)
            file_pid = _safe_pid(m.group(1).strip()) if m else ""
            if file_pid and file_pid == target_pid:
                return content
        except Exception:
            continue

    return ""


def _build_plan_from_system(pid: str, cfg) -> Dict[str, Any]:
    detail_text = _load_detail_txt_by_pid(cfg.data_analysis.title_optimize_detail_dir, pid)

    title = _extract_title_v1(detail_text)

    raw_attr_names = str(getattr(cfg.upload, "optimize_attribute_names", "") or "")
    attr_names = [x.strip() for x in re.split(r"[,，;；\n]+", raw_attr_names) if x.strip()]
    attr_keywords = _parse_attr_keywords(detail_text)
    split_groups = _split_evenly(attr_keywords, len(attr_names)) if attr_names else []

    attrs: Dict[str, str] = {}
    for i, attr_name in enumerate(attr_names):
        kws = split_groups[i] if i < len(split_groups) else []
        if kws:
            attrs[attr_name] = ", ".join(kws)

    selling_points = _parse_selling_lines(detail_text)

    logger.info(f"[优化数据读取] PID={_safe_pid(pid)}")
    logger.info(f"[优化数据读取] 标题V1={title}")
    logger.info(f"[优化数据读取] 属性关键词({len(attr_keywords)}): {attr_keywords}")
    logger.info(f"[优化数据读取] 属性目标字段: {attr_names}")
    for i, attr_name in enumerate(attr_names):
        kws = split_groups[i] if i < len(split_groups) else []
        logger.info(f"[优化数据分配] {attr_name} <= {kws}")
    logger.info(f"[优化数据读取] 卖点行数={len(selling_points)}")
    for idx, sp in enumerate(selling_points, start=1):
        logger.info(f"[优化数据读取] 卖点{idx}: {sp}")

    return {
        "product_id": _safe_pid(pid),
        "title": title,
        "attributes_to_update": attrs,
        "selling_points": selling_points,
        "detail_text_found": bool(detail_text),
    }


def _load_plan_by_pid(output_dir: str, pid: str) -> Optional[Dict[str, Any]]:
    if not output_dir or not os.path.isdir(output_dir):
        return None

    files = [
        os.path.join(output_dir, n)
        for n in os.listdir(output_dir)
        if n.lower().endswith(".json")
    ]
    files.sort(key=lambda p: os.path.getmtime(p), reverse=True)

    for fp in files:
        try:
            with open(fp, "r", encoding="utf-8") as f:
                data = json.load(f)
        except Exception:
            continue

        candidates: List[Dict[str, Any]] = []
        if isinstance(data, dict):
            if "product_id" in data or "产品ID" in data or "itemId" in data:
                candidates = [data]
            else:
                # 兼容 {"results": [...]} / {"items": [...]} 等结构
                for key in ["results", "items", "data", "list"]:
                    v = data.get(key)
                    if isinstance(v, list):
                        candidates = [x for x in v if isinstance(x, dict)]
                        break
        elif isinstance(data, list):
            candidates = [x for x in data if isinstance(x, dict)]

        for item in candidates:
            plan = _extract_plan_from_obj(item)
            if str(plan.get("product_id", "")).strip() == str(pid).strip():
                return plan

    return None


def _load_all_plans(output_dir: str) -> List[Dict[str, Any]]:
    plans: List[Dict[str, Any]] = []
    if not output_dir or not os.path.isdir(output_dir):
        return plans

    files = [
        os.path.join(output_dir, n)
        for n in os.listdir(output_dir)
        if n.lower().endswith(".json")
    ]
    files.sort(key=lambda p: os.path.getmtime(p), reverse=True)

    seen = set()
    for fp in files:
        try:
            with open(fp, "r", encoding="utf-8") as f:
                data = json.load(f)
        except Exception:
            continue

        candidates: List[Dict[str, Any]] = []
        if isinstance(data, dict):
            if "product_id" in data or "产品ID" in data or "itemId" in data:
                candidates = [data]
            else:
                for key in ["results", "items", "data", "list"]:
                    v = data.get(key)
                    if isinstance(v, list):
                        candidates = [x for x in v if isinstance(x, dict)]
                        break
        elif isinstance(data, list):
            candidates = [x for x in data if isinstance(x, dict)]

        for item in candidates:
            plan = _extract_plan_from_obj(item)
            pid = str(plan.get("product_id", "")).strip()
            if not pid or pid in seen:
                continue
            seen.add(pid)
            plans.append(plan)

    return plans


def _ensure_result_excel(path: str):
    os.makedirs(os.path.dirname(path), exist_ok=True)
    if os.path.exists(path):
        return
    wb = Workbook()
    ws = wb.active
    ws.title = "结果"
    ws.append([
        "序号",
        "产品ID",
        "优化日期",
        "标题（原标题和优化后的标题）",
        "属性（修改的属性）",
        "卖点",
        "操作结果",
        "错误信息",
    ])
    wb.save(path)


def _append_result_excel(
    path: str,
    index_num: int,
    product_id: str,
    optimize_date: str,
    title_before: str,
    title_after: str,
    attrs_text: str,
    selling_points_text: str,
    result_text: str,
    error_text: str,
):
    _ensure_result_excel(path)
    wb = load_workbook(path)
    ws = wb.active

    title_mix = f"原标题: {title_before} | 优化后: {title_after}".strip()
    ws.append([
        index_num,
        product_id,
        optimize_date,
        title_mix,
        attrs_text,
        selling_points_text,
        result_text,
        error_text,
    ])

    if result_text == "失败":
        red_fill = PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid")
        row_idx = ws.max_row
        for col in range(1, 9):
            ws.cell(row=row_idx, column=col).fill = red_fill

    wb.save(path)


class ProductOptimizer:
    def __init__(self):
        self.driver = None

    def setup_driver(self):
        options = Options()
        options.add_argument("--window-size=1280,720")
        options.add_experimental_option("excludeSwitches", ["enable-automation"])
        options.add_argument("--disable-blink-features=AutomationControlled")
        service = Service(ChromeDriverManager().install())
        self.driver = webdriver.Chrome(service=service, options=options)
        self.driver.set_page_load_timeout(30)

    def wait_and_login(self):
        default_url = getattr(config, 'DEFAULT_POSTING_URL', "https://post.alibaba.com/product/publish.htm")
        self.driver.get(default_url)
        time.sleep(2)

        cookie_file = getattr(config, 'COOKIE_FILE', None)
        if cookie_file and os.path.exists(cookie_file):
            try:
                with open(cookie_file, 'rb') as f:
                    cookies = pickle.load(f)
                for cookie in cookies:
                    if 'domain' in cookie and 'alibaba.com' in cookie['domain']:
                        cookie['domain'] = '.alibaba.com'
                    try:
                        self.driver.add_cookie(cookie)
                    except Exception:
                        pass
                self.driver.refresh()
                time.sleep(2)
            except Exception:
                pass

        try:
            WebDriverWait(self.driver, 5).until(
                EC.presence_of_element_located((By.XPATH, "//button[contains(text(), '发布产品')]"))
            )
            return True
        except TimeoutException:
            pass

        # 服务模式下不做 input() 阻塞，无法自动登录时判失败
        try:
            WebDriverWait(self.driver, 8).until(EC.presence_of_element_located((By.ID, "fm-login-id")))
            logger.error("检测到登录页，Cookie 失效，请先在系统中更新 Cookie")
            return False
        except TimeoutException:
            return True

    def open_page(self, product_id: str):
        url = f"https://post.alibaba.com/product/publish.htm?spm=a2700.micro_product_manager.0.0.5d083e5f9FxaQ6&itemId={product_id}"
        self.driver.get(url)
        time.sleep(2)
        try:
            WebDriverWait(self.driver, getattr(config, 'PAGE_LOAD_WAIT', 20)).until(
                EC.presence_of_element_located((By.XPATH, "//h2//span[text()='基本信息']"))
            )
            return True
        except TimeoutException:
            return False

    def optimize_one(self, plan: Dict[str, Any], task: Optional[TaskInfo] = None) -> Dict[str, Any]:
        pid = str(plan.get("product_id") or "").strip()
        if not pid:
            return {"ok": False, "title_before": "", "title_after": "", "attrs_text": "", "selling_points_text": "", "error": "缺少产品ID"}

        if task:
            task.current_step = f"正在优化产品: {pid}"

        if not self.open_page(pid):
            logger.error(f"产品 {pid} 打开失败")
            return {"ok": False, "title_before": "", "title_after": "", "attrs_text": "", "selling_points_text": "", "error": "打开产品页面失败"}

        close_all_popups(self.driver)

        non_editable_reason = _detect_non_editable_reason(self.driver)
        if non_editable_reason:
            logger.warning(f"产品 {pid} 跳过优化: {non_editable_reason}")
            return {"ok": False, "title_before": "", "title_after": "", "attrs_text": "", "selling_points_text": "", "error": non_editable_reason}

        title_before = ""
        try:
            title_el = WebDriverWait(self.driver, 8).until(EC.presence_of_element_located((By.XPATH, "//input[@id='productTitle']")))
            title_before = str(title_el.get_attribute("value") or "").strip()
        except Exception:
            pass

        title_val = str(plan.get("title") or "").strip()
        changed = False

        if title_val:
            ok_title = fill_text_field(self.driver, "//input[@id='productTitle']", title_val, "商品名称")
            changed = changed or ok_title
        time.sleep(getattr(config, 'NORMAL_SLEEP', 0.3))

        attr_task = plan.get("attributes_to_update") or {}
        if isinstance(attr_task, dict) and attr_task:
            logger.info(f"[属性执行] PID={pid} 计划属性={list(attr_task.keys())}")
            for attr_name, keyword in attr_task.items():
                attr_config = getattr(config, "ALL_ATTRIBUTES_CONFIG", {}).get(attr_name)
                if attr_config:
                    logger.info(f"[属性执行] 填写 {attr_name} => {keyword}")
                    fill_single_attribute(self.driver, attr_name, attr_config, actual_values=[keyword])
                    changed = True
                else:
                    logger.warning(f"[属性执行] 跳过 {attr_name}：未在系统属性配置中找到")
                time.sleep(getattr(config, 'NORMAL_SLEEP', 0.3))
        else:
            logger.warning(f"[属性执行] PID={pid} 无属性优化数据")

        points_val = plan.get("selling_points") or []
        if isinstance(points_val, list) and points_val:
            try:
                WebDriverWait(self.driver, 5).until(
                    EC.presence_of_element_located((By.ID, "struct-textDesc"))
                )
            except Exception:
                self.driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
                time.sleep(2)
            sp_ok = fill_selling_points(self.driver, points_val)
            changed = changed or bool(sp_ok)

        title_after = title_val or title_before

        attrs_text = ""
        if isinstance(attr_task, dict) and attr_task:
            attrs_text = "; ".join([f"{k}: {v}" for k, v in attr_task.items()])

        selling_points_text = ""
        if isinstance(points_val, list) and points_val:
            selling_points_text = "\n".join([str(x) for x in points_val])

        if not changed:
            logger.warning(f"产品 {pid} 没有可优化内容，跳过提交")
            return {
                "ok": False,
                "title_before": title_before,
                "title_after": title_after,
                "attrs_text": attrs_text,
                "selling_points_text": selling_points_text,
                "error": "无可优化内容，已跳过提交",
            }

        # 提交方式严格复用自动发品
        ok, primary_id = submit_and_verify(self.driver)
        if ok:
            logger.info(f"产品优化提交成功: pid={pid}, primaryId={primary_id}")
            return {
                "ok": True,
                "title_before": title_before,
                "title_after": title_after,
                "attrs_text": attrs_text,
                "selling_points_text": selling_points_text,
                "error": "",
            }

        logger.error(f"产品优化提交失败: pid={pid}")
        return {
            "ok": False,
            "title_before": title_before,
            "title_after": title_after,
            "attrs_text": attrs_text,
            "selling_points_text": selling_points_text,
            "error": "页面提交失败或校验未通过",
        }


def _list_all_pids_from_detail_dir(detail_dir: str) -> List[str]:
    if not detail_dir or not os.path.isdir(detail_dir):
        return []

    day_dirs = [d for d in os.listdir(detail_dir) if re.fullmatch(r"\d{6}", str(d or "")) and os.path.isdir(os.path.join(detail_dir, d))]
    if day_dirs:
        day_dirs.sort(reverse=True)
        base_dir = os.path.join(detail_dir, day_dirs[0])
    else:
        base_dir = detail_dir

    pids: List[str] = []
    seen = set()
    txt_files = [os.path.join(base_dir, n) for n in os.listdir(base_dir) if str(n).lower().endswith(".txt")]
    txt_files.sort(key=lambda p: os.path.getmtime(p), reverse=True)

    for fp in txt_files:
        if not os.path.isfile(fp):
            continue
        try:
            with open(fp, "r", encoding="utf-8") as f:
                content = f.read()
            m = re.search(r"产品ID\s*:\s*([^\n\r]+)", content)
            pid = _safe_pid(m.group(1).strip()) if m else ""
            if pid and pid not in seen:
                seen.add(pid)
                pids.append(pid)
        except Exception:
            continue

    return pids


def _resolve_statistics_file(cfg) -> str:
    output_file = str(getattr(cfg.data_analysis, "output_file", "") or "").strip()
    if output_file and os.path.exists(output_file):
        return output_file

    # 兜底：在同目录找最近的 xlsx
    base_dir = os.path.dirname(output_file) if output_file else ""
    if base_dir and os.path.isdir(base_dir):
        cands = [
            os.path.join(base_dir, n)
            for n in os.listdir(base_dir)
            if str(n).lower().endswith((".xlsx", ".xls")) and not str(n).startswith("~$")
        ]
        cands.sort(key=lambda p: os.path.getmtime(p), reverse=True)
        for fp in cands:
            if os.path.exists(fp):
                return fp
    return ""


def _get_latest_5_weeks(cfg) -> List[str]:
    statistics_file = _resolve_statistics_file(cfg)
    if not statistics_file:
        return []

    try:
        df0 = pd.read_excel(statistics_file, sheet_name="全店曝光次数", nrows=0)
    except Exception:
        return []

    cols = [str(c).strip() for c in df0.columns]
    weeks = [c for c in cols if re.fullmatch(r"\d{6}-\d{6}", c)]
    weeks.sort(reverse=True)
    return weeks[:5]


def _extract_latest_5_week_values_from_statistics(cfg, product_id: str) -> List[Dict[str, str]]:
    """仅从“全店曝光次数”sheet 读取最近5周历史（新到旧）"""
    statistics_file = _resolve_statistics_file(cfg)
    if not statistics_file:
        return []

    try:
        df = pd.read_excel(statistics_file, sheet_name="全店曝光次数", dtype=str)
    except Exception:
        return []

    cols = [str(c).strip() for c in df.columns]
    if not cols:
        return []

    pid_col = "产品ID" if "产品ID" in cols else cols[0]
    target_pid = _safe_pid(product_id)

    row = None
    for _, r in df.iterrows():
        pid = _safe_pid(r.get(pid_col))
        if pid == target_pid:
            row = r
            break

    if row is None:
        return []

    week_cols = [c for c in cols if re.fullmatch(r"\d{6}-\d{6}", c)]
    week_cols.sort(reverse=True)
    week_cols = week_cols[:5]

    out: List[Dict[str, str]] = []
    for wk in week_cols:
        out.append({"week": wk, "value": str(row.get(wk) or "")})
    return out


def _load_new_product_flags(cfg) -> Dict[str, bool]:
    """从“新发链接监控”读取发品<=30天产品（兼容两种格式：发品天数/发品日期）"""
    fp = str(getattr(cfg.data_analysis, "new_links_file_path", "") or "").strip()
    if not fp or not os.path.exists(fp):
        return {}

    try:
        xls = pd.ExcelFile(fp)
        if not xls.sheet_names:
            return {}
        df = pd.read_excel(fp, sheet_name=xls.sheet_names[0], dtype=str)
    except Exception:
        return {}

    cols = [str(c).strip() for c in df.columns]
    if not cols:
        return {}

    # 1) 产品ID列识别
    pid_col = "产品ID" if "产品ID" in cols else ""
    if not pid_col:
        # 优先找“值像产品ID”的列
        best_col = ""
        best_score = -1
        for c in cols:
            score = 0
            for v in df[c].head(200).tolist():
                if re.search(r"\d{10,20}", str(v or "")):
                    score += 1
            if score > best_score:
                best_score = score
                best_col = c
        pid_col = best_col or cols[0]

    # 2) 发品天数列识别
    days_col = "发品天数" if "发品天数" in cols else ""
    if not days_col:
        for c in cols:
            if "发品" in c and "天" in c:
                days_col = c
                break

    # 3) 发品日期列识别（如 260411）
    date_col = ""
    for c in cols:
        if "发品" in c and ("日期" in c or "时间" in c):
            date_col = c
            break
    if not date_col:
        # 兜底：找到最像 yymmdd 的列
        best_col = ""
        best_score = -1
        for c in cols:
            score = 0
            for v in df[c].head(200).tolist():
                if re.fullmatch(r"\d{6}", str(v or "").strip()):
                    score += 1
            if score > best_score:
                best_score = score
                best_col = c
        if best_score > 0:
            date_col = best_col

    today = pd.Timestamp.now().normalize()
    flags: Dict[str, bool] = {}

    for _, r in df.iterrows():
        pid = _safe_pid(r.get(pid_col))
        if not pid:
            continue

        days: Optional[int] = None

        # 优先直接用发品天数
        if days_col:
            raw_days = str(r.get(days_col) or "").strip()
            m = re.search(r"-?\d+", raw_days)
            if m:
                try:
                    days = int(m.group(0))
                except Exception:
                    days = None

        # 无发品天数时，用发品日期反推
        if days is None and date_col:
            raw_date = str(r.get(date_col) or "").strip().replace("'", "")
            digits = re.sub(r"\D", "", raw_date)
            dt = None
            if len(digits) == 6:
                try:
                    y = 2000 + int(digits[:2])
                    mth = int(digits[2:4])
                    d = int(digits[4:6])
                    dt = pd.Timestamp(year=y, month=mth, day=d)
                except Exception:
                    dt = None
            elif len(digits) == 8:
                try:
                    y = int(digits[:4])
                    mth = int(digits[4:6])
                    d = int(digits[6:8])
                    dt = pd.Timestamp(year=y, month=mth, day=d)
                except Exception:
                    dt = None
            if dt is not None:
                days = int((today - pd.Timestamp(dt).normalize()).days)

        if days is None:
            continue

        flags[pid] = days <= 30

    return flags


def get_optimize_list(limit: int = 300) -> Dict[str, Any]:
    """返回优化产品列表（用于前端“优化产品列表”模块）"""
    cfg = get_config()
    result_excel = os.path.join(
        cfg.upload.optimize_output_dir,
        "产品优化结果记录.xlsx",
    )

    rows: List[Dict[str, Any]] = []
    history_weeks: List[str] = _get_latest_5_weeks(cfg)
    new_flags = _load_new_product_flags(cfg)

    if os.path.exists(result_excel):
        try:
            wb = load_workbook(result_excel)
            ws = wb.active
            for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=1):
                seq, pid, optimize_date, title_mix, attrs_text, _selling_points, _result_text, _error = (list(row) + [""] * 8)[:8]
                pid_s = _safe_pid(pid)
                if not pid_s:
                    continue

                title_before = ""
                title_after = ""
                mix = str(title_mix or "")
                m = re.search(r"原标题\s*:\s*(.*?)\s*\|\s*优化后\s*:\s*(.*)", mix)
                if m:
                    title_before = str(m.group(1) or "").strip()
                    title_after = str(m.group(2) or "").strip()

                attrs = [x.strip() for x in str(attrs_text or "").split(";") if x.strip()]

                history = _extract_latest_5_week_values_from_statistics(cfg, pid_s)
                if not history and history_weeks:
                    history = [{"week": wk, "value": ""} for wk in history_weeks]

                rows.append({
                    "product_id": pid_s,
                    "optimize_date": str(optimize_date or ""),
                    "title_before": title_before,
                    "title_after": title_after,
                    "attrs": attrs,
                    "history": history,
                    "is_new_product": bool(new_flags.get(pid_s, False)),
                    "result": str(_result_text or ""),
                })
        except Exception:
            pass

    # 最新优化结果放在前面（Excel 为追加写入，末尾为最新）
    rows = list(reversed(rows))
    rows = rows[: max(0, int(limit or 0))] if limit is not None else rows
    return {"rows": rows, "history_weeks": history_weeks}


def run_optimize_product_task(task: TaskInfo, manual_product_ids: Optional[List[str]] = None):
    """任务入口：供 /upload/optimize/start 调用"""
    global config
    config = _build_compat_config()  # 每次启动刷新系统配置

    cfg = get_config()
    plans: List[Dict[str, Any]] = []

    ids = [str(x).strip() for x in (manual_product_ids or []) if str(x).strip()]
    if not ids:
        ids = _list_all_pids_from_detail_dir(cfg.data_analysis.title_optimize_detail_dir)

    for pid in ids:
        plan = _build_plan_from_system(pid, cfg)
        plans.append(plan)

    if not plans:
        task.current_step = "未找到可执行的优化数据"
        task.total = 0
        task.progress = 0
        return

    optimizer = ProductOptimizer()
    success_count = 0
    result_excel = getattr(config, "OPTIMIZE_RESULT_EXCEL", os.path.join(cfg.upload.optimize_output_dir, "产品优化结果记录.xlsx"))

    try:
        task.total = len(plans)
        task.progress = 0
        task.current_step = "启动浏览器"
        optimizer.setup_driver()

        task.current_step = "登录中"
        if not optimizer.wait_and_login():
            raise Exception("登录失败，请在 Cookie 管理中更新有效 Cookie")

        for i, plan in enumerate(plans):
            if task.should_stop():
                task.current_step = "收到停止信号，任务结束"
                break

            task.wait_if_paused()
            result = optimizer.optimize_one(plan, task=task)
            ok = bool(result.get("ok"))
            pid = str(plan.get("product_id") or "").strip()
            title_after = str(result.get("title_after") or "")
            title_before = str(result.get("title_before") or "")
            attrs_text = str(result.get("attrs_text") or "")
            selling_points_text = str(result.get("selling_points_text") or "")
            err = str(result.get("error") or "")

            _append_result_excel(
                result_excel,
                i + 1,
                pid,
                time.strftime("%Y-%m-%d"),
                title_before,
                title_after,
                attrs_text,
                selling_points_text,
                "成功" if ok else "失败",
                err,
            )

            if ok:
                success_count += 1

            task.progress = i + 1

        task.current_step = f"优化任务完成：成功 {success_count}/{task.total}"

    except Exception as e:
        logger.error(f"优化任务异常: {e}")
        task.error = str(e)
        raise
    finally:
        if optimizer.driver:
            try:
                optimizer.driver.quit()
            except Exception:
                pass


def _load_recent_history_map(cfg) -> Tuple[List[str], Dict[str, Dict[str, str]]]:
    """读取综合数据分析-产品数据历史（最近5周，最新在前）"""
    week_cols: List[str] = []
    pid_map: Dict[str, Dict[str, str]] = {}

    xlsx = getattr(cfg.data_analysis, "output_file", "")
    if not xlsx or not os.path.exists(xlsx):
        return week_cols, pid_map

    try:
        wb = load_workbook(xlsx, data_only=True)
    except Exception:
        return week_cols, pid_map

    ws = wb["全店曝光次数"] if "全店曝光次数" in wb.sheetnames else wb[wb.sheetnames[0]]
    headers = [str(c.value or "").strip() for c in ws[1]]
    if not headers:
        return week_cols, pid_map

    pid_idx = -1
    for i, h in enumerate(headers):
        if h == "产品ID":
            pid_idx = i
            break
    if pid_idx < 0:
        return week_cols, pid_map

    all_week = [h for h in headers if re.fullmatch(r"\d{6}-\d{6}", h)]
    all_week.sort(reverse=True)
    week_cols = all_week[:5]

    col_index = {h: i for i, h in enumerate(headers)}
    for row in ws.iter_rows(min_row=2, values_only=True):
        raw_pid = row[pid_idx] if pid_idx < len(row) else ""
        pid = _safe_pid(raw_pid)
        if not pid:
            continue
        rec: Dict[str, str] = {}
        for wk in week_cols:
            idx = col_index.get(wk, -1)
            if idx < 0 or idx >= len(row):
                rec[wk] = ""
            else:
                v = row[idx]
                rec[wk] = "" if v is None else str(v)
        pid_map[pid] = rec

    return week_cols, pid_map


def get_optimize_product_list(limit: int = 200) -> Dict[str, Any]:
    """获取优化产品列表（用于前端列表展示）"""
    cfg = get_config()
    result_excel = getattr(cfg.upload, "optimize_result_excel", os.path.join(cfg.upload.optimize_output_dir, "产品优化结果记录.xlsx"))

    rows: List[Dict[str, Any]] = []
    if os.path.exists(result_excel):
        try:
            wb = load_workbook(result_excel, data_only=True)
            ws = wb.active
            headers = [str(c.value or "").strip() for c in ws[1]]
            idx = {h: i for i, h in enumerate(headers)}

            for r in ws.iter_rows(min_row=2, values_only=True):
                pid = _safe_pid(r[idx.get("产品ID", -1)] if idx.get("产品ID", -1) >= 0 else "")
                if not pid:
                    continue

                optimize_date = str(r[idx.get("优化日期", -1)] or "") if idx.get("优化日期", -1) >= 0 else ""
                title_mix = str(r[idx.get("标题（原标题和优化后的标题）", -1)] or "") if idx.get("标题（原标题和优化后的标题）", -1) >= 0 else ""
                attrs_text = str(r[idx.get("属性（修改的属性）", -1)] or "") if idx.get("属性（修改的属性）", -1) >= 0 else ""
                result_text = str(r[idx.get("操作结果", -1)] or "") if idx.get("操作结果", -1) >= 0 else ""
                if result_text != "成功":
                    continue

                title_before = ""
                title_after = ""
                m = re.search(r"原标题\s*:\s*(.*?)\s*\|\s*优化后\s*:\s*(.*)", title_mix)
                if m:
                    title_before = str(m.group(1) or "").strip()
                    title_after = str(m.group(2) or "").strip()
                else:
                    title_after = title_mix

                attrs_list = [x.strip() for x in re.split(r"[;；]+", attrs_text) if x.strip()]

                rows.append({
                    "product_id": pid,
                    "optimize_date": optimize_date,
                    "title_before": title_before,
                    "title_after": title_after,
                    "attrs": attrs_list,
                    "result": result_text,
                })
        except Exception:
            rows = []

    # 新数据在前
    rows = list(reversed(rows))
    if limit > 0:
        rows = rows[:limit]

    week_cols, history_map = _load_recent_history_map(cfg)
    for item in rows:
        pid = str(item.get("product_id") or "")
        hist = history_map.get(pid, {})
        item["history"] = [{"week": w, "value": str(hist.get(w, ""))} for w in week_cols]

    return {
        "columns": ["产品ID", "优化时间", "优化标题", "优化属性", "历史数据"],
        "history_weeks": week_cols,
        "rows": rows,
    }


def get_today_failed_optimize_ids() -> Dict[str, Any]:
    """从配置输出目录的结果文件中读取当天失败的产品ID"""
    cfg = get_config()
    result_excel = getattr(cfg.upload, "optimize_result_excel", os.path.join(cfg.upload.optimize_output_dir, "产品优化结果记录.xlsx"))
    today = time.strftime("%Y-%m-%d")
    if not os.path.exists(result_excel):
        return {"date": today, "ids": []}

    def _norm_date(v: Any) -> str:
        s = str(v or "").strip()
        if not s:
            return ""
        return s[:10].replace("/", "-").replace(".", "-")

    ids: List[str] = []
    try:
        wb = load_workbook(result_excel, data_only=True)
        ws = wb.active
        headers = [str(c.value or "").strip() for c in ws[1]]
        idx = {h: i for i, h in enumerate(headers)}
        pid_idx = idx.get("产品ID", -1)
        date_idx = idx.get("优化日期", -1)
        result_idx = idx.get("操作结果", -1)
        if pid_idx < 0 or date_idx < 0 or result_idx < 0:
            return {"date": today, "ids": []}

        for r in ws.iter_rows(min_row=2, values_only=True):
            row_date = _norm_date(r[date_idx] if date_idx < len(r) else "")
            row_result = str(r[result_idx] or "").strip()
            if row_date != today or "失败" not in row_result:
                continue
            pid = _safe_pid(r[pid_idx] if pid_idx < len(r) else "")
            if pid and pid not in ids:
                ids.append(pid)
    except Exception:
        pass

    return {"date": today, "ids": ids}


def delete_optimize_product_record(product_id: str, optimize_date: Optional[str] = None) -> Dict[str, Any]:
    """删除优化结果记录"""
    cfg = get_config()
    result_excel = getattr(cfg.upload, "optimize_result_excel", os.path.join(cfg.upload.optimize_output_dir, "产品优化结果记录.xlsx"))
    if not os.path.exists(result_excel):
        raise FileNotFoundError("优化结果文件不存在")

    target_pid = _safe_pid(product_id)
    if not target_pid:
        raise ValueError("产品ID不能为空")

    wb = load_workbook(result_excel)
    ws = wb.active
    headers = [str(c.value or "").strip() for c in ws[1]]
    idx = {h: i for i, h in enumerate(headers)}
    pid_idx = idx.get("产品ID", -1)
    date_idx = idx.get("优化日期", -1)
    if pid_idx < 0:
        raise ValueError("结果文件缺少产品ID列")

    removed = 0
    kept_rows = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        row_pid = _safe_pid(r[pid_idx] if pid_idx < len(r) else "")
        row_date = str(r[date_idx] or "") if date_idx >= 0 and date_idx < len(r) else ""
        should_delete = row_pid == target_pid and (not optimize_date or row_date == optimize_date)
        if should_delete:
            removed += 1
            continue
        kept_rows.append(list(r))

    if removed == 0:
        return {"deleted": 0, "product_id": target_pid}

    ws.delete_rows(2, ws.max_row)
    for row in kept_rows:
        ws.append(row)
    wb.save(result_excel)
    return {"deleted": removed, "product_id": target_pid}
