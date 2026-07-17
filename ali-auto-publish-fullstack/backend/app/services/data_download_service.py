# -*- coding: utf-8 -*-
"""
数据下载服务层
重构自: 后台数据下载/*.py

【如何修改】
- 修改产品参谋下载逻辑 → 修改 _download_product_ranking()
- 修改日数据下载逻辑 → 修改 _download_daily_data()
- 修改店铺数据下载逻辑 → 修改 _download_store_overview()
- 修改关键词下载逻辑 → 修改 _download_keywords()
- 添加新的下载类型 → 在 run_download_task() 中添加新分支
"""
import os
import re
import json
import time
import gc
import pickle
import base64
import shutil
import threading
import asyncio
import tempfile
import requests
import pandas as pd
from urllib.parse import urlparse, parse_qs, quote
from datetime import datetime, timedelta
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill
from typing import Dict, List, Optional, Any

from app.core.settings import get_config
from app.core.task_manager import TaskInfo
from app.core.logger import setup_logger
from app.services.automation.browser_manager import BrowserManager

logger = setup_logger("data_download")
ARK_API_BASE_URL = "https://ark.cn-beijing.volces.com/api/v3"
ARK_API_ENDPOINT = "/responses"

# 轻量缓存：避免店铺运营页重复读大Excel
_STORE_OVERVIEW_CACHE: Dict[str, Dict] = {}

# 关键词采集最大页数（防止异常死循环）
MAX_KEYWORD_PAGES = 50

KEYWORD_GET_API_BASE = "https://mydata.alibaba.com/self/.json"
KEYWORD_GET_REFERER = "https://data.alibaba.com/traffic/keyword"
KEYWORD_API_PAGE_SIZE = 50
KEYWORD_API_MAX_SELECTED = 30
KEYWORD_API_FIELD_MAP = {
    "keyword": "词",
    "sumShowCnt": "搜索曝光次数",
    "sumClickCnt": "点击量",
    "ctr": "点击率",
    "sumP4pShowCnt": "外贸直通车曝光",
    "sumP4pClickCnt": "外贸直通车点击次数",
    "avgSumShowCnt": "Top10 平均曝光",
    "avgSumClickCnt": "Top10 平均点击",
    "pv": "关键词指数",
    "resultCnt": "卖家规模指数",
    "statDate": "统计日期",
    "isP4pKeyword": "是否为 P4P 推广关键词",
    "p4pCampaignStatus": "P4P 推广状态",
    "keywordDetails": "关键词变体",
}
KEYWORD_SUMMARY_METRICS = {
    "搜索曝光次数": "搜索曝光次数",
    "点击量": "点击量",
    "关键词指数": "关键词指数",
    "卖家规模指数": "卖家规模指数",
    "Top10 平均曝光": "Top10 平均曝光",
    "Top10 平均点击": "Top10 平均点击",
}
KEYWORD_ANOMALY_METRICS = {
    "搜索曝光次数异动": "搜索曝光次数",
    "点击量异动": "点击量",
    "关键词指数异动": "关键词指数",
}

SHOP_TRENDS_API_BASE = "https://mydata.alibaba.com/self/.json"
SHOP_TRENDS_REFERER = "https://data.alibaba.com/"
SHOP_TRENDS_DEFAULT_CATE_ID = "201650701"
# 单次请求即返回一批完整周期（勿按 selected 逐条当成单日/单周）：
# 日 selected=0 → 约30天；周 selected=1 → 约10周；月 selected=1 → 约6个月
SHOP_TRENDS_START_SELECTED = {"day": 0, "week": 1, "month": 1}

SHOP_TRENDS_METRIC_BINDINGS = [
    ("shopUv", "店铺访问人数"),
    ("shopPv", "店铺访问次数"),
    ("searchImpls", "搜索曝光次数"),
    ("searchClicks", "搜索点击次数"),
    ("fbUv", "询盘人数"),
    ("fbPv", "询盘个数"),
    ("tmUv", "TM咨询人数"),
    ("ordCnt", "信保交易订单个数"),
    ("ordAmt", "信保交易金额(美元)"),
    ("totalClkCnt", "全店点击次数"),
    ("busByrCnt", "商机人数"),
    ("semiMgtImpsCnt", "半托管曝光次数"),
    ("highQualityFbUvRate", "L1+买家询盘人数占比"),
    ("abCnt30d", "近30天商机人数"),
    ("replyRate", "及时回复率"),
    ("avgReplyTime", "平均回复时长"),
    ("fst5minReplyRate30d", "极速回复率"),
    ("totalImpsCnt", "全店曝光次数"),
    ("totalBusCnt", "全店品的商机量"),
    ("p4pExposureCnt", "营销曝光次数"),
    ("p4pClickCnt", "营销点击次数"),
    ("natureExposureCnt", "自然曝光量"),
    ("natureClickCnt", "自然点击量"),
]
SHOP_TRENDS_P4P_ALIAS = {
    "p4pExposureCnt": "全站推曝光次数",
    "p4pClickCnt": "全站推点击次数",
}
SHOP_TRENDS_ZERO_METRICS = [
    "加购笔数",
    "加购人数",
    "标准推广曝光次数",
    "标准推广点击次数",
    "全站推商机量",
    "自然商机量",
]
SHOP_TRENDS_METRIC_ORDER = [
    "店铺访问人数",
    "店铺访问次数",
    "搜索曝光次数",
    "搜索点击次数",
    "询盘人数",
    "询盘个数",
    "TM咨询人数",
    "信保交易订单个数",
    "信保交易金额(美元)",
    "全店点击次数",
    "商机人数",
    "半托管曝光次数",
    "L1+买家询盘人数占比",
    "近30天商机人数",
    "加购笔数",
    "加购人数",
    "商机转化率",
    "及时回复率",
    "平均回复时长",
    "极速回复率",
    "全店曝光次数",
    "全店品的商机量",
    "营销曝光次数",
    "营销点击次数",
    "标准推广曝光次数",
    "标准推广点击次数",
    "全站推曝光次数",
    "全站推点击次数",
    "全站推商机量",
    "自然曝光量",
    "自然点击量",
    "自然商机量",
]


def _find_system_browser_candidates() -> List[str]:
    """返回当前机器上可能可用的系统浏览器可执行文件路径。"""
    candidates = []
    env_candidates = [
        os.getenv("CHROME_PATH", "").strip(),
        os.getenv("EDGE_PATH", "").strip(),
        os.getenv("BROWSER_PATH", "").strip(),
    ]
    candidates.extend([p for p in env_candidates if p])

    program_files = [os.getenv("ProgramFiles", ""), os.getenv("ProgramFiles(x86)", "")]
    browser_rel_paths = [
        r"Google\Chrome\Application\chrome.exe",
        r"Microsoft\Edge\Application\msedge.exe",
    ]
    for root in program_files:
        if not root:
            continue
        for rel in browser_rel_paths:
            candidates.append(os.path.join(root, rel))

    # Windows 常见用户级安装路径
    local_appdata = os.getenv("LOCALAPPDATA", "")
    if local_appdata:
        candidates.extend([
            os.path.join(local_appdata, r"Google\Chrome\Application\chrome.exe"),
            os.path.join(local_appdata, r"Microsoft\Edge\Application\msedge.exe"),
        ])

    return [p for p in candidates if p and os.path.exists(p)]


def _launch_playwright_browser(p, *, headless: bool, slow_mo: int = 0, args: Optional[List[str]] = None):
    """优先使用系统 Chrome/Edge，找不到再回退到 Playwright 默认 Chromium。"""
    launch_kwargs = {
        "headless": headless,
    }
    if slow_mo:
        launch_kwargs["slow_mo"] = slow_mo
    if args:
        launch_kwargs["args"] = args

    for browser_path in _find_system_browser_candidates():
        try:
            logger.info(f"尝试使用系统浏览器启动 Playwright: {browser_path}")
            return p.chromium.launch(executable_path=browser_path, **launch_kwargs)
        except Exception as e:
            logger.warning(f"系统浏览器启动失败，回退下一个候选: {browser_path} | {e}")

    logger.info("未找到系统浏览器，尝试使用 Playwright 默认 Chromium")
    return p.chromium.launch(**launch_kwargs)


# ===================== 统一入口 =====================

def run_download_task(
    task: TaskInfo,
    task_type: str,
    date_range: Optional[str] = None,
    period_type: str = "week",
    download_options: Optional[Dict[str, Any]] = None,
):
    """
    数据下载主任务

    Args:
        task: TaskInfo 对象
        task_type: 下载类型
        date_range: 日期范围
        period_type: 周期类型 (day/week/month)
        download_options: 本次任务可选覆盖参数（如 big_keywords / dropdown_keywords）
    """
    download_options = download_options or {}
    cfg = get_config()
    task.current_step = f"初始化 {task_type} 下载..."

    try:
        if task_type == "product360":
            _download_product360_full(task, cfg)
        elif task_type == "product360_crawler":
            _download_product360_crawler(task, cfg)
        elif task_type == "product360_parser":
            _download_product360_parser(task, cfg)
        elif task_type == "product_ranking":
            _download_product_ranking(task, cfg, date_range)
        elif task_type == "daily_data":
            _download_daily_data(task, cfg, date_range)
            if not task.should_stop():
                from app.services.analysis_service import run_analysis_chain_after_daily_download

                run_analysis_chain_after_daily_download(task)
        elif task_type == "store_overview":
            _download_store_overview(task, cfg, period_type)
        elif task_type == "traffic_channel":
            _download_traffic_channel(task, cfg)
        elif task_type == "product_operate":
            _download_product_operate(task, cfg)
        elif task_type == "store_image_collect":
            _download_store_images(task, cfg)
        elif task_type == "keyword_crawler":
            _download_keywords(task, cfg)
        elif task_type == "keyword_parser":
            _parse_keywords(task, cfg)
        elif task_type == "industry_keyword":
            _download_and_merge_industry_keywords(
                task,
                cfg,
                big_keywords_override=download_options.get("big_keywords"),
            )
        elif task_type == "industry_keyword_dropdown":
            _download_industry_keyword_dropdown(
                task,
                cfg,
                dropdown_keywords_override=download_options.get("dropdown_keywords"),
            )
        elif task_type == "keyword":
            _download_keywords(task, cfg)
        else:
            raise ValueError(f"未知的下载类型: {task_type}")

        if task_type not in ("industry_keyword", "industry_keyword_dropdown"):
            step = (task.current_step or "").strip()
            if (not step) or step.startswith("初始化") or step.endswith("..."):
                task.current_step = "下载完成"
    except Exception as e:
        logger.error(f"下载任务异常: {e}")
        task.error = str(e)
        raise


# ===================== 产品360：采集 + 解析（一键） =====================

def _download_product360_full(task: TaskInfo, cfg):
    """产品360：先采集 JSON，完成后立即解析生成 Excel 总报告。"""
    logger.info("产品360：开始采集")
    task.current_step = "正在采集产品数据（JSON）..."
    _download_product360_crawler(task, cfg)
    if task.should_stop():
        task.current_step = "已停止（采集阶段）"
        return
    logger.info("产品360：采集完成，开始解析")
    task.current_step = "正在解析 JSON 并生成 Excel..."
    _download_product360_parser(task, cfg)


# ===================== 产品360：采集 =====================

def _download_product360_crawler(task: TaskInfo, cfg):
    """
    产品360 - 模式1：抓取数据并保存 JSON
    来源脚本: cs_产品数据关键词_渠道_曝光国家_正式发布.py
    """
    task.current_step = "启动浏览器..."
    logger.info("产品360采集 - 启动浏览器")

    download_dir = _normalize_download_root(cfg.data_download.product360_output_dir)
    logger.info(f"产品360采集当前工作目录: {os.getcwd()}")
    logger.info(f"产品360采集下载根目录: {download_dir}")
    json_output_dir = _normalize_product360_dir(getattr(cfg.data_download, "product360_json_dir", "") or os.path.join(download_dir, "Json文件"))
    keyword_json_dir = _normalize_product360_dir(getattr(cfg.data_download, "product360_keyword_json_dir", "") or os.path.join(download_dir, "关键词json"))
    logger.info(f"产品360采集 JSON 输出目录: {json_output_dir}")
    logger.info(f"产品360采集关键词目录: {keyword_json_dir}")
    os.makedirs(download_dir, exist_ok=True)
    os.makedirs(json_output_dir, exist_ok=True)
    os.makedirs(keyword_json_dir, exist_ok=True)
    logger.info(f"产品360采集 JSON 输出目录: {json_output_dir}")
    logger.info(f"产品360采集关键词目录: {keyword_json_dir}")

    login_url = cfg.data_download.login_url
    target_url = "https://data.alibaba.com/product/overview"

    cfg.data_download.headless = True
    browser = BrowserManager()
    if not browser.setup():
        raise Exception("浏览器启动失败")

    driver = browser.driver

    from selenium.webdriver.common.by import By
    from selenium.webdriver.support.ui import WebDriverWait
    from selenium.webdriver.support import expected_conditions as EC

    wait = WebDriverWait(driver, 12)
    logger.info("产品360采集浏览器已启动，开始初始化页面")

    # 模式1提速参数（尽量快，同时保留基本稳定性）
    SLEEP_TINY = 0.04
    SLEEP_SHORT = 0.18
    SLEEP_MEDIUM = 0.45

    def _safe_click(elem):
        try:
            elem.click()
        except Exception:
            driver.execute_script("arguments[0].click();", elem)
        time.sleep(SLEEP_TINY)

    def _get_cookie_dict():
        return {c['name']: c['value'] for c in driver.get_cookies()}

    session = requests.Session()

    def _api_get(url):
        cookies = _get_cookie_dict()
        ctoken = cookies.get('_m_h5_tk', '').split('_')[0] or cookies.get('ctoken')
        if ctoken:
            separator = '&' if '?' in url else '?'
            url = f"{url}{separator}ctoken={ctoken}"
        headers = {"User-Agent": "Mozilla/5.0", "Referer": "https://data.alibaba.com/"}
        try:
            r = session.get(url, headers=headers, cookies=cookies, timeout=(5, 15))
            r.raise_for_status()
            return r.json() if r.text else {}
        except (requests.exceptions.RequestException, json.JSONDecodeError):
            return {}

    def _save_jsons(product_id, detail, region, channel):
        product_dir = os.path.join(json_output_dir, str(product_id))
        os.makedirs(product_dir, exist_ok=True)
        files_to_save = {
            "1_产品详情.json": detail,
            "2_访客地域.json": region,
            "3_流量来源.json": channel,
        }
        for filename, data in files_to_save.items():
            file_path = os.path.join(product_dir, filename)
            with open(file_path, 'w', encoding='utf-8') as f:
                json.dump(data, f, ensure_ascii=False, indent=4)
        logger.info(f"保存 JSON: {product_id}")

    def _save_keyword_json(product_id, filename, data):
        product_dir = os.path.join(keyword_json_dir, str(product_id))
        os.makedirs(product_dir, exist_ok=True)
        file_path = os.path.join(product_dir, filename)
        with open(file_path, 'w', encoding='utf-8') as f:
            json.dump(data, f, ensure_ascii=False, indent=4)

    def _collect_and_save_product_data(product_id: str):
        logger.info(f"开始采集产品360数据: {product_id}")
        # 1) 关键词分析（API分页）
        for page in [1, 2]:
            kw_exp_url = (
                "https://hz-mydata.alibaba.com/self/.json?action=OneAction&"
                "iName=vip/product/360/wordAnalysis/content&isVip=true&terminalType=TOTAL&"
                "statisticType=os&selected=1&statisticsType=week&prodId={product_id}&"
                "orderField=searchImps&orderDirection=desc&pageCount={page}"
            ).format(product_id=product_id, page=page)
            kw_exp_json = _api_get(kw_exp_url)

            if not isinstance(kw_exp_json, dict):
                break
            inner_data = kw_exp_json.get("data", {})
            word_list = []
            if isinstance(inner_data, dict):
                word_list = inner_data.get("list", [])
            elif isinstance(inner_data, list):
                word_list = inner_data
            if not word_list:
                break

            _save_keyword_json(product_id, f"关键词_曝光降序_第{page}页.json", kw_exp_json)

        page = 1
        while page <= MAX_KEYWORD_PAGES:
            kw_clk_url = (
                "https://hz-mydata.alibaba.com/self/.json?action=OneAction&"
                "iName=vip/product/360/wordAnalysis/content&isVip=true&terminalType=TOTAL&"
                "statisticType=os&selected=1&statisticsType=week&prodId={product_id}&"
                "orderField=searchClicks&orderDirection=desc&pageCount={page}"
            ).format(product_id=product_id, page=page)
            kw_clk_json = _api_get(kw_clk_url)

            if not isinstance(kw_clk_json, dict):
                break
            inner_data = kw_clk_json.get("data", {})
            word_list = []
            if isinstance(inner_data, dict):
                word_list = inner_data.get("list", [])
            elif isinstance(inner_data, list):
                word_list = inner_data
            if not word_list:
                break

            _save_keyword_json(product_id, f"关键词_点击降序_第{page}页.json", kw_clk_json)

            last_item = word_list[-1] if word_list else {}
            last_clicks = 0
            if isinstance(last_item, dict):
                last_clicks = int(last_item.get("searchClicks", 0) or 0)
            if last_clicks == 0:
                break

            page += 1

        # 2) 产品详情
        detail_url = (
            "https://mydata.alibaba.com/self/.json?action=OneAction&"
            "iName=buyerPortrait/diagnoseData&prodId={product_id}"
        ).format(product_id=product_id)
        detail_json = _api_get(detail_url)

        # 3) 访客地域（与页面观测参数对齐）
        region_url = (
            "https://mydata.alibaba.com/self/.json?action=OneAction&"
            "iName=vip/product/360/regionAnalysis&region=os&prodId={product_id}&"
            "terminalType=TOTAL&statisticType=os&selected=1&statisticsType=week"
        ).format(product_id=product_id)
        region_json = _api_get(region_url)

        # 4) 流量来源
        channel_url = (
            "https://hz-mydata.alibaba.com/self/.json?action=OneAction&"
            "iName=vip/product/360/channelAnalysis&isVip=true&region=os&"
            "isMyselfUpgraded=true&terminalType=TOTAL&statisticType=os&selected=1&"
            "statisticsType=week&prodId={product_id}"
        ).format(product_id=product_id)
        channel_json = _api_get(channel_url)

        _save_jsons(product_id, detail_json, region_json, channel_json)

    failed_product_ids: set = set()

    def _click_all_analysis_buttons():
        index = 0
        while True:
            if task.should_stop():
                return
            task.wait_if_paused()
            try:
                buttons = driver.find_elements(By.CSS_SELECTOR, "button.analysis-button")
                if index >= len(buttons):
                    break

                btn = buttons[index]
                driver.execute_script("arguments[0].scrollIntoView({block:'center'});", btn)
                time.sleep(SLEEP_SHORT)
                logger.info(f"处理产品 {index + 1}/{len(buttons)}")

                driver.execute_script("arguments[0].click();", btn)
                drawer = wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, ".next-drawer")))
                time.sleep(SLEEP_MEDIUM)

                product_id = None
                try:
                    id_element = drawer.find_element(By.XPATH,
                        ".//div[contains(@class, 'detail-left')]//li[contains(text(), 'ID:')]")
                    match = re.search(r'\d+', id_element.text)
                    product_id = match.group(0)
                except Exception:
                    logger.warning("获取 Product ID 失败，跳过")
                    wait.until(EC.element_to_be_clickable((By.CSS_SELECTOR, "a.next-drawer-close"))).click()
                    wait.until(EC.invisibility_of_element_located((By.CSS_SELECTOR, ".next-drawer")))
                    index += 1
                    continue

                # 轻量切到关键词分析tab（不强依赖，失败不阻断）
                try:
                    kw_tab = driver.find_element(By.XPATH, "//div[text()='关键词分析']")
                    _safe_click(kw_tab)
                    time.sleep(SLEEP_SHORT)
                except Exception:
                    pass

                # 尝试对齐页面tab状态（失败不阻断，后续重试会补）
                try:
                    wait.until(EC.element_to_be_clickable((By.XPATH, "//span[text()='访客地域']"))).click()
                    time.sleep(SLEEP_SHORT)
                except Exception:
                    pass
                try:
                    wait.until(EC.element_to_be_clickable((By.XPATH, "//span[text()='流量来源']"))).click()
                    time.sleep(SLEEP_SHORT)
                except Exception:
                    pass

                _collect_and_save_product_data(product_id)

                wait.until(EC.element_to_be_clickable((By.CSS_SELECTOR, "a.next-drawer-close"))).click()
                wait.until(EC.invisibility_of_element_located((By.CSS_SELECTOR, ".next-drawer")))
                time.sleep(SLEEP_TINY)

                index += 1
                time.sleep(SLEEP_TINY)
            except Exception as e:
                logger.error(f"处理第 {index + 1} 个产品异常: {e}")
                if product_id:
                    failed_product_ids.add(str(product_id))
                index += 1
                try:
                    if driver.find_elements(By.CSS_SELECTOR, ".next-drawer"):
                        driver.find_element(By.CSS_SELECTOR, "a.next-drawer-close").click()
                except Exception:
                    pass

    def _click_product_next_page():
        try:
            current_page = driver.find_element(By.CSS_SELECTOR, ".next-pagination-display em").text
            next_btn = wait.until(
                EC.presence_of_element_located((By.CSS_SELECTOR, ".detail-table-pagination button.next-next")))
            if not next_btn.is_enabled():
                return False
            driver.execute_script("arguments[0].click();", next_btn)
            wait.until(
                lambda d: d.find_element(By.CSS_SELECTOR, ".next-pagination-display em").text != current_page)
            time.sleep(2)
            return True
        except Exception:
            return False

    try:
        # 登录（使用统一 cookie）
        driver.get(target_url)
        logger.info(f"产品360采集已打开目标页: {driver.current_url}")
        time.sleep(2)
        browser._load_cookies()
        driver.refresh()
        logger.info(f"产品360采集刷新后页面: {driver.current_url}")
        time.sleep(2)

        try:
            wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, ".product-datepicker")))
            logger.info("已通过 Cookie 登录")
        except Exception:
            logger.warning("Cookie 未登录，等待手动登录")
            driver.get(login_url)
            end_time = time.time() + 300
            while time.time() < end_time:
                if "data.alibaba.com" in (driver.current_url or ""):
                    break
                time.sleep(0.4)
            browser._save_cookies()
            driver.get(target_url)
            wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, ".product-datepicker")))

        logger.info("开始采集当前日期...")
        while True:
            if task.should_stop():
                break
            task.wait_if_paused()
            _click_all_analysis_buttons()
            logger.info("翻页...")
            if not _click_product_next_page():
                logger.info("当前日期全部完成")
                break

        # 二次补采：对首次采集异常的产品ID再尝试一次
        if failed_product_ids and not task.should_stop():
            retry_ids = sorted([pid for pid in failed_product_ids if str(pid).strip()])
            logger.info(f"开始二次补采异常产品，共 {len(retry_ids)} 个")
            retry_success = 0
            retry_failed: List[str] = []

            for pid in retry_ids:
                if task.should_stop():
                    break
                task.wait_if_paused()
                try:
                    _collect_and_save_product_data(str(pid))
                    retry_success += 1
                    logger.info(f"二次补采成功: {pid}")
                except Exception as e:
                    retry_failed.append(str(pid))
                    logger.error(f"二次补采失败 {pid}: {e}")

            logger.info(
                f"二次补采完成: 成功 {retry_success} / {len(retry_ids)}，失败 {len(retry_failed)}"
            )
            if retry_failed:
                logger.warning(f"二次补采仍失败产品ID: {', '.join(retry_failed)}")

    finally:
        try:
            driver.quit()
        except Exception:
            pass
        logger.info("采集结束")


# ===================== 产品360：解析 =====================

def _download_product360_parser(task: TaskInfo, cfg):
    """
    产品360 - 模式2：读取 JSON 生成总报告
    来源脚本: cs_产品数据关键词_渠道_曝光国家_正式发布.py
    """
    download_dir = _normalize_download_root(cfg.data_download.product360_output_dir)
    json_output_dir = _normalize_product360_dir(getattr(cfg.data_download, "product360_json_dir", "") or os.path.join(download_dir, "Json文件"))
    keyword_json_dir = _normalize_product360_dir(getattr(cfg.data_download, "product360_keyword_json_dir", "") or os.path.join(download_dir, "关键词json"))
    excel_output_dir = _normalize_product360_dir(getattr(cfg.data_download, "product360_excel_result_dir", "") or os.path.join(download_dir, "Excel结果"))
    os.makedirs(json_output_dir, exist_ok=True)
    os.makedirs(keyword_json_dir, exist_ok=True)
    os.makedirs(excel_output_dir, exist_ok=True)

    consolidated_excel_name = "产品数据总报告.xlsx"

    task.current_step = "扫描 JSON 文件..."
    product_folders = [
        f for f in os.listdir(json_output_dir)
        if os.path.isdir(os.path.join(json_output_dir, f))
    ]
    logger.info(f"JSON目录: {json_output_dir}")
    logger.info(f"发现产品文件夹数量: {len(product_folders)}")
    if not product_folders:
        task.current_step = "未找到 JSON 文件"
        return

    logger.info(f"开始解析 {len(product_folders)} 个产品 JSON")

    all_details_data = []
    all_regions_data = []
    all_channels_data = []
    all_keywords_data = []

    country_code_map = _get_country_code_map()

    for product_id in product_folders:
        if task.should_stop():
            break
        task.wait_if_paused()
        try:
            with open(os.path.join(json_output_dir, product_id, "1_产品详情.json"), 'r', encoding='utf-8') as f:
                detail_json = json.load(f)
            with open(os.path.join(json_output_dir, product_id, "2_访客地域.json"), 'r', encoding='utf-8') as f:
                region_json = json.load(f)
            with open(os.path.join(json_output_dir, product_id, "3_流量来源.json"), 'r', encoding='utf-8') as f:
                channel_json = json.load(f)

            detail_list = detail_json.get('data', [])
            if detail_list and isinstance(detail_list, list):
                all_details_data.append(detail_list[0])

            region_list = region_json.get('data', [])
            if region_list and isinstance(region_list, list):
                for item in region_list:
                    item['productId'] = product_id
                    all_regions_data.append(item)

            channel_list = channel_json.get('data', [])
            if channel_list and isinstance(channel_list, list):
                for channel_group in channel_list:
                    trends_list = channel_group.get('trends', [])
                    if trends_list and isinstance(trends_list, list):
                        for trend_item in trends_list:
                            trend_item['productId'] = product_id
                            all_channels_data.append(trend_item)

            # 关键词解析（完整移植：读取 关键词json/{productId} 下所有分页文件，按关键词去重）
            keyword_product_dir = os.path.join(keyword_json_dir, product_id)
            if os.path.exists(keyword_product_dir):
                product_keywords: Dict[str, Dict] = {}
                kw_files = sorted([f for f in os.listdir(keyword_product_dir) if f.endswith(".json")])
                for kw_file in kw_files:
                    try:
                        with open(os.path.join(keyword_product_dir, kw_file), 'r', encoding='utf-8') as f:
                            kw_json = json.load(f)

                        inner_data = kw_json.get('data', {})
                        word_list = []
                        if isinstance(inner_data, dict):
                            word_list = inner_data.get('list', [])
                        elif isinstance(inner_data, list):
                            word_list = inner_data

                        if not word_list:
                            continue

                        for item in word_list:
                            kw = item.get('searchKeyword') if isinstance(item, dict) else None
                            if not kw:
                                continue
                            if kw in product_keywords:
                                continue

                            p4p_state = item.get('p4pState', {}) if isinstance(item, dict) else {}
                            product_keywords[kw] = {
                                '产品ID': product_id,
                                '关键词': kw,
                                '搜索曝光次数': item.get('searchImps') if isinstance(item, dict) else None,
                                '搜索点击次数': item.get('searchClicks') if isinstance(item, dict) else None,
                                '标准推广曝光次数': item.get('p4pExposureCnt') if isinstance(item, dict) else None,
                                '标准推广点击次数': item.get('p4pClickCnt') if isinstance(item, dict) else None,
                                '商品详情页访问人数': item.get('detailUv') if isinstance(item, dict) else None,
                                '店内询盘人数': item.get('inquiryUv') if isinstance(item, dict) else None,
                                '店内 TM 咨询人数': item.get('tmUv') if isinstance(item, dict) else None,
                                'P4P是否允许添加': p4p_state.get('isAllowAdded') if isinstance(p4p_state, dict) else None,
                                'P4P是否为P4P关键词': p4p_state.get('isP4pKeyword') if isinstance(p4p_state, dict) else None,
                                'P4P推广活动状态': p4p_state.get('p4pCampaignStatus') if isinstance(p4p_state, dict) else None,
                            }
                    except Exception as e:
                        logger.warning(f"解析关键词文件失败: {kw_file} | {e}")

                if product_keywords:
                    all_keywords_data.extend(product_keywords.values())
        except Exception as e:
            logger.warning(f"解析产品 {product_id} 失败: {e}")

    report_path = os.path.join(excel_output_dir, consolidated_excel_name)
    task.current_step = "生成总报告..."

    try:
        with pd.ExcelWriter(report_path, engine='xlsxwriter') as writer:
            if all_details_data:
                df_details = pd.DataFrame(all_details_data)
                detail_map = _get_detail_field_map()
                df_details_renamed = df_details.rename(columns=detail_map)
                final_detail_headers = _get_detail_headers_order()
                existing_headers = [h for h in final_detail_headers if h in df_details_renamed.columns]
                df_details_final = df_details_renamed[existing_headers]
                df_details_final.to_excel(writer, sheet_name='产品详细信息', index=False)

            if all_keywords_data:
                df_keywords = pd.DataFrame(all_keywords_data)
                kw_columns = [
                    '产品ID', '关键词', '搜索曝光次数', '搜索点击次数', '标准推广曝光次数',
                    '标准推广点击次数', '商品详情页访问人数', '店内询盘人数', '店内 TM 咨询人数',
                    'P4P是否允许添加', 'P4P是否为P4P关键词', 'P4P推广活动状态'
                ]
                existing_kw_cols = [c for c in kw_columns if c in df_keywords.columns]
                df_keywords[existing_kw_cols].to_excel(writer, sheet_name='关键词', index=False)

            if all_regions_data:
                df_regions = pd.DataFrame(all_regions_data)
                df_regions['国家(中文)'] = df_regions['countryId'].apply(
                    lambda code: country_code_map.get(str(code).upper(), str(code)))
                df_regions = df_regions.rename(columns={'countryId': '国家代码', 'uv': '访客数(UV)', 'productId': '产品ID'})
                final_region_columns = ['产品ID', '国家代码', '国家(中文)', '访客数(UV)']
                existing_columns = [col for col in final_region_columns if col in df_regions.columns]
                df_regions[existing_columns].to_excel(writer, sheet_name='访客地域', index=False)

            if all_channels_data:
                df_channels = pd.DataFrame(all_channels_data)
                df_channels = df_channels.rename(
                    columns={'statDate': '日期', 'channelType': '流量渠道类型', 'detailUv': '店铺访问人数',
                             'crtOrdUv': '询盘人数', 'tmUv': 'TM咨询人数', 'productId': '产品ID'})
                final_channel_columns = ['产品ID', '日期', '流量渠道类型', '店铺访问人数', '询盘人数', 'TM咨询人数']
                existing_columns = [col for col in final_channel_columns if col in df_channels.columns]
                df_channels[existing_columns].to_excel(writer, sheet_name='流量来源', index=False)

        logger.info(f"总报告已生成: {report_path}")
    except Exception as e:
        logger.error(f"生成总报告失败: {e}")


def _keyword_stat_date_to_short(stat_date: str) -> str:
    text = str(stat_date or "").strip()
    m = re.search(r"(\d{4})-(\d{2})-(\d{2})", text)
    if m:
        return f"{m.group(1)[2:]}{m.group(2)}{m.group(3)}"
    key = _normalize_traffic_channel_date_key(text)
    return key


def _keyword_stat_date_to_filename(stat_date: str) -> str:
    text = str(stat_date or "").strip()
    m = re.search(r"(\d{4})-(\d{2})-(\d{2})", text)
    if m:
        return f"{m.group(1)}-{m.group(2)}-{m.group(3)}"
    return text


def _format_keyword_details(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, list):
        parts = []
        for item in value:
            if isinstance(item, dict):
                alias = str(item.get("keywordAlias") or item.get("alias") or "").strip()
                if alias:
                    parts.append(alias)
            elif item:
                parts.append(str(item))
        return " | ".join(parts)
    if isinstance(value, dict):
        return json.dumps(value, ensure_ascii=False)
    return str(value)


_KEYWORD_NUMERIC_API_KEYS = {
    "sumShowCnt",
    "sumClickCnt",
    "ctr",
    "sumP4pShowCnt",
    "sumP4pClickCnt",
    "avgSumShowCnt",
    "avgSumClickCnt",
    "pv",
    "resultCnt",
}


def _keyword_record_to_row(item: Dict) -> Dict:
    row: Dict[str, Any] = {}
    for api_key, col_name in KEYWORD_API_FIELD_MAP.items():
        if api_key == "keywordDetails":
            row[col_name] = _format_keyword_details(item.get(api_key))
            continue
        val = item.get(api_key)
        if api_key == "keyword":
            row[col_name] = str(val or "").strip()
        elif api_key == "statDate":
            row[col_name] = val
        elif api_key in _KEYWORD_NUMERIC_API_KEYS:
            row[col_name] = pd.to_numeric(val, errors="coerce") if val is not None and val != "" else val
        else:
            row[col_name] = val
    if "词" not in row or not str(row.get("词") or "").strip():
        row["词"] = str(item.get("keyword") or "").strip()
    return row


def _build_keyword_get_url(
    *,
    selected: int,
    page_no: int,
    ctoken: str,
    statistics_type: str = "week",
    order_by: str = "sumShowCnt",
    order_model: str = "desc",
) -> str:
    params = {
        "action": "OneAction",
        "iName": "vip/traffic/keyword/getKeywords",
        "isVip": "true",
        "statisticsType": statistics_type,
        "selected": str(selected),
        "statisticType": "os",
        "orderBy": order_by,
        "orderModel": order_model,
        "pageSize": str(KEYWORD_API_PAGE_SIZE),
        "pageNO": str(page_no),
        "ctoken": ctoken,
    }
    query = "&".join(f"{k}={quote(str(v))}" for k, v in params.items())
    return f"{KEYWORD_GET_API_BASE}?{query}"


def _fetch_keyword_get_page(
    session: requests.Session,
    *,
    selected: int,
    page_no: int,
    ctoken: str,
) -> Dict:
    url = _build_keyword_get_url(selected=selected, page_no=page_no, ctoken=ctoken)
    headers = {
        "User-Agent": (
            "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
            "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"
        ),
        "Referer": KEYWORD_GET_REFERER,
        "Accept": "application/json, text/plain, */*",
    }
    resp = session.get(url, headers=headers, timeout=(8, 30))
    resp.raise_for_status()
    payload = resp.json()
    if not isinstance(payload, dict):
        raise Exception(f"关键词接口返回异常: selected={selected} page={page_no}")
    code = payload.get("code")
    if code not in (None, 0, 200, "0", "200") and payload.get("data") is None:
        raise Exception(f"关键词接口返回失败: selected={selected} page={page_no} | code={code}")
    data = payload.get("data") if isinstance(payload, dict) else {}
    block = data.get("keywords") if isinstance(data, dict) else {}
    if not isinstance(block, dict):
        return {"items": [], "total": 0}
    items = block.get("data") or []
    if not isinstance(items, list):
        items = []
    total = int(block.get("total") or 0)
    return {"items": [x for x in items if isinstance(x, dict)], "total": total}


def _get_keyword_existing_stop_date(download_dir: str) -> str:
    if not download_dir or not os.path.isdir(download_dir):
        return ""
    dates: List[str] = []
    for name in os.listdir(download_dir):
        m = re.search(r"getKeywords-(\d{4})-(\d{2})-(\d{2})", name)
        if m:
            dates.append(f"{m.group(1)[2:]}{m.group(2)}{m.group(3)}")
    if not dates:
        return ""
    return sorted(dates)[-1]


def _save_keyword_week_excel(download_dir: str, stat_date: str, df: pd.DataFrame) -> str:
    file_date = _keyword_stat_date_to_filename(stat_date)
    filename = f"getKeywords-{file_date}.xlsx"
    path = os.path.join(download_dir, filename)
    df.to_excel(path, index=False, engine="openpyxl")
    return path


def _keyword_metric_series(df: pd.DataFrame, column_name: str) -> pd.Series:
    """按「词」去重后的指标序列（index=词），用于批量透视。"""
    if df is None or df.empty or "词" not in df.columns or column_name not in df.columns:
        return pd.Series(dtype=float)
    work = df[["词", column_name]].copy()
    work["词"] = work["词"].map(lambda x: str(x).strip() if pd.notna(x) else "")
    work = work[work["词"] != ""]
    work[column_name] = pd.to_numeric(work[column_name], errors="coerce").fillna(0)
    return work.drop_duplicates(subset=["词"], keep="last").set_index("词")[column_name]


def _keyword_metric_wide_table(all_data: Dict[str, pd.DataFrame], column_name: str, sorted_dates: List[str]) -> pd.DataFrame:
    series_list = [_keyword_metric_series(all_data[d], column_name).rename(d) for d in sorted_dates]
    if not series_list:
        return pd.DataFrame(columns=["词"])
    wide = pd.concat(series_list, axis=1).fillna(0)
    wide.index.name = "词"
    return wide.reset_index()


def _keyword_anomaly_wide_table(metric_wide: pd.DataFrame, sorted_dates: List[str]) -> pd.DataFrame:
    if len(sorted_dates) < 2:
        return pd.DataFrame(columns=["词"])
    indexed = metric_wide.set_index("词")
    parts: List[pd.Series] = []
    for i in range(1, len(sorted_dates)):
        curr_d = sorted_dates[i]
        prev_d = sorted_dates[i - 1]
        if curr_d not in indexed.columns or prev_d not in indexed.columns:
            continue
        parts.append((indexed[curr_d] - indexed[prev_d]).rename(f"{curr_d}异动"))
    if not parts:
        return pd.DataFrame(columns=["词"])
    out = pd.concat(parts, axis=1)
    out.index.name = "词"
    return out.reset_index()


def _build_keyword_summary_and_anomaly(
    all_data: Dict[str, pd.DataFrame],
    output_dir: str,
    task: Optional[TaskInfo] = None,
) -> Dict[str, str]:
    if not all_data:
        raise Exception("没有可分析的关键词数据")

    sorted_dates = sorted(all_data.keys(), reverse=True)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    summary_file = os.path.join(output_dir, f"关键词数据汇总_{ts}.xlsx")
    change_file = os.path.join(output_dir, f"关键词异动分析_{ts}.xlsx")

    if task:
        task.current_step = "生成关键词汇总..."
    t0 = time.time()
    with pd.ExcelWriter(summary_file, engine="openpyxl") as writer:
        for sheet_name, column_name in KEYWORD_SUMMARY_METRICS.items():
            _keyword_metric_wide_table(all_data, column_name, sorted_dates).to_excel(
                writer, sheet_name=sheet_name, index=False
            )
    logger.info(f"关键词汇总生成完成，耗时 {time.time() - t0:.1f}s")

    if task:
        task.current_step = "生成异动分析..."
    t1 = time.time()
    with pd.ExcelWriter(change_file, engine="openpyxl") as writer:
        for sheet_name, column_name in KEYWORD_ANOMALY_METRICS.items():
            metric_wide = _keyword_metric_wide_table(all_data, column_name, sorted_dates)
            _keyword_anomaly_wide_table(metric_wide, sorted_dates).to_excel(
                writer, sheet_name=sheet_name, index=False
            )
    logger.info(f"关键词异动分析生成完成，耗时 {time.time() - t1:.1f}s")

    logger.info(f"关键词分析完成：{summary_file} | {change_file}")
    return {"summary_file": summary_file, "change_file": change_file}


# ===================== 关键词解析（模式2，兼容入口） =====================

def _parse_keywords(task: TaskInfo, cfg):
    """与下载合并为同一流程"""
    _download_keywords(task, cfg)


# ===================== 行业关键词：下载+整合 =====================

def _parse_cookie_file(cookie_file_path: str) -> str:
    """解析配置管理 Cookie 管理保存的 cookies，返回请求头可用 Cookie 字符串。"""
    def _build_cookie_str(cookie_data) -> str:
        cookie_str = ""
        if isinstance(cookie_data, list):
            for item in cookie_data:
                if isinstance(item, dict) and "name" in item and "value" in item:
                    cookie_str += f"{item['name']}={item['value']};"
                elif isinstance(item, (tuple, list)) and len(item) >= 2:
                    cookie_str += f"{item[0]}={item[1]};"
        elif isinstance(cookie_data, dict):
            if isinstance(cookie_data.get("cookies"), list):
                for item in cookie_data.get("cookies") or []:
                    if isinstance(item, dict) and "name" in item and "value" in item:
                        cookie_str += f"{item['name']}={item['value']};"
            else:
                cookie_str = "; ".join([f"{k}={v}" for k, v in cookie_data.items()])
        return cookie_str.strip(";")

    try:
        with open(cookie_file_path, "rb") as f:
            cookie_data = pickle.load(f)
        parsed = _build_cookie_str(cookie_data)
        if parsed:
            return parsed
    except Exception:
        pass

    # 兼容 Cookie 管理旁路排查文件（cookies.pkl.json）或直接 JSON 的 cookies 文件
    for json_path in [f"{cookie_file_path}.json", cookie_file_path]:
        try:
            if not os.path.exists(json_path):
                continue
            with open(json_path, "r", encoding="utf-8") as f:
                payload = json.load(f)
            parsed = _build_cookie_str(payload)
            if parsed:
                return parsed
        except Exception:
            continue
    return ""


def _download_single_industry_keyword(keyword: str, cookie_str: str, save_path: str) -> bool:
    encoded_keyword = quote(keyword)
    url = (
        "https://hz-mydata.alibaba.com/self/excelDownload.do"
        "?&iName=excel/download/vip/kwIndex/searchWords"
        "&terminalType=TOTAL"
        "&dateType=30d"
        "&countryId=TOTAL"
        f"&queryRaw={encoded_keyword}"
        "&nd=30d"
    )
    headers = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
        "Referer": "https://hz-mydata.alibaba.com/",
        "Cookie": cookie_str.encode("utf-8").decode("latin-1"),
    }
    try:
        response = requests.get(url, headers=headers, stream=True, timeout=20, allow_redirects=False, verify=False)
        if response.status_code == 302:
            logger.warning(f"行业关键词下载被重定向(可能Cookie失效): {keyword} -> {response.headers.get('location')}")
            return False
        if response.status_code != 200:
            logger.warning(f"行业关键词下载HTTP失败: {keyword} | status={response.status_code}")
            return False
        if len(response.content) < 1024:
            logger.warning(f"行业关键词下载内容过小: {keyword} | bytes={len(response.content)}")
            return False
        with open(save_path, "wb") as f:
            for chunk in response.iter_content(chunk_size=1024):
                if chunk:
                    f.write(chunk)
        return True
    except Exception:
        return False


def _extract_date_from_filename(filename: str) -> Optional[str]:
    match = re.search(r"(\d{8})", filename)
    if match:
        return match.group(1)
    match = re.search(r"(\d{6})", filename)
    if match:
        date_str = match.group(1)
        year_prefix = "20" if int(date_str[:2]) < 50 else "19"
        return year_prefix + date_str
    return None


def _parse_numeric_value(value):
    if pd.isna(value):
        return 0
    value_str = str(value).strip()
    if value_str in {"", "0"}:
        return 0
    try:
        if "%" in value_str:
            return float(value_str.replace("%", ""))
        return float(value_str)
    except Exception:
        return 0



def _extract_industry_source_keyword(filename: str) -> str:
    """从“源关键词_YYYYMMDD.xlsx”文件名还原源关键词。"""
    stem = os.path.splitext(os.path.basename(str(filename or "")))[0]
    source = re.sub(r"[_-]\d{8}$", "", stem).strip(" _-")
    return source or stem


def _merge_industry_keywords(task: TaskInfo, cfg, source_folder: str, output_file: str):
    metrics_map = {
        "搜索指数": "搜索指数",
        "搜索涨幅": "搜索涨幅",
        "点击率": "点击率",
    }
    keyword_col_name = "关键词"
    red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

    files = [os.path.join(source_folder, n) for n in os.listdir(source_folder) if str(n).lower().endswith(".xlsx")]
    files = [f for f in files if os.path.basename(f) != os.path.basename(output_file)]
    if not files:
        raise FileNotFoundError(f"未在目录找到行业关键词Excel: {source_folder}")

    # 键使用 (源关键词, 关键词)，避免不同源关键词下载结果被揉成一个无法区分的总词池。
    all_data = {sheet: {} for sheet in metrics_map.values()}
    task.current_step = "整合行业关键词数据"
    for idx, file_path in enumerate(files, start=1):
        if task.should_stop():
            break
        task.wait_if_paused()
        task.progress = idx
        task.total = len(files)
        filename = os.path.basename(file_path)
        source_keyword = _extract_industry_source_keyword(filename)
        date_str = _extract_date_from_filename(filename)
        if not date_str:
            mtime = os.path.getmtime(file_path)
            date_str = datetime.fromtimestamp(mtime).strftime("%Y%m%d")
        try:
            df = pd.read_excel(file_path, header=5)
            df.columns = [str(c).strip() for c in df.columns]
        except Exception as e:
            logger.warning(f"行业关键词整合读取失败 {filename}: {e}")
            continue

        if keyword_col_name not in df.columns:
            continue

        for _, row in df.iterrows():
            keyword = row[keyword_col_name]
            if pd.isna(keyword) or str(keyword).strip() == "":
                continue
            keyword = str(keyword).strip()
            data_key = (source_keyword, keyword)

            for source_col, target_sheet in metrics_map.items():
                if source_col not in df.columns:
                    continue
                value = row[source_col]
                if pd.isna(value) or str(value).strip() == "":
                    continue

                if data_key not in all_data[target_sheet]:
                    all_data[target_sheet][data_key] = {}

                current_val = _parse_numeric_value(value)
                if date_str in all_data[target_sheet][data_key]:
                    existing_val = _parse_numeric_value(all_data[target_sheet][data_key][date_str])
                    if current_val > existing_val:
                        all_data[target_sheet][data_key][date_str] = value
                else:
                    all_data[target_sheet][data_key][date_str] = value

    os.makedirs(os.path.dirname(output_file) or source_folder, exist_ok=True)
    with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
        for metric_name, sheet_name in metrics_map.items():
            data_dict = all_data[sheet_name]
            if not data_dict:
                continue

            df_wide = pd.DataFrame.from_dict(data_dict, orient="index")
            date_cols = sorted(df_wide.columns, reverse=True)
            df_wide = df_wide[date_cols]
            df_wide.index = pd.MultiIndex.from_tuples(df_wide.index, names=["源关键词", "关键词"])
            df_wide.reset_index(inplace=True)
            df_wide.to_excel(writer, sheet_name=sheet_name, index=False)

            worksheet = writer.sheets[sheet_name]
            max_row = len(df_wide) + 1
            first_date_col = 3
            last_date_col = worksheet.max_column

            for row_idx in range(2, max_row + 1):
                row_values = []
                for col_idx in range(first_date_col, last_date_col + 1):
                    cell = worksheet.cell(row=row_idx, column=col_idx)
                    if cell.value is not None and str(cell.value).strip() != "":
                        num_val = _parse_numeric_value(cell.value)
                        row_values.append((num_val, col_idx))
                if len(row_values) > 1:
                    unique_values = set(val for val, _ in row_values)
                    if len(unique_values) > 1:
                        _, max_col_idx = max(row_values, key=lambda x: x[0])
                        worksheet.cell(row=row_idx, column=max_col_idx).fill = red_fill


def _download_and_merge_industry_keywords(task: TaskInfo, cfg, big_keywords_override: Optional[str] = None):
    """行业关键词下载：先下载，再自动整合。"""
    icfg = cfg.industry_keyword
    save_folder = _normalize_download_root(getattr(icfg, "save_folder", "") or "")
    output_file_raw = _normalize_download_root(getattr(icfg, "output_file", "") or "")
    big_keywords_text = str(
        big_keywords_override
        if big_keywords_override is not None
        else getattr(icfg, "big_keywords", "") or ""
    )
    delay_seconds = float(getattr(icfg, "delay_seconds", 2.0) or 2.0)
    # 统一复用“配置管理 -> Cookie管理”配置的 cookie
    cookie_file = _normalize_download_root(str(getattr(cfg.paths, "cookie_file", "") or "").strip())

    if not save_folder:
        raise ValueError("行业关键词保存目录未配置")
    if not big_keywords_text.strip():
        raise ValueError("行业关键词大词未配置")
    if not cookie_file:
        raise ValueError("未配置 Cookie 文件，请先在配置管理 -> Cookie管理中完成登录并保存")

    os.makedirs(save_folder, exist_ok=True)

    # 兼容“输出文件”填成目录的情况：自动补默认文件名，避免 Permission denied(目录当文件写)
    if not output_file_raw:
        output_file = os.path.join(save_folder, "关键词数据总表_宽表版.xlsx")
    else:
        normalized = output_file_raw.strip()
        lower_name = os.path.basename(normalized).lower()
        if os.path.isdir(normalized) or not lower_name.endswith((".xlsx", ".xls")):
            output_file = os.path.join(normalized, "关键词数据总表_宽表版.xlsx")
        else:
            output_file = normalized

    keyword_list = [k.strip() for k in re.split(r"[,，;；\n\r]+", big_keywords_text) if k.strip()]
    if not keyword_list:
        raise ValueError("行业关键词大词解析为空")

    logger.info(f"行业关键词下载任务开始，共 {len(keyword_list)} 个大词: {', '.join(keyword_list)}")

    cookie_str = _parse_cookie_file(cookie_file)
    if not cookie_str:
        raise ValueError("Cookie解析失败，请在配置管理 -> Cookie管理中重新登录并保存Cookie")

    task.current_step = "行业关键词下载中"
    task.total = len(keyword_list)
    task.progress = 0
    today_str = datetime.now().strftime("%Y%m%d")

    success_count = 0
    for i, kw in enumerate(keyword_list, start=1):
        if task.should_stop():
            break
        task.wait_if_paused()
        task.current_step = f"下载行业关键词: {kw}"
        filename = f"{kw}_{today_str}.xlsx"
        save_path = os.path.join(save_folder, filename)
        ok = _download_single_industry_keyword(kw, cookie_str, save_path)
        if ok:
            success_count += 1
            logger.info(f"行业关键词下载成功: {kw} -> {filename}")
        else:
            logger.warning(f"行业关键词下载失败: {kw}")
        task.progress = i
        time.sleep(max(0.0, delay_seconds))

    if success_count == 0:
        raise ValueError("行业关键词下载全部失败，请检查Cookie有效性与关键词")

    task.current_step = "行业关键词下载完成，开始关键词整合"
    _merge_industry_keywords(task, cfg, save_folder, output_file)
    task.current_step = f"行业关键词任务完成：下载成功 {success_count}/{len(keyword_list)}，整合完成"



def _load_existing_dropdown_rows(output_file: str) -> List[Dict[str, Any]]:
    """读取已有下拉词结果，并兼容旧的“原词、US”两列表格。"""
    if not output_file or not os.path.exists(output_file):
        return []
    rows: List[Dict[str, Any]] = []
    try:
        wb = load_workbook(output_file, read_only=True, data_only=True)
        ws = wb.active
        headers = [str(cell.value or "").strip() for cell in ws[1]]
        header_map = {name: idx for idx, name in enumerate(headers) if name}
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row:
                continue

            def get_value(name: str, fallback_index: Optional[int] = None):
                idx = header_map.get(name)
                if idx is not None and idx < len(row):
                    return row[idx]
                if fallback_index is not None and fallback_index < len(row):
                    return row[fallback_index]
                return ""

            origin = str(get_value("原词", 0) or "").strip()
            word = str(get_value("下拉词") or get_value("US", 1) or "").strip()
            source = str(get_value("源关键词") or origin).strip()
            heat = _parse_numeric_value(get_value("关键词热度"))
            if source or origin or word:
                rows.append({"源关键词": source, "原词": origin, "关键词热度": heat, "US": word})
        wb.close()
    except Exception as e:
        logger.warning(f"读取已有下拉词文件失败: {output_file} | {e}")
    return rows


def _build_industry_keyword_source_lookup(output_file: str) -> Dict[str, List[Dict[str, Any]]]:
    """按行业关键词建立其所属源关键词池和最新热度的映射。"""
    lookup: Dict[str, List[Dict[str, Any]]] = {}
    table = get_industry_keyword_latest(output_file)
    latest_col = str(table.get("latest_col") or "")
    for row in table.get("rows", []) or []:
        keyword = str((row or {}).get("关键词", "") or "").strip()
        if not keyword:
            continue
        source = str((row or {}).get("源关键词", "") or keyword).strip()
        heat = _parse_numeric_value((row or {}).get(latest_col)) if latest_col else 0
        key = keyword.lower()
        values = lookup.setdefault(key, [])
        identity = source.lower()
        if any(str(v.get("source", "")).lower() == identity for v in values):
            continue
        values.append({"source": source, "heat": heat})
    return lookup


def _download_industry_keyword_dropdown(task: TaskInfo, cfg, dropdown_keywords_override: Optional[str] = None):
    """行业关键词下拉词下载，并保留源关键词池及原词热度。"""
    icfg = cfg.industry_keyword
    raw_keywords = str(
        dropdown_keywords_override
        if dropdown_keywords_override is not None
        else getattr(icfg, "dropdown_keywords", "") or ""
    )
    output_file_raw = _normalize_download_root(str(getattr(icfg, "dropdown_output_file", "") or ""))
    industry_output_file = _normalize_download_root(str(getattr(icfg, "output_file", "") or ""))
    cookie_file = _normalize_download_root(str(getattr(cfg.paths, "cookie_file", "") or "").strip())

    if not raw_keywords.strip():
        raise ValueError("下拉词关键词未配置，请在配置区域填写或从表格选择后再执行")
    if not output_file_raw.strip():
        raise ValueError("下拉词输出文件未配置")
    if not cookie_file:
        raise ValueError("未配置 Cookie 文件，请先在配置管理 -> Cookie管理中完成登录并保存")

    if os.path.isdir(output_file_raw) or not os.path.basename(output_file_raw).lower().endswith((".xlsx", ".xls")):
        output_file = os.path.join(output_file_raw, "下拉词结果.xlsx")
    else:
        output_file = output_file_raw
    os.makedirs(os.path.dirname(output_file) or ".", exist_ok=True)

    cookie_str = _parse_cookie_file(cookie_file)
    if not cookie_str:
        raise ValueError("Cookie解析失败，请在配置管理 -> Cookie管理中重新登录并保存Cookie")

    keywords = [k.strip() for k in re.split(r"[,，;；\n\r]+", raw_keywords) if k.strip()]
    unique_keywords: List[str] = []
    seen = set()
    for kw in keywords:
        key = kw.lower()
        if key in seen:
            continue
        seen.add(key)
        unique_keywords.append(kw)
    keywords = unique_keywords
    if not keywords:
        raise ValueError("下拉词关键词解析为空")

    logger.info(f"行业关键词下拉词任务开始，共 {len(keywords)} 个原词: {', '.join(keywords)}")

    source_lookup = _build_industry_keyword_source_lookup(industry_output_file)
    delay_seconds = float(getattr(icfg, "delay_seconds", 2.0) or 2.0)
    batch_origin_keys = {kw.lower() for kw in keywords}
    existing_rows = _load_existing_dropdown_rows(output_file)
    preserved_rows = [
        row
        for row in existing_rows
        if str((row or {}).get("原词", "")).strip().lower() not in batch_origin_keys
    ]

    task.current_step = "下拉词下载中"
    task.total = len(keywords)
    task.progress = 0

    url_tpl = (
        "https://open-s.alibaba.com/openservice/associationSuggestionViewService"
        "?keywords={}&name=home_login_first_screen&bizScene=pcHomeProducts&lang=en&country=US"
    )
    headers = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/134.0.0.0 Safari/537.36",
        "Accept-Language": "en-US,en;q=0.9",
        "Referer": "https://www.alibaba.com/",
        "Origin": "https://www.alibaba.com",
        "Cookie": cookie_str.encode("utf-8").decode("latin-1"),
    }

    # 输出保留：源关键词池、下载原词、原词最新热度、下拉词。
    wb = Workbook()
    ws = wb.active
    ws.title = "下拉词"
    ws.append(["源关键词", "原词", "关键词热度", "US"])

    exist_words = set()
    green_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
    total_count = 0
    preserved_origin_count = len({str((r or {}).get("原词", "")).strip().lower() for r in preserved_rows if str((r or {}).get("原词", "")).strip()})

    for row in preserved_rows:
        source = str((row or {}).get("源关键词", "") or (row or {}).get("原词", "")).strip()
        origin = str((row or {}).get("原词", "")).strip()
        heat = _parse_numeric_value((row or {}).get("关键词热度"))
        word = str((row or {}).get("US", "") or (row or {}).get("下拉词", "")).strip()
        key = (source.lower(), word.lower())
        if not word or key in exist_words:
            continue
        ws.append([source, origin, heat, word])
        exist_words.add(key)
        total_count += 1

    new_count = 0
    for idx, kw in enumerate(keywords, start=1):
        if task.should_stop():
            break
        task.wait_if_paused()
        task.current_step = f"下载下拉词: {kw} ({idx}/{len(keywords)})"
        task.progress = idx
        logger.info(f"行业关键词下拉词下载 ({idx}/{len(keywords)}): {kw}")

        suggests: List[str] = []
        try:
            url = url_tpl.format(quote(kw))
            resp = requests.get(url, headers=headers, timeout=15)
            payload = resp.json() if resp.ok else {}
            for item in (((payload or {}).get("data") or {}).get("list") or []):
                word = str((item or {}).get("suggestKeyword") or "").strip()
                if word:
                    suggests.append(word)
            if not suggests:
                logger.warning(f"下拉词无结果: {kw} | status={resp.status_code}")
        except Exception as e:
            logger.warning(f"下拉词请求失败: {kw} | {e}")

        source_infos = source_lookup.get(kw.lower()) or [{"source": kw, "heat": 0}]
        wrote = 0
        for info in source_infos:
            source = str((info or {}).get("source", "") or kw).strip()
            heat = _parse_numeric_value((info or {}).get("heat"))
            for word in suggests:
                key = (source.lower(), word.lower())
                if key in exist_words:
                    continue
                ws.append([source, kw, heat, word])
                exist_words.add(key)
                total_count += 1
                new_count += 1
                wrote += 1
                if wrote == 1:
                    ws.cell(row=ws.max_row, column=1).fill = green_fill

        if idx < len(keywords):
            time.sleep(max(0.0, delay_seconds))

    wb.save(output_file)
    task.current_step = (
        f"下拉词下载完成：本次 {len(keywords)} 个原词，新增 {new_count} 条，"
        f"保留历史 {preserved_origin_count} 个原词，合计 {total_count} 条，文件 {os.path.basename(output_file)}"
    )
    logger.info(f"行业关键词{task.current_step}")


# ===================== 其他下载（原占位） =====================

def _download_product_ranking(task: TaskInfo, cfg, date_range: Optional[str]):
    """
    产品参谋 - 产品排名/访客/渠道下载
    重构自: cs_产品参谋_产品排名，访客，渠道下载.py
    """
    task.current_step = "启动浏览器..."

    # TODO: 从原脚本迁移具体的 Selenium 抓取逻辑
    from app.services.automation.browser_manager import BrowserManager
    browser = BrowserManager()

    try:
        if not browser.setup():
            raise Exception("浏览器启动失败")

        task.current_step = "登录中..."
        if not browser.login():
            raise Exception("登录失败")

        task.current_step = "正在抓取数据..."
        logger.info("产品排名数据下载 - 待实现具体抓取逻辑")

    finally:
        browser.quit()


def _dismiss_page_overlays(driver) -> None:
    """关闭日期选择器等留下的遮罩，避免导出按钮 click intercepted。"""
    from selenium.webdriver.common.by import By
    from selenium.webdriver.common.keys import Keys
    from selenium.webdriver.common.action_chains import ActionChains

    for _ in range(3):
        try:
            for sel in (
                ".next-overlay-backdrop",
                ".next-overlay-wrapper.open .next-overlay-backdrop",
                ".next-overlay-wrapper.open",
            ):
                for el in driver.find_elements(By.CSS_SELECTOR, sel):
                    try:
                        driver.execute_script("arguments[0].click();", el)
                    except Exception:
                        pass
        except Exception:
            pass
        try:
            ActionChains(driver).send_keys(Keys.ESCAPE).perform()
        except Exception:
            pass
        time.sleep(0.15)

    try:
        backdrops = driver.find_elements(By.CSS_SELECTOR, ".next-overlay-backdrop")
        if backdrops:
            driver.execute_script(
                "document.querySelectorAll('.next-overlay-backdrop,.next-overlay-wrapper.open')"
                ".forEach(el=>{try{el.click()}catch(e){}});"
            )
            time.sleep(0.2)
    except Exception:
        pass


def _click_product_export(wait, driver) -> None:
    """点击产品参谋导出按钮（优先 JS 点击，避免遮罩拦截）。"""
    from selenium.webdriver.common.by import By
    from selenium.webdriver.support import expected_conditions as EC

    last_err = None
    for attempt in range(4):
        _dismiss_page_overlays(driver)
        try:
            export_btn = wait.until(
                EC.presence_of_element_located((By.CSS_SELECTOR, "a.product-effective-download"))
            )
            driver.execute_script("arguments[0].scrollIntoView({block:'center'});", export_btn)
            time.sleep(0.15)
            driver.execute_script("arguments[0].click();", export_btn)
            return
        except Exception as e:
            last_err = e
            time.sleep(0.35)
    raise last_err or RuntimeError("无法点击导出按钮")


def _finalize_products_download(daily_download_dir: str, current_date: str, new_file: Optional[str]) -> bool:
    """将浏览器下载的文件规范为 Products-YYYY-MM-DD.xls。"""
    if not new_file:
        return False
    target_name = f"Products-{current_date}.xls"
    src_path = os.path.join(daily_download_dir, new_file)
    if not os.path.isfile(src_path):
        return False
    if new_file.lower() in ("downloads.htm", "downloads.html") or new_file.endswith(".crdownload"):
        try:
            os.remove(src_path)
        except Exception:
            pass
        return False
    target_path = os.path.join(daily_download_dir, target_name)
    if os.path.abspath(src_path) != os.path.abspath(target_path):
        if os.path.exists(target_path):
            os.remove(target_path)
        os.rename(src_path, target_path)
    return True


# ===================== 产品参谋日期周期控件 =====================
# DOM 参考: 产品参谋周期接环.txt / cs_产品数据下载_正式发布.py
# - 周期下拉: .datepicker-container .next-select-trigger
# - 周期菜单: ul.static-menu > li > span (日/周/月)
# - 日期输入: input[placeholder='请选择日期']
# - 日历单元: td.next-calendar-cell[title='YYYY-MM-DD'] 或 div.next-calendar-date
# - 翻页按钮: button.next-prev / .datepicker-arrow button.next-pagination-item.next-prev


def _legacy_overview_setup_day_mode(wait, driver) -> bool:
    """
    老脚本 cs_产品数据下载_正式发布.py 的日粒度初始化流程（已验证可下载）。
    """
    from selenium.webdriver.common.by import By
    from selenium.webdriver.support import expected_conditions as EC
    from selenium.webdriver.common.action_chains import ActionChains

    try:
        date_trigger = wait.until(
            EC.element_to_be_clickable((By.CSS_SELECTOR, ".datepicker-container .next-select-trigger"))
        )
        date_trigger.click()
        time.sleep(1)

        day_option = wait.until(
            EC.element_to_be_clickable((By.XPATH, "//ul[@class='static-menu']/li[span[text()='日']]"))
        )
        day_option.click()
        time.sleep(1)

        wait.until(EC.presence_of_element_located((By.CLASS_NAME, "next-date-picker")))
        date_cells = wait.until(
            EC.presence_of_all_elements_located(
                (
                    By.XPATH,
                    "//td[not(contains(@class, 'next-disabled'))]//div[@class='next-calendar-date']",
                )
            )
        )
        picked = False
        for elem in reversed(date_cells):
            date_text = (elem.text or "").strip()
            if date_text.isdigit():
                logger.info(f"老脚本流程选中日历日: {date_text}")
                elem.click()
                picked = True
                break
        if not picked:
            logger.warning("老脚本流程：未找到可选日历日期")
            return False

        blank = driver.find_element(By.CSS_SELECTOR, "body")
        ActionChains(driver).move_to_element(blank).click().perform()
        time.sleep(2)
        return bool(_overview_extract_date(driver))
    except Exception as e:
        logger.warning(f"老脚本日粒度流程失败: {e}")
        return False


def _legacy_overview_go_prev_day(
    driver,
    wait,
    prev_date: str,
    task: Optional[TaskInfo] = None,
) -> Optional[str]:
    """老脚本翻页：button.next-prev"""
    from selenium.webdriver.common.by import By

    try:
        prev_btn = driver.find_element(By.CSS_SELECTOR, "button.next-prev")
        cls = (prev_btn.get_attribute("class") or "").lower()
        disabled_attr = str(prev_btn.get_attribute("disabled") or "").lower()
        if "disabled" in cls or disabled_attr in {"true", "disabled"} or not prev_btn.is_enabled():
            return None
        prev_btn.click()
        if not _sleep_with_stop(task, 3):
            return None
        actual = _overview_extract_date(driver)
        if actual and actual != prev_date:
            return actual
    except Exception:
        pass
    return None


def _overview_close_date_picker(driver) -> None:
    """关闭周期弹层与日历遮罩。"""
    _dismiss_page_overlays(driver)
    from selenium.webdriver.common.by import By

    try:
        body = driver.find_element(By.CSS_SELECTOR, "body")
        driver.execute_script("arguments[0].click();", body)
    except Exception:
        pass
    time.sleep(0.35)
    _dismiss_page_overlays(driver)


def _overview_open_period_popup(wait, driver) -> bool:
    """点击时间选择框，展开 select-date-popup。"""
    from selenium.webdriver.common.by import By
    from selenium.webdriver.support import expected_conditions as EC

    _overview_close_date_picker(driver)
    openers = [
        ".product-datepicker .datepicker-container .next-select-trigger",
        ".datepicker-container .next-select-trigger",
        ".product-datepicker .next-select-trigger",
    ]
    for sel in openers:
        try:
            el = wait.until(EC.element_to_be_clickable((By.CSS_SELECTOR, sel)))
            driver.execute_script("arguments[0].scrollIntoView({block:'center'});", el)
            try:
                el.click()
            except Exception:
                driver.execute_script("arguments[0].click();", el)
            time.sleep(0.4)
            if driver.find_elements(
                By.CSS_SELECTOR,
                "ul.static-menu, .select-date-popup, .next-select-popup-wrap, .next-date-picker",
            ):
                return True
        except Exception:
            continue
    return False


def _overview_get_period_type(driver) -> Optional[str]:
    """读取当前周期粒度：日 / 周 / 月。"""
    from selenium.webdriver.common.by import By

    for li in driver.find_elements(By.CSS_SELECTOR, "ul.static-menu li.active"):
        for span in li.find_elements(By.TAG_NAME, "span"):
            text = (span.text or "").strip()
            if text in {"日", "周", "月"}:
                return text

    try:
        label = driver.find_element(By.CSS_SELECTOR, ".datepicker-container .next-input-label span")
        text = (label.get_attribute("title") or label.text or "").strip()
        if re.search(r"第\d+周", text):
            return "周"
        if re.fullmatch(r"\d{4}-\d{2}-\d{2}", text):
            return "日"
    except Exception:
        pass

    try:
        date_input = driver.find_element(By.CSS_SELECTOR, "input[placeholder='请选择日期']")
        if (date_input.get_attribute("value") or "").strip():
            return "日"
    except Exception:
        pass
    return None


def _overview_extract_date(driver) -> Optional[str]:
    """从页面读取当前选中日期 YYYY-MM-DD（日粒度优先读 input）。"""
    from selenium.webdriver.common.by import By

    candidates: list[str] = []
    try:
        el = driver.find_element(By.CSS_SELECTOR, "input[placeholder='请选择日期']")
        candidates.append((el.get_attribute("value") or "").strip())
    except Exception:
        pass

    try:
        label = driver.find_element(By.CSS_SELECTOR, ".next-input-label span[title*='更新周期']")
        for raw in ((label.text or "").strip(), (label.get_attribute("title") or "").strip()):
            if raw:
                candidates.append(raw)
    except Exception:
        pass

    try:
        el = driver.find_element(By.CSS_SELECTOR, ".next-date-picker-panel-input input")
        candidates.append((el.get_attribute("value") or "").strip())
    except Exception:
        pass

    for cell in driver.find_elements(By.CSS_SELECTOR, "td.next-calendar-cell.next-selected[title]"):
        candidates.append((cell.get_attribute("title") or "").strip())

    for sel in (
        ".datepicker-container .next-input-label span",
        ".product-datepicker .next-input-label span",
        ".next-select-trigger",
    ):
        try:
            for el in driver.find_elements(By.CSS_SELECTOR, sel):
                for raw in ((el.text or "").strip(), (el.get_attribute("title") or "").strip()):
                    if raw:
                        candidates.append(raw)
        except Exception:
            pass

    for text in candidates:
        if not text:
            continue
        if re.fullmatch(r"\d{4}-\d{2}-\d{2}", text):
            return text
        dates = re.findall(r"\d{4}-\d{2}-\d{2}", text)
        if dates:
            return dates[-1]
    return None


def _overview_select_period_type(wait, driver, period: str) -> bool:
    """在 static-menu 中切换周期：日 / 周 / 月。"""
    from selenium.webdriver.common.by import By
    from selenium.webdriver.support import expected_conditions as EC

    if _overview_get_period_type(driver) == period:
        return True

    if not _overview_open_period_popup(wait, driver):
        return False

    xpath_candidates = [
        f"//ul[@class='static-menu']/li[span[text()='{period}']]",
        (
            f"//div[contains(@class,'select-date-popup') or contains(@class,'next-select-popup-wrap')]"
            f"//ul[contains(@class,'static-menu')]"
            f"//li[.//span[normalize-space(text())='{period}']]"
        ),
        f"//ul[contains(@class,'static-menu')]//li[.//span[normalize-space(text())='{period}']]",
    ]
    option = None
    last_err: Optional[Exception] = None
    for xpath in xpath_candidates:
        try:
            option = wait.until(EC.element_to_be_clickable((By.XPATH, xpath)))
            break
        except Exception as e:
            last_err = e
            continue
    if option is None:
        logger.warning(f"选择周期「{period}」失败: {last_err}")
        return False

    try:
        option.click()
    except Exception:
        driver.execute_script("arguments[0].click();", option)
    time.sleep(0.5)

    try:
        wait.until(
            lambda d: _overview_get_period_type(d) == period
            or bool(d.find_elements(By.CSS_SELECTOR, "input[placeholder='请选择日期']"))
        )
    except Exception:
        pass
    return _overview_get_period_type(driver) == period or bool(
        driver.find_elements(By.CSS_SELECTOR, "input[placeholder='请选择日期']")
    )


def _overview_open_calendar_panel(wait, driver) -> bool:
    """在弹层内打开日历面板。"""
    from selenium.webdriver.common.by import By
    from selenium.webdriver.support import expected_conditions as EC

    if driver.find_elements(By.CSS_SELECTOR, ".next-calendar-table, .next-date-picker-body"):
        return True

    if not _overview_open_period_popup(wait, driver):
        return False

    openers = [
        "input[placeholder='请选择日期']",
        ".next-date-picker-trigger",
        ".next-date-picker-input",
    ]
    for sel in openers:
        try:
            el = wait.until(EC.element_to_be_clickable((By.CSS_SELECTOR, sel)))
            try:
                el.click()
            except Exception:
                driver.execute_script("arguments[0].click();", el)
            time.sleep(0.35)
            if driver.find_elements(By.CSS_SELECTOR, ".next-calendar-table, .next-date-picker-body"):
                return True
        except Exception:
            continue
    return False


def _overview_list_selectable_dates(driver) -> list[str]:
    """列出日历中所有可选日期（td[title] 且非 disabled）。"""
    from selenium.webdriver.common.by import By

    dates: list[str] = []
    for cell in driver.find_elements(
        By.CSS_SELECTOR,
        "td.next-calendar-cell[title]:not(.next-disabled)",
    ):
        title = (cell.get_attribute("title") or "").strip()
        aria_disabled = (cell.get_attribute("aria-disabled") or "").strip().lower()
        if title and re.fullmatch(r"\d{4}-\d{2}-\d{2}", title) and aria_disabled not in {"true", "1"}:
            dates.append(title)
    return sorted(set(dates))


def _overview_read_calendar_month_year(driver) -> Optional[tuple[int, int]]:
    from selenium.webdriver.common.by import By

    for sel in (
        ".next-calendar-panel-header-full",
        ".next-calendar-header-full",
        ".next-calendar-panel-header",
    ):
        for el in driver.find_elements(By.CSS_SELECTOR, sel):
            text = (el.text or "").strip()
            m = re.search(r"(\d{4})\D+(\d{1,2})", text)
            if m:
                return int(m.group(1)), int(m.group(2))
    return None


def _overview_click_calendar_month_nav(driver, direction: str) -> bool:
    from selenium.webdriver.common.by import By

    selectors = (
        [
            ".next-calendar-btn-prev-month",
            ".next-calendar-btn-prev",
            "button.next-calendar-btn-prev-month",
        ]
        if direction == "prev"
        else [
            ".next-calendar-btn-next-month",
            ".next-calendar-btn-next",
            "button.next-calendar-btn-next-month",
        ]
    )
    for sel in selectors:
        for btn in driver.find_elements(By.CSS_SELECTOR, sel):
            if not btn.is_displayed():
                continue
            try:
                driver.execute_script("arguments[0].click();", btn)
                time.sleep(0.35)
                return True
            except Exception:
                continue
    return False


def _overview_select_calendar_date(wait, driver, target_date: str) -> bool:
    """通过 td[title] 选择指定日期。"""
    from selenium.webdriver.common.by import By

    try:
        target_dt = datetime.strptime(target_date, "%Y-%m-%d")
    except Exception:
        return False

    if not _overview_open_calendar_panel(wait, driver):
        return False

    for _ in range(36):
        cells = driver.find_elements(
            By.CSS_SELECTOR,
            f"td.next-calendar-cell[title='{target_date}']:not(.next-disabled)",
        )
        if cells:
            cell = cells[-1]
            try:
                driver.execute_script("arguments[0].scrollIntoView({block:'center'});", cell)
                driver.execute_script("arguments[0].click();", cell)
            except Exception:
                cell.click()
            time.sleep(0.55)
            _overview_close_date_picker(driver)
            return _overview_extract_date(driver) == target_date

        cur = _overview_read_calendar_month_year(driver)
        if not cur or (cur[0], cur[1]) > (target_dt.year, target_dt.month):
            if not _overview_click_calendar_month_nav(driver, "prev"):
                break
        elif cur and (cur[0], cur[1]) < (target_dt.year, target_dt.month):
            if not _overview_click_calendar_month_nav(driver, "next"):
                break
        else:
            break

    _overview_close_date_picker(driver)
    return False


def _overview_select_latest_available_date(wait, driver) -> Optional[str]:
    """选择日历中最新可选日期。"""
    from selenium.webdriver.common.by import By
    from selenium.webdriver.support import expected_conditions as EC
    from selenium.webdriver.common.action_chains import ActionChains

    if not _overview_open_calendar_panel(wait, driver):
        return None

    try:
        wait.until(EC.presence_of_element_located((By.CLASS_NAME, "next-date-picker")))
        date_cells = driver.find_elements(
            By.XPATH,
            "//td[not(contains(@class, 'next-disabled'))]//div[@class='next-calendar-date']",
        )
        for elem in reversed(date_cells):
            if (elem.text or "").strip().isdigit():
                try:
                    elem.click()
                except Exception:
                    driver.execute_script("arguments[0].click();", elem)
                time.sleep(0.55)
                blank = driver.find_element(By.CSS_SELECTOR, "body")
                ActionChains(driver).move_to_element(blank).click().perform()
                time.sleep(0.5)
                current = _overview_extract_date(driver)
                if current:
                    return current
                break
    except Exception:
        pass

    dates = _overview_list_selectable_dates(driver)
    if not dates:
        current = _overview_extract_date(driver)
        _overview_close_date_picker(driver)
        return current

    latest = dates[-1]
    if _overview_select_calendar_date(wait, driver, latest):
        return latest

    current = _overview_extract_date(driver)
    if current:
        return current
    return latest


def _overview_set_date_input(wait, driver, target_date: str) -> bool:
    """通过弹层内 YYYY-MM-DD 输入框设置日期。"""
    from selenium.webdriver.common.by import By
    from selenium.webdriver.common.keys import Keys
    from selenium.webdriver.support import expected_conditions as EC

    if not _overview_open_calendar_panel(wait, driver):
        if not _overview_open_period_popup(wait, driver):
            return False

    input_selectors = [
        ".next-date-picker-panel-input input[placeholder='YYYY-MM-DD']",
        "input[placeholder='YYYY-MM-DD']",
        "input[placeholder='请选择日期']",
    ]
    for sel in input_selectors:
        try:
            inp = wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, sel)))
            driver.execute_script("arguments[0].scrollIntoView({block:'center'});", inp)
            inp.click()
            time.sleep(0.1)
            inp.send_keys(Keys.CONTROL, "a")
            inp.send_keys(target_date)
            inp.send_keys(Keys.ENTER)
            time.sleep(0.75)
            _overview_close_date_picker(driver)
            if _overview_extract_date(driver) == target_date:
                return True
        except Exception:
            continue
    return False


def _overview_click_period_pagination(
    driver,
    wait,
    direction: str,
    prev_date: Optional[str] = None,
    task: Optional[TaskInfo] = None,
) -> Optional[str]:
    """
    点击 .datepicker-arrow 中的上一页/下一页按钮切换周期。
    direction: prev / next
    """
    from selenium.webdriver.common.by import By

    btn_suffix = "next-prev" if direction == "prev" else "next-next"
    aria_hint = "上一页" if direction == "prev" else "下一页"
    selectors = [
        f"button.{btn_suffix}",
        f".datepicker-arrow button.{btn_suffix}",
        f".datepicker-arrow button.next-pagination-item.{btn_suffix}",
        f".product-datepicker .datepicker-arrow button.{btn_suffix}",
        f"button.next-pagination-item.{btn_suffix}",
        f"button[aria-label*='{aria_hint}']",
    ]

    for _attempt in range(5):
        if task and task.should_stop():
            return None
        btn = None
        has_control = False
        for sel in selectors:
            for elem in driver.find_elements(By.CSS_SELECTOR, sel):
                has_control = True
                cls = (elem.get_attribute("class") or "").lower()
                disabled_attr = str(elem.get_attribute("disabled") or "").lower()
                aria_disabled = str(elem.get_attribute("aria-disabled") or "").lower()
                if "disabled" in cls or disabled_attr in {"true", "disabled"} or aria_disabled == "true":
                    continue
                btn = elem
                break
            if btn:
                break

        if not btn:
            if has_control:
                return None
            if not _sleep_with_stop(task, 0.5):
                return None
            continue

        marker_before = _overview_period_marker(driver)
        try:
            driver.execute_script("arguments[0].scrollIntoView({block:'center'});", btn)
            try:
                btn.click()
            except Exception:
                driver.execute_script("arguments[0].click();", btn)
        except Exception:
            if not _sleep_with_stop(task, 0.5):
                return None
            continue

        if not _sleep_with_stop(task, 0.75):
            return None
        new_date = _overview_extract_date(driver)
        marker_after = _overview_period_marker(driver)
        if prev_date:
            if new_date and new_date != prev_date:
                if not _sleep_with_stop(task, 0.35):
                    return None
                return new_date
        elif marker_after and marker_after != marker_before:
            return new_date or "__moved__"

    return None


def _overview_click_prev_period(
    driver,
    wait,
    prev_date: Optional[str] = None,
    task: Optional[TaskInfo] = None,
) -> str:
    """返回 moved / end / fail。"""
    from selenium.webdriver.common.by import By

    has_prev = bool(
        driver.find_elements(By.CSS_SELECTOR, ".datepicker-arrow button.next-pagination-item.next-prev")
    )
    new_date = _overview_click_period_pagination(driver, wait, "prev", prev_date, task)
    if new_date:
        return "moved"

    if has_prev:
        for elem in driver.find_elements(By.CSS_SELECTOR, ".datepicker-arrow button.next-pagination-item.next-prev"):
            cls = (elem.get_attribute("class") or "").lower()
            disabled_attr = str(elem.get_attribute("disabled") or "").lower()
            if "disabled" not in cls and disabled_attr not in {"true", "disabled"}:
                return "fail"
        return "end"
    return "fail"


def _overview_switch_to_day_granularity(wait, driver, target_url: str) -> bool:
    """切换到日粒度并选中最新可选日期。"""
    from selenium.webdriver.common.by import By

    logger.info("切换到日粒度")

    for attempt, refresh in enumerate((False, True)):
        if refresh:
            logger.info("首次切换日粒度未成功，刷新页面后重试")
            driver.get(target_url)
            try:
                wait.until(lambda d: d.find_elements(By.CSS_SELECTOR, ".product-datepicker"))
            except Exception:
                logger.warning("切换到日粒度失败：页面未加载 product-datepicker")
                return False
            time.sleep(0.8)

        _overview_close_date_picker(driver)

        if _legacy_overview_setup_day_mode(wait, driver):
            current = _overview_extract_date(driver)
            if current:
                logger.info(f"日粒度当前日期(老脚本流程): {current}")
                _overview_close_date_picker(driver)
                return True

        if not _overview_select_period_type(wait, driver, "日"):
            logger.warning(f"选择日粒度失败 (attempt {attempt + 1})")
            continue

        current = _overview_extract_date(driver)
        if current:
            logger.info(f"日粒度当前日期(切换后): {current}")
            _overview_close_date_picker(driver)
            return True

        latest = _overview_select_latest_available_date(wait, driver)
        _overview_close_date_picker(driver)
        if latest:
            logger.info(f"日粒度选中最新日期: {latest}")
            return True

        logger.warning(f"切换到日粒度后无法读取日期 (attempt {attempt + 1})")

    return False


def _overview_get_period_label(driver) -> str:
    """读取周期控件当前展示文本（如 2026 第19周 (PST)）。"""
    from selenium.webdriver.common.by import By

    try:
        el = driver.find_element(By.CSS_SELECTOR, ".datepicker-container .next-input-label span")
        return (el.get_attribute("title") or el.text or "").strip()
    except Exception:
        return ""


def _overview_period_marker(driver) -> str:
    date_part = _overview_extract_date(driver) or ""
    label_part = _overview_get_period_label(driver) or ""
    return f"{date_part}|{label_part}"


def _get_daily_staging_download_dir() -> str:
    """Chrome 落盘用纯英文临时目录，避免中文路径导致「出了点问题」。"""
    staging = os.path.join(tempfile.gettempdir(), "ali_product_daily_dl")
    os.makedirs(staging, exist_ok=True)
    return staging


def _set_browser_download_dir(driver, download_dir: str) -> None:
    """设置 Chrome 下载目录（prefs + CDP 双保险）。"""
    abs_path = os.path.abspath(download_dir)
    os.makedirs(abs_path, exist_ok=True)
    for path in (abs_path, abs_path.replace("\\", "/")):
        payload = {"behavior": "allow", "downloadPath": path, "eventsEnabled": True}
        for cmd in ("Page.setDownloadBehavior", "Browser.setDownloadBehavior"):
            try:
                driver.execute_cdp_cmd(cmd, payload)
            except Exception:
                try:
                    driver.execute_cdp_cmd(cmd, {"behavior": "allow", "downloadPath": path})
                except Exception:
                    pass


def _promote_staging_xls_to_target(
    staging_dir: str,
    final_dir: str,
    current_date: str,
    old_staging_files: set,
) -> bool:
    """把临时目录里新下载的 xls 移动到用户配置目录。"""
    target = os.path.join(final_dir, f"Products-{current_date}.xls")
    os.makedirs(final_dir, exist_ok=True)

    staged_exact = os.path.join(staging_dir, f"Products-{current_date}.xls")
    if os.path.isfile(staged_exact):
        if os.path.exists(target):
            os.remove(target)
        shutil.move(staged_exact, target)
        return True

    try:
        names = sorted(os.listdir(staging_dir), key=lambda n: os.path.getmtime(os.path.join(staging_dir, n)), reverse=True)
    except Exception:
        names = []
    for name in names:
        if name in old_staging_files:
            continue
        if not name.lower().endswith(".xls") or name.endswith(".crdownload"):
            continue
        src = os.path.join(staging_dir, name)
        if not os.path.isfile(src) or os.path.getsize(src) < 256:
            continue
        if os.path.exists(target):
            os.remove(target)
        shutil.move(src, target)
        return True
    return os.path.isfile(target)


def _is_probable_export_bytes(data: bytes) -> bool:
    if not data or len(data) < 256:
        return False
    head = data[:400].lower()
    if b"<html" in head or b"<!doctype" in head:
        return False
    if data[:4] == b"\xd0\xcf\x11\xe0":  # .xls OLE
        return True
    if data[:2] == b"PK":  # .xlsx
        return True
    return len(data) >= 1024


def _is_export_network_response(mime: str, url: str, status: int = 200) -> bool:
    if status < 200 or status >= 400:
        return False
    mime_l = (mime or "").lower()
    url_l = (url or "").lower()
    if any(
        x in mime_l
        for x in (
            "excel",
            "spreadsheet",
            "ms-excel",
            "octet-stream",
            "x-msdownload",
            "vnd.ms",
            "force-download",
        )
    ):
        return True
    return url_l.endswith(".xls") or "export" in url_l or "/download" in url_l


def _write_export_bytes(dest_path: str, data: bytes) -> bool:
    if not _is_probable_export_bytes(data):
        return False
    tmp = dest_path + ".part"
    with open(tmp, "wb") as f:
        f.write(data)
    if os.path.exists(dest_path):
        os.remove(dest_path)
    os.replace(tmp, dest_path)
    return os.path.isfile(dest_path)


def _js_probe_export_link(driver) -> Optional[str]:
    """在页面上下文中解析导出链接。"""
    try:
        return driver.execute_script(
            """
            const a = document.querySelector('a.product-effective-download');
            if (!a) return null;
            const href = (a.href || '').trim();
            if (href && !href.startsWith('javascript') && href !== '#') return href;
            const dataHref = (a.getAttribute('data-href') || a.getAttribute('data-url') || '').trim();
            if (dataHref && dataHref.startsWith('http')) return dataHref;
            return null;
            """
        )
    except Exception:
        return None


def _export_via_inpage_fetch(driver, dest_path: str) -> bool:
    """在已登录页面内用 fetch 拉取导出文件，不经过 Chrome 下载管理器。"""
    url = _js_probe_export_link(driver)
    if not url:
        return False
    logger.info(f"页面内 fetch 导出: {url[:120]}")
    try:
        driver.set_script_timeout(120)
        b64_data = driver.execute_async_script(
            """
            const callback = arguments[arguments.length - 1];
            const url = arguments[0];
            fetch(url, { credentials: 'include', redirect: 'follow' })
              .then(async (resp) => {
                if (!resp.ok) throw new Error('status ' + resp.status);
                const ct = (resp.headers.get('content-type') || '').toLowerCase();
                const buf = await resp.arrayBuffer();
                if (buf.byteLength < 256) throw new Error('too small');
                if (ct.includes('text/html') && buf.byteLength < 80000) throw new Error('html');
                const bytes = new Uint8Array(buf);
                let binary = '';
                const step = 0x8000;
                for (let i = 0; i < bytes.length; i += step) {
                  binary += String.fromCharCode.apply(null, bytes.subarray(i, i + step));
                }
                callback(btoa(binary));
              })
              .catch(() => callback(null));
            """,
            url,
        )
        if not b64_data:
            return False
        data = base64.b64decode(b64_data)
        if _write_export_bytes(dest_path, data):
            logger.info(f"页面 fetch 保存成功: {dest_path}")
            return True
    except Exception as e:
        logger.warning(f"页面内 fetch 失败: {e}")
    return False


def _export_via_cdp_network_body(driver, wait, dest_path: str, task: TaskInfo) -> bool:
    """点击导出后通过 CDP Network.getResponseBody 直接读取响应体。"""
    from selenium.webdriver.common.by import By
    from selenium.webdriver.support import expected_conditions as EC

    pending: dict[str, str] = {}
    saved = {"ok": False}
    lock = threading.Lock()

    def on_response(event: dict) -> None:
        try:
            params = event.get("params") or {}
            resp = params.get("response") or {}
            rid = params.get("requestId")
            if not rid:
                return
            if _is_export_network_response(
                resp.get("mimeType") or "",
                resp.get("url") or "",
                int(resp.get("status") or 0),
            ):
                with lock:
                    pending[rid] = resp.get("url") or rid
        except Exception:
            pass

    def on_loading_finished(event: dict) -> None:
        if saved["ok"]:
            return
        try:
            rid = (event.get("params") or {}).get("requestId")
            if not rid:
                return
            with lock:
                if rid not in pending:
                    return
            body = driver.execute_cdp_cmd("Network.getResponseBody", {"requestId": rid})
            raw = body.get("body") or ""
            data = base64.b64decode(raw) if body.get("base64Encoded") else raw.encode("latin-1", "ignore")
            if _write_export_bytes(dest_path, data):
                logger.info(f"CDP 网络体保存成功: {dest_path}")
                saved["ok"] = True
            with lock:
                pending.pop(rid, None)
        except Exception:
            with lock:
                pending.pop(rid, None)

    try:
        driver.add_cdp_listener("Network.responseReceived", on_response)
        driver.add_cdp_listener("Network.loadingFinished", on_loading_finished)
    except Exception as e:
        logger.warning(f"无法注册 CDP 监听: {e}")
        return False

    try:
        driver.execute_cdp_cmd("Network.enable", {})
    except Exception:
        pass

    try:
        export_btn = wait.until(
            EC.element_to_be_clickable((By.CSS_SELECTOR, "a.product-effective-download"))
        )
        driver.execute_script("arguments[0].scrollIntoView({block:'center'});", export_btn)
        time.sleep(0.12)
        export_btn.click()
    except Exception as e:
        logger.warning(f"点击导出失败: {e}")
        return False

    deadline = time.time() + 75
    while time.time() < deadline:
        if task.should_stop():
            return False
        if saved["ok"] or (os.path.isfile(dest_path) and os.path.getsize(dest_path) > 256):
            return True
        if not _sleep_with_stop(task, 0.35):
            return False
    return saved["ok"] or (os.path.isfile(dest_path) and os.path.getsize(dest_path) > 256)


def _resolve_export_download_url(export_el) -> Optional[str]:
    """从导出按钮解析真实下载 URL（若有）。"""
    for attr in ("href", "data-href", "data-url", "data-download-url"):
        val = (export_el.get_attribute(attr) or "").strip()
        if val.startswith("http"):
            return val
        if val.startswith("//"):
            return "https:" + val
    onclick = export_el.get_attribute("onclick") or ""
    match = re.search(r"https?://[^'\"\\s)]+", onclick)
    if match:
        return match.group(0)
    return None


def _http_download_with_browser_session(driver, url: str, dest_path: str) -> bool:
    """用当前浏览器 Cookie 直连下载，绕过 Chrome 下载安全拦截。"""
    headers = {
        "User-Agent": driver.execute_script("return navigator.userAgent;"),
        "Referer": driver.current_url or "https://data.alibaba.com/product/overview",
        "Origin": "https://data.alibaba.com",
        "Accept": "application/vnd.ms-excel,application/octet-stream,*/*",
    }
    session = requests.Session()
    for cookie in driver.get_cookies():
        domain = cookie.get("domain") or ""
        session.cookies.set(
            cookie["name"],
            cookie["value"],
            domain=domain,
            path=cookie.get("path") or "/",
        )
    with session.get(url, headers=headers, stream=True, timeout=120, allow_redirects=True) as resp:
        resp.raise_for_status()
        content_type = (resp.headers.get("content-type") or "").lower()
        if "text/html" in content_type and int(resp.headers.get("content-length", 0) or 0) < 2048:
            return False
        tmp_path = dest_path + ".part"
        with open(tmp_path, "wb") as f:
            for chunk in resp.iter_content(chunk_size=65536):
                if chunk:
                    f.write(chunk)
        if not os.path.isfile(tmp_path) or os.path.getsize(tmp_path) < 128:
            try:
                os.remove(tmp_path)
            except Exception:
                pass
            return False
        if os.path.exists(dest_path):
            os.remove(dest_path)
        os.replace(tmp_path, dest_path)
    return os.path.isfile(dest_path)


def _capture_xls_url_from_performance(driver) -> Optional[str]:
    """从 Chrome performance 日志中抓取最近一次 xls/export 请求 URL。"""
    try:
        logs = driver.get_log("performance")
    except Exception:
        return None
    candidates: list[str] = []
    for entry in reversed(logs[-200:]):
        try:
            msg = json.loads(entry.get("message", "{}")).get("message", {})
        except Exception:
            continue
        method = msg.get("method")
        params = msg.get("params") or {}
        url = ""
        mime = ""
        if method == "Network.responseReceived":
            resp = params.get("response") or {}
            url = resp.get("url") or ""
            mime = (resp.get("mimeType") or "").lower()
        elif method == "Network.requestWillBeSent":
            url = (params.get("request") or {}).get("url") or ""
        if not url.startswith("http"):
            continue
        lower = url.lower()
        if (
            lower.endswith(".xls")
            or "export" in lower
            or "download" in lower
            or "excel" in mime
            or "spreadsheet" in mime
            or "ms-excel" in mime
        ):
            candidates.append(url)
    return candidates[0] if candidates else None


def _try_http_export_urls(driver, urls: list[str], dest_path: str) -> bool:
    seen: set[str] = set()
    for url in urls:
        if not url or url in seen:
            continue
        seen.add(url)
        try:
            logger.info(f"HTTP 拉取导出: {url[:140]}")
            if _http_download_with_browser_session(driver, url, dest_path):
                with open(dest_path, "rb") as f:
                    if _is_probable_export_bytes(f.read(4096)):
                        return True
                try:
                    os.remove(dest_path)
                except Exception:
                    pass
        except Exception as e:
            logger.warning(f"HTTP 拉取失败: {url[:80]} | {e}")
    return False


def _daily_export_products_file(
    driver,
    wait,
    download_dir: str,
    current_date: str,
    old_files: set,
    task: TaskInfo,
) -> bool:
    """
    导出单日 Products 文件：只点击一次导出。
    Chrome 先落到英文临时目录，再移动到用户目录；并抓取 downloadWillBegin URL 做 HTTP 回退。
    """
    from selenium.webdriver.common.by import By
    from selenium.webdriver.support import expected_conditions as EC

    target = os.path.join(download_dir, f"Products-{current_date}.xls")
    if os.path.isfile(target):
        return True

    staging_dir = _get_daily_staging_download_dir()
    old_staging = set(os.listdir(staging_dir))
    captured_urls: list[str] = []
    url_lock = threading.Lock()
    cdp_saved = {"ok": False}
    pending: dict[str, str] = {}
    pending_lock = threading.Lock()

    def on_download_begin(event: dict) -> None:
        url = (event.get("params") or {}).get("url") or ""
        if url.startswith("http"):
            with url_lock:
                captured_urls.append(url)
            logger.info(f"捕获下载 URL: {url[:140]}")

    def on_response(event: dict) -> None:
        try:
            params = event.get("params") or {}
            resp = params.get("response") or {}
            rid = params.get("requestId")
            if not rid:
                return
            if _is_export_network_response(
                resp.get("mimeType") or "",
                resp.get("url") or "",
                int(resp.get("status") or 0),
            ):
                with pending_lock:
                    pending[rid] = resp.get("url") or rid
        except Exception:
            pass

    def on_loading_finished(event: dict) -> None:
        if cdp_saved["ok"]:
            return
        try:
            rid = (event.get("params") or {}).get("requestId")
            if not rid:
                return
            with pending_lock:
                if rid not in pending:
                    return
            body = driver.execute_cdp_cmd("Network.getResponseBody", {"requestId": rid})
            raw = body.get("body") or ""
            data = base64.b64decode(raw) if body.get("base64Encoded") else raw.encode("latin-1", "ignore")
            if _write_export_bytes(target, data):
                cdp_saved["ok"] = True
                logger.info(f"CDP 响应体保存成功: {target}")
            with pending_lock:
                pending.pop(rid, None)
        except Exception:
            with pending_lock:
                pending.pop(rid, None)

    for evt, handler in (
        ("Browser.downloadWillBegin", on_download_begin),
        ("Network.responseReceived", on_response),
        ("Network.loadingFinished", on_loading_finished),
    ):
        try:
            driver.add_cdp_listener(evt, handler)
        except Exception as e:
            logger.warning(f"CDP 监听注册失败 {evt}: {e}")

    _set_browser_download_dir(driver, staging_dir)
    logger.info(f"Chrome 下载暂存目录: {staging_dir}")

    try:
        driver.execute_cdp_cmd("Network.enable", {})
    except Exception:
        pass

    # 有直链时先 HTTP，避免触发 Chrome 下载栏
    direct_url = _js_probe_export_link(driver)
    if direct_url and _try_http_export_urls(driver, [direct_url], target):
        return True
    if _export_via_inpage_fetch(driver, target):
        return True

    try:
        export_btn = wait.until(
            EC.element_to_be_clickable((By.CSS_SELECTOR, "a.product-effective-download"))
        )
        driver.execute_script("arguments[0].scrollIntoView({block:'center'});", export_btn)
        time.sleep(0.15)
        export_btn.click()
        logger.info(f"已点击导出: Products-{current_date}.xls")
    except Exception as e:
        logger.warning(f"点击导出失败: {e}")
        return False

    deadline = time.time() + 70
    while time.time() < deadline:
        if task.should_stop():
            return False
        if cdp_saved["ok"] or os.path.isfile(target):
            return True
        if _promote_staging_xls_to_target(staging_dir, download_dir, current_date, old_staging):
            logger.info(f"暂存目录落盘成功: {target}")
            return True
        if not _sleep_with_stop(task, 0.5):
            return False

    with url_lock:
        urls = list(captured_urls)
    if _try_http_export_urls(driver, urls, target):
        return True

    perf_url = _capture_xls_url_from_performance(driver)
    if perf_url and _try_http_export_urls(driver, [perf_url], target):
        return True

    try:
        export_btn = driver.find_element(By.CSS_SELECTOR, "a.product-effective-download")
        extra = _resolve_export_download_url(export_btn)
        if extra and _try_http_export_urls(driver, [extra], target):
            return True
    except Exception:
        pass

    logger.warning(
        f"导出失败 {current_date}：Chrome 下载栏若显示「出了点问题」，"
        f"请检查暂存目录 {staging_dir} 是否有 xls"
    )
    return os.path.isfile(target)


def _sleep_with_stop(task: Optional[TaskInfo], seconds: float, step: float = 0.25) -> bool:
    """分段 sleep，便于响应停止信号。返回 False 表示收到停止请求。"""
    end = time.time() + max(0.0, seconds)
    while time.time() < end:
        if task and task.should_stop():
            return False
        time.sleep(min(step, max(0.05, end - time.time())))
    return True


def _download_daily_data(task: TaskInfo, cfg, date_range: Optional[str]):
    """
    产品参谋 - 日数据下载
    来源脚本: cs_产品数据下载_正式发布.py
    """
    task.current_step = "启动浏览器..."
    logger.info("日数据下载 - 启动浏览器")

    daily_download_dir = _normalize_download_root(cfg.data_download.daily_output_dir)
    weekly_download_dir = _normalize_download_root(getattr(cfg.data_download, "weekly_output_dir", "") or daily_download_dir)
    os.makedirs(daily_download_dir, exist_ok=True)
    os.makedirs(weekly_download_dir, exist_ok=True)

    login_url = cfg.data_download.login_url or (
        "https://login.alibaba.com/newlogin/icbuLogin.htm"
        "?return_url=https%3A%2F%2Fdata.alibaba.com%2F"
    )
    target_url = "https://data.alibaba.com/product/overview"

    browser = BrowserManager()
    if not browser.setup():
        raise Exception("浏览器启动失败")

    driver = browser.driver

    from selenium.webdriver.common.by import By
    from selenium.webdriver.support.ui import WebDriverWait
    from selenium.webdriver.support import expected_conditions as EC
    from selenium.webdriver.common.action_chains import ActionChains

    wait = WebDriverWait(driver, 20)

    try:
        driver.execute_cdp_cmd("Page.setDownloadBehavior", {
            "behavior": "allow",
            "downloadPath": weekly_download_dir,
        })
        logger.info(f"下载目录已设置(周): {weekly_download_dir}")
    except Exception as e:
        logger.warning(f"设置下载目录失败: {e}")

    def _wait_for_file(filepath: str, timeout: int = 30) -> bool:
        start = time.time()
        while time.time() - start < timeout:
            if os.path.exists(filepath):
                return True
            time.sleep(0.5)
        return False

    def _wait_for_new_products_file(target_dir: str, old_files: set, timeout: int = 60) -> Optional[str]:
        """等待导出产生新文件，返回新文件名"""
        start = time.time()
        while time.time() - start < timeout:
            current_files = set(os.listdir(target_dir))
            new_files = [f for f in (current_files - old_files) if f.lower().endswith('.xls')]
            if new_files:
                products_files = [f for f in new_files if re.match(r"^Products-\d{4}-\d{2}-\d{2}(?: \(\d+\))?\.xls$", f)]
                return (products_files[0] if products_files else new_files[0])
            time.sleep(1)
        return None

    def _extract_date_from_label() -> Optional[str]:
        """兼容不同页面结构，尽量提取当前日期（优先取范围右侧日期）"""
        candidates = []
        try:
            el = driver.find_element(By.CSS_SELECTOR, "input[placeholder='请选择日期']")
            candidates.append((el.get_attribute("value") or "").strip())
        except Exception:
            pass

        selectors = [
            ".next-input-label span",
            ".next-select-trigger",
            ".datepicker-container",
            ".next-input-wrapper",
        ]
        for sel in selectors:
            try:
                for el in driver.find_elements(By.CSS_SELECTOR, sel):
                    txt = (el.text or "").strip()
                    title = (el.get_attribute("title") or "").strip()
                    if txt:
                        candidates.append(txt)
                    if title:
                        candidates.append(title)
            except Exception:
                pass

        for text in candidates:
            if not text:
                continue
            dates = re.findall(r"\d{4}-\d{2}-\d{2}", text)
            if dates:
                return dates[-1]
        return None

    try:
        task.current_step = "加载 Cookie 登录..."
        driver.get("https://www.alibaba.com")
        time.sleep(0.8)
        browser._load_cookies()
        driver.refresh()
        time.sleep(0.8)

        driver.get(target_url)
        try:
            wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, ".product-datepicker")))
            logger.info("已通过 Cookie 登录")
        except Exception:
            logger.warning("Cookie 未登录，等待手动登录")
            driver.get(login_url)
            end_time = time.time() + 300
            while time.time() < end_time:
                if "data.alibaba.com" in (driver.current_url or ""):
                    break
                time.sleep(0.3)
            browser._save_cookies()
            driver.get(target_url)
            wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, ".product-datepicker")))

        try:
            driver.execute_cdp_cmd("Page.setDownloadBehavior", {
                "behavior": "allow",
                "downloadPath": os.path.abspath(weekly_download_dir)
            })
        except Exception as e:
            logger.warning(f"设置下载目录失败: {e}")

        existing_weekly_files = set(os.listdir(weekly_download_dir))
        existing_daily_files = set(os.listdir(daily_download_dir))
        downloaded_weekly_dates = set()
        downloaded_daily_dates = set()

        def is_weekly_downloaded(date_str: str) -> bool:
            raw_name = f"Products-{date_str}.xls"
            week_name = _week_name_from_end_date(date_str)
            return (
                raw_name in existing_weekly_files
                or (week_name in existing_weekly_files if week_name else False)
                or date_str in downloaded_weekly_dates
            )

        def is_daily_downloaded(date_str: str) -> bool:
            raw_name = f"Products-{date_str}.xls"
            return raw_name in existing_daily_files or date_str in downloaded_daily_dates

        def _week_name_from_end_date(date_str: str) -> Optional[str]:
            try:
                end_dt = datetime.strptime(date_str, "%Y-%m-%d")
                start_dt = end_dt - timedelta(days=6)
                return f"{start_dt.strftime('%y%m%d')}-{end_dt.strftime('%y%m%d')}.xls"
            except Exception:
                return None

        def _rename_weekly_file_if_needed(src_name: str) -> tuple[str, bool]:
            """仅用于周数据阶段：Products-YYYY-MM-DD(.xls 或 (n).xls) -> YYMMDD-YYMMDD.xls
            返回: (最终文件名, 是否命中已存在周文件)
            """
            m = re.match(r"^Products-(\d{4}-\d{2}-\d{2})(?: \(\d+\))?\.xls$", src_name)
            if not m:
                return src_name, False

            date_str = m.group(1)
            src_path = os.path.join(weekly_download_dir, src_name)
            if not os.path.exists(src_path):
                return src_name, False

            week_name = _week_name_from_end_date(date_str)
            if not week_name:
                return src_name, False

            dst_path = os.path.join(weekly_download_dir, week_name)
            if os.path.abspath(src_path) == os.path.abspath(dst_path):
                return os.path.basename(dst_path), False

            if os.path.exists(dst_path):
                # 目标周文件已存在时，删除本次重复的 Products 文件，避免目录里堆积未重命名文件
                try:
                    os.remove(src_path)
                    logger.info(f"周文件已存在，删除重复文件: {src_name}")
                except Exception as e:
                    logger.warning(f"周文件已存在但删除重复失败: {src_name} | {e}")
                return os.path.basename(dst_path), True

            try:
                os.rename(src_path, dst_path)
                logger.info(f"周文件重命名: {src_name} -> {week_name}")
                return os.path.basename(dst_path), False
            except Exception as e:
                logger.warning(f"周文件重命名失败: {src_name} -> {week_name} | {e}")
                return src_name, False

        def get_current_date() -> Optional[str]:
            current_date = None
            try:
                date_input = driver.find_element(By.CSS_SELECTOR, "input[placeholder='请选择日期']")
                current_date = (date_input.get_attribute("value") or "").strip()
            except Exception:
                pass
            if not current_date or not re.match(r"\d{4}-\d{2}-\d{2}", current_date):
                current_date = _extract_date_from_label()
            return current_date

        def click_prev_until_changed(prev_date: Optional[str]) -> str:
            """返回: moved / end / fail"""
            for attempt in range(6):
                try:
                    prev_btn = None
                    has_prev_control = False
                    selectors = [
                        ".product-datepicker .datepicker-arrow button.next-prev",
                        ".datepicker-container .datepicker-arrow button.next-prev",
                        ".product-datepicker button.next-pagination-item.next-prev",
                        "button.next-prev",
                        "button.next-btn.next-prev",
                        "button.next-pagination-item.next-prev",
                        "button[aria-label='Prev']",
                        "button[aria-label='Previous']",
                    ]
                    for selector in selectors:
                        elems = driver.find_elements(By.CSS_SELECTOR, selector)
                        if elems:
                            has_prev_control = True
                        for elem in elems:
                            cls = (elem.get_attribute("class") or "").lower()
                            disabled_attr = str(elem.get_attribute("disabled") or "").lower()
                            if "disabled" in cls or disabled_attr in {"true", "disabled"}:
                                continue
                            prev_btn = elem
                            break
                        if prev_btn:
                            break

                    if not prev_btn:
                        xpath_candidates = [
                            "//button[contains(@class,'next-prev') and not(@disabled)]",
                            "//button[contains(@class,'next-pagination-item') and contains(@class,'next-prev') and not(@disabled)]",
                            "//button[@aria-label='Prev' and not(@disabled)]",
                            "//button[@aria-label='Previous' and not(@disabled)]",
                        ]
                        for xp in xpath_candidates:
                            elems = driver.find_elements(By.XPATH, xp)
                            if elems:
                                prev_btn = elems[0]
                                break

                    if not prev_btn:
                        if has_prev_control:
                            return "end"
                        if attempt < 5:
                            time.sleep(0.8)
                            continue
                        return "fail"

                    try:
                        driver.execute_script("arguments[0].scrollIntoView({block:'center'});", prev_btn)
                    except Exception:
                        pass

                    try:
                        prev_btn.click()
                    except Exception:
                        driver.execute_script("arguments[0].click();", prev_btn)

                    changed = False
                    if prev_date:
                        try:
                            wait.until(lambda d: (
                                (d.find_element(By.CSS_SELECTOR, "input[placeholder='请选择日期']").get_attribute("value") or "").strip() != prev_date
                            ))
                            changed = True
                        except Exception:
                            time.sleep(1.2)
                            now_date = get_current_date()
                            changed = bool(now_date and now_date != prev_date)
                    else:
                        time.sleep(0.8)
                        changed = True

                    if not changed:
                        return "fail"

                    time.sleep(0.4)
                    return "moved"
                except Exception:
                    if attempt < 5:
                        time.sleep(0.8)
                        continue
            return "fail"

        # 阶段1：先下载周数据（不切到“日”，不依赖日期控件）
        task.current_step = "下载周数据..."
        logger.info("日数据任务阶段1：先下载周数据（不切日期粒度）")
        while True:
            if task.should_stop():
                break
            task.wait_if_paused()

            try:
                old_files = set(os.listdir(weekly_download_dir))
                export_btn = wait.until(EC.element_to_be_clickable((By.CSS_SELECTOR, "a.product-effective-download")))
                export_btn.click()
                logger.info("开始下载周数据当前页")
            except Exception as e:
                logger.warning(f"周数据导出失败: {e}，停止周数据下载")
                break

            new_file = _wait_for_new_products_file(weekly_download_dir, old_files, 60)
            if not new_file:
                logger.info("周数据未检测到新文件（到顶/已下载），结束周数据并切换到日数据流程")
                break

            # 周数据文件重命名（支持 Products-YYYY-MM-DD.xls 与 Products-YYYY-MM-DD (n).xls）
            final_name, hit_existing_week = _rename_weekly_file_if_needed(new_file)
            m = re.match(r"^Products-(\d{4}-\d{2}-\d{2})(?: \(\d+\))?\.xls$", new_file)
            if m:
                downloaded_weekly_dates.add(m.group(1))
            existing_weekly_files = set(os.listdir(weekly_download_dir))
            logger.info(f"周数据下载完成: {final_name}")

            if hit_existing_week:
                logger.info("周数据已存在，切换到日数据流程")
                break

            move_result = click_prev_until_changed(None)
            if move_result == "end":
                logger.info("周数据：上一页已不可点击，周数据下载结束")
                break
            if move_result != "moved":
                logger.warning("周数据：翻页失败，结束周数据下载")
                break

        # 阶段2：切换到「日」并下载（对齐 cs_产品数据下载_正式发布.py）
        task.current_step = "切换到日粒度并下载日数据..."
        logger.info("日数据任务阶段2：刷新页面并切换到日粒度")

        driver.get(target_url)
        wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, ".product-datepicker")))
        time.sleep(1)

        try:
            driver.execute_cdp_cmd("Page.setDownloadBehavior", {
                "behavior": "allow",
                "downloadPath": os.path.abspath(daily_download_dir),
            })
            logger.info(f"下载目录已设置(日): {daily_download_dir}")
        except Exception as e:
            logger.warning(f"设置日数据下载目录失败: {e}")

        day_mode_ok = False
        try:
            logger.info("1. 点击日期选择框...")
            date_trigger = wait.until(
                EC.element_to_be_clickable(
                    (By.CSS_SELECTOR, ".datepicker-container .next-select-trigger")
                )
            )
            date_trigger.click()
            time.sleep(1)

            logger.info("2. 选择周期为「日」...")
            day_option = wait.until(
                EC.element_to_be_clickable(
                    (By.XPATH, "//ul[@class='static-menu']/li[span[text()='日']]")
                )
            )
            day_option.click()
            time.sleep(1)

            logger.info("3. 点击最新可选日期...")
            wait.until(EC.presence_of_element_located((By.CLASS_NAME, "next-date-picker")))
            date_cells = wait.until(
                EC.presence_of_all_elements_located(
                    (
                        By.XPATH,
                        "//td[not(contains(@class, 'next-disabled'))]"
                        "//div[@class='next-calendar-date']",
                    )
                )
            )
            if not date_cells:
                raise Exception("未找到任何可选日期格子")

            for elem in reversed(date_cells):
                date_text = (elem.text or "").strip()
                if date_text.isdigit():
                    elem.click()
                    logger.info(f"选中日期: {date_text}")
                    day_mode_ok = True
                    break
            if not day_mode_ok:
                raise Exception("未找到有效的可选日期")

            blank = driver.find_element(By.CSS_SELECTOR, "body")
            ActionChains(driver).move_to_element(blank).click().perform()
            time.sleep(2)

            if not get_current_date():
                raise Exception("切换日粒度后无法读取当前日期")
            logger.info(f"已切换到日粒度，当前日期: {get_current_date()}")
        except Exception as e:
            logger.error(f"切换到日粒度失败: {e}")
            try:
                driver.save_screenshot(
                    os.path.join(daily_download_dir, "error_select_date.png")
                )
            except Exception:
                pass

        if day_mode_ok:
            logger.info("开始日数据循环下载")

            while True:
                if task.should_stop():
                    break
                task.wait_if_paused()

                current_date = get_current_date()
                if not current_date:
                    logger.warning("无法从任何地方获取有效日期，结束日数据下载")
                    break

                logger.info(f"当前页面日期: {current_date}")
                task.current_step = f"下载日数据 {current_date}..."

                if is_daily_downloaded(current_date):
                    logger.info(f"日数据已存在 {current_date}，停止下载")
                    break

                try:
                    export_btn = wait.until(
                        EC.element_to_be_clickable(
                            (By.CSS_SELECTOR, "a.product-effective-download")
                        )
                    )
                    export_btn.click()
                    logger.info(f"→ 下载 Products-{current_date}.xls")

                    target = os.path.join(
                        daily_download_dir, f"Products-{current_date}.xls"
                    )
                    if _wait_for_file(target, 60):
                        existing_daily_files.add(f"Products-{current_date}.xls")
                        downloaded_daily_dates.add(current_date)
                        logger.info("下载完成")
                    else:
                        logger.warning(f"下载超时: {current_date}")
                except Exception as e:
                    logger.warning(f"日数据导出失败: {e}")

                move_result = click_prev_until_changed(current_date)
                if move_result == "end":
                    logger.info("日数据：日期控件上一页已禁用，下载结束")
                    break
                if move_result != "moved":
                    logger.warning("日数据：无法切换到前一天，结束日数据下载")
                    break
                time.sleep(3)

        logger.info(
            f"周数据新增 {len(downloaded_weekly_dates)} 个，"
            f"日数据新增 {len(downloaded_daily_dates)} 个"
        )

    finally:
        try:
            driver.quit()
        except Exception:
            pass
        logger.info("日数据下载结束")


def _store_parse_period_start(period_label: str) -> Optional[str]:
    try:
        clean_label = str(period_label).strip().replace(" (PST)", "").replace("（PST）", "")
        if "~" in clean_label:
            start_str = clean_label.split("~")[0].strip()
        else:
            start_str = clean_label.split()[0].strip()
        if re.match(r"^\d{4}-\d{2}$", start_str):
            dt = datetime.strptime(f"{start_str}-01", "%Y-%m-%d")
        else:
            dt = datetime.strptime(start_str, "%Y-%m-%d")
        return dt.strftime("%Y%m%d")
    except Exception:
        return None


def _store_parse_period_range(period_label: str):
    try:
        clean = str(period_label).strip().replace(" (PST)", "").replace("（PST）", "")
        if "~" in clean:
            left, right = [x.strip() for x in clean.split("~", 1)]
            start_dt = datetime.strptime(left, "%Y-%m-%d")
            end_dt = datetime.strptime(right, "%Y-%m-%d")
        else:
            d = datetime.strptime(clean.split()[0].strip(), "%Y-%m-%d")
            start_dt = d
            end_dt = d
        return start_dt, end_dt
    except Exception:
        return None, None


def _store_to_week_range_label(period_label: str) -> Optional[str]:
    start_dt, end_dt = _store_parse_period_range(period_label)
    ref = end_dt or start_dt
    if not ref:
        return None
    sunday = ref - pd.Timedelta(days=(ref.weekday() + 1) % 7)
    saturday = sunday + pd.Timedelta(days=6)
    return f"{sunday.strftime('%y%m%d')}-{saturday.strftime('%y%m%d')}"


def _store_aggregate_weekly(df: pd.DataFrame) -> pd.DataFrame:
    if df is None or df.empty:
        return pd.DataFrame(columns=["指标名称"])

    metric_col = str(df.columns[0])
    period_cols = [c for c in df.columns[1:]]

    week_cols: Dict[str, List[str]] = {}
    week_end_map: Dict[str, str] = {}
    week_coverage: Dict[str, set] = {}

    for c in period_cols:
        raw = str(c)
        wk = _store_to_week_range_label(raw)
        start_dt, end_dt = _store_parse_period_range(raw)
        if not wk or not start_dt or not end_dt:
            continue

        week_cols.setdefault(wk, []).append(c)
        week_end_map[wk] = wk.split("-")[1]
        week_coverage.setdefault(wk, set())

        ref = end_dt
        sunday = ref - pd.Timedelta(days=(ref.weekday() + 1) % 7)
        saturday = sunday + pd.Timedelta(days=6)

        d = max(start_dt, sunday)
        end_in_week = min(end_dt, saturday)
        while d <= end_in_week:
            week_coverage[wk].add(d.strftime("%Y-%m-%d"))
            d = d + pd.Timedelta(days=1)

    if not week_cols:
        out = df[[metric_col]].copy()
        out.columns = ["指标名称"]
        return out

    complete_weeks = [wk for wk in week_cols.keys() if len(week_coverage.get(wk, set())) >= 7]
    if not complete_weeks:
        out = df[[metric_col]].copy()
        out.columns = ["指标名称"]
        return out

    out = pd.DataFrame()
    out["指标名称"] = df[metric_col]
    for wk in complete_weeks:
        cols = week_cols[wk]
        nums = df[cols].apply(pd.to_numeric, errors="coerce").fillna(0)
        out[wk] = nums.sum(axis=1)

    ordered_weeks = sorted(list(complete_weeks), key=lambda w: week_end_map.get(w, "000000"), reverse=True)
    out = out[["指标名称"] + ordered_weeks]
    return out


def _store_build_summary_sheet(ws, my_weekly: pd.DataFrame, ind_weekly: pd.DataFrame, peer_weekly: pd.DataFrame):
    ws.merge_cells("A1:A2")
    ws["A1"] = "指标名称"
    ws.merge_cells("B1:E1")
    ws["B1"] = "我的数据"
    ws.merge_cells("F1:I1")
    ws["F1"] = "行业平均"
    ws.merge_cells("J1:M1")
    ws["J1"] = "同行优秀"

    ws["B2"] = "我的最新"
    ws["C2"] = "我的上周"
    ws["D2"] = "我的差值"
    ws["E2"] = "我的环比"
    ws["F2"] = "行业最新"
    ws["G2"] = "行业上周"
    ws["H2"] = "行业差值"
    ws["I2"] = "行业环比"
    ws["J2"] = "同行最新"
    ws["K2"] = "同行上周"
    ws["L2"] = "同行差值"
    ws["M2"] = "同行环比"

    for c in ["A1", "B1", "F1", "J1", "B2", "C2", "D2", "E2", "F2", "G2", "H2", "I2", "J2", "K2", "L2", "M2"]:
        ws[c].font = Font(bold=True)
        ws[c].alignment = Alignment(horizontal="center", vertical="center")

    thin = Side(style="thin", color="000000")

    def _set_outer_border(r1: int, c1: int, r2: int, c2: int):
        for r in range(r1, r2 + 1):
            for c in range(c1, c2 + 1):
                cell = ws.cell(row=r, column=c)
                left = thin if c == c1 else cell.border.left
                right = thin if c == c2 else cell.border.right
                top = thin if r == r1 else cell.border.top
                bottom = thin if r == r2 else cell.border.bottom
                cell.border = Border(left=left, right=right, top=top, bottom=bottom)

    _set_outer_border(1, 1, 2, 1)
    _set_outer_border(1, 2, 2, 5)
    _set_outer_border(1, 6, 2, 9)
    _set_outer_border(1, 10, 2, 13)

    my_map = my_weekly.set_index("指标名称") if not my_weekly.empty else pd.DataFrame()
    ind_map = ind_weekly.set_index("指标名称") if not ind_weekly.empty else pd.DataFrame()
    peer_map = peer_weekly.set_index("指标名称") if not peer_weekly.empty else pd.DataFrame()

    indicators = []
    for dfx in [my_weekly, ind_weekly, peer_weekly]:
        if not dfx.empty:
            indicators.extend([str(x) for x in dfx["指标名称"].tolist()])
    indicators = list(dict.fromkeys(indicators))

    def _latest_prev(dfm: pd.DataFrame, idx_name: str):
        if dfm is None or dfm.empty or idx_name not in dfm.index:
            return 0.0, 0.0
        cols = [c for c in dfm.columns if c != "指标名称"]
        if not cols:
            return 0.0, 0.0
        latest = pd.to_numeric(dfm.loc[idx_name, cols[0]], errors="coerce")
        prev = pd.to_numeric(dfm.loc[idx_name, cols[1]], errors="coerce") if len(cols) > 1 else 0
        return float(latest if pd.notna(latest) else 0), float(prev if pd.notna(prev) else 0)

    row_idx = 3
    for name in indicators:
        ws.cell(row=row_idx, column=1, value=name)

        my_latest, my_prev = _latest_prev(my_map, name)
        ind_latest, ind_prev = _latest_prev(ind_map, name)
        peer_latest, peer_prev = _latest_prev(peer_map, name)

        vals = [
            my_latest, my_prev, my_latest - my_prev, ((my_latest - my_prev) / my_prev if my_prev else 0),
            ind_latest, ind_prev, ind_latest - ind_prev, ((ind_latest - ind_prev) / ind_prev if ind_prev else 0),
            peer_latest, peer_prev, peer_latest - peer_prev, ((peer_latest - peer_prev) / peer_prev if peer_prev else 0),
        ]
        for j, v in enumerate(vals, start=2):
            ws.cell(row=row_idx, column=j, value=v)
        row_idx += 1

    for col in [5, 9, 13]:
        for r in range(3, row_idx):
            ws.cell(row=r, column=col).number_format = "0.00%"

    for col_letter in ["A", "B", "C", "D", "E", "F", "G", "H", "I", "J", "K", "L", "M"]:
        ws.column_dimensions[col_letter].width = 14


def _store_save_weekly_summary_workbook(
    scfg,
    base_path: str,
    my_df: pd.DataFrame,
    ind_df: pd.DataFrame,
    peer_df: pd.DataFrame,
):
    weekly_my = _store_aggregate_weekly(my_df)
    weekly_ind = _store_aggregate_weekly(ind_df)
    weekly_peer = _store_aggregate_weekly(peer_df)

    out_path = (scfg.summary_output_path or "").strip()
    if not out_path:
        out_path = os.path.join(os.path.dirname(base_path), "运营数据_周汇总.xlsx")

    os.makedirs(os.path.dirname(out_path), exist_ok=True)

    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        weekly_my.to_excel(writer, sheet_name="合并数据", index=False)
        weekly_ind.to_excel(writer, sheet_name="行业平均", index=False)
        weekly_peer.to_excel(writer, sheet_name="同行优秀", index=False)

    wb = load_workbook(out_path)
    if "总结" in wb.sheetnames:
        del wb["总结"]
    ws = wb.create_sheet("总结")
    _store_build_summary_sheet(ws, weekly_my, weekly_ind, weekly_peer)
    wb.save(out_path)
    logger.info(f"店铺周汇总已生成: {out_path}")


def _store_period_save_path(base_path: str, period: str) -> str:
    """日数据用主路径；周/月写入带后缀的独立文件。"""
    base = (base_path or "").strip()
    if not base or period == "day":
        return base
    root, ext = os.path.splitext(base)
    if not ext:
        ext = ".xlsx"
    suffix = {"week": "_周", "month": "_月"}.get(period, f"_{period}")
    return f"{root}{suffix}{ext}"


def _store_get_latest_date_from_excel(file_path: str) -> Optional[str]:
    try:
        if not os.path.exists(file_path):
            return None
        df = pd.read_excel(file_path, sheet_name="合并数据")
        if df.shape[1] <= 1:
            return None
        valid_dates = []
        for col in df.columns[1:]:
            d = _store_parse_period_start(str(col))
            if d:
                valid_dates.append(d)
        return max(valid_dates) if valid_dates else None
    except Exception:
        return None


def _store_merge_dataframes(new_data: Dict[str, Dict], existing_df: pd.DataFrame) -> pd.DataFrame:
    if not new_data:
        return existing_df if not existing_df.empty else pd.DataFrame(columns=["指标名称"])

    sorted_items = sorted(
        new_data.items(),
        key=lambda x: _store_parse_period_start(x[0]) or "00000000",
        reverse=True,
    )

    new_df = pd.DataFrame()
    for period_label, metrics in sorted_items:
        ser = pd.Series(metrics, name=period_label)
        new_df = pd.concat([new_df, ser], axis=1)

    new_df = new_df.reset_index()
    new_df.columns = ["指标名称"] + list(new_df.columns[1:])

    if existing_df.empty:
        return new_df

    existing_indexed = existing_df.set_index(existing_df.columns[0])
    existing_indexed.index.name = "指标名称"
    new_indexed = new_df.set_index("指标名称")

    merged_indexed = pd.concat([new_indexed, existing_indexed], axis=1)
    merged = merged_indexed.reset_index()

    cols = merged.columns.tolist()
    cols.insert(0, cols.pop(cols.index("指标名称")))
    return merged[cols]


def _store_order_metrics_df(df: pd.DataFrame) -> pd.DataFrame:
    if df is None or df.empty:
        return pd.DataFrame(columns=["指标名称"])
    metric_col = str(df.columns[0])
    order_map = {name: idx for idx, name in enumerate(SHOP_TRENDS_METRIC_ORDER)}
    out = df.copy()
    out["_sort"] = out[metric_col].map(lambda x: order_map.get(str(x).strip(), 9999))
    out = out.sort_values("_sort", kind="stable").drop(columns=["_sort"])
    if metric_col != "指标名称":
        out = out.rename(columns={metric_col: "指标名称"})
    return out


def _shop_trends_to_float(val) -> float:
    try:
        if val is None or (isinstance(val, float) and pd.isna(val)):
            return 0.0
        return float(val)
    except Exception:
        return 0.0


def _shop_trends_metrics_from_item(item: Dict) -> tuple:
    def _build(getter):
        metrics: Dict[str, float] = {}
        for api_key, cn_name in SHOP_TRENDS_METRIC_BINDINGS:
            metrics[cn_name] = _shop_trends_to_float(getter(api_key))
            alias = SHOP_TRENDS_P4P_ALIAS.get(api_key)
            if alias:
                metrics[alias] = metrics[cn_name]
        for name in SHOP_TRENDS_ZERO_METRICS:
            metrics[name] = 0.0
        shop_uv = metrics.get("店铺访问人数") or 0.0
        bus_byr = metrics.get("商机人数") or 0.0
        metrics["商机转化率"] = (bus_byr / shop_uv) if shop_uv > 0 else 0.0
        return metrics

    current = _build(lambda k: item.get(k))
    industry = _build(lambda k: item.get(f"{k}RivalAvg"))
    peer = _build(lambda k: item.get(f"{k}RivalGood"))
    return current, industry, peer


def _shop_trends_batch_start_keys(batch: List[Dict], statistics_type: str) -> List[str]:
    keys: List[str] = []
    for item in batch:
        label = _shop_trends_period_label(item, statistics_type)
        key = _store_parse_period_start(label)
        if key:
            keys.append(key)
    return keys


def _shop_trends_period_label(item: Dict, statistics_type: str) -> str:
    raw_range = str(item.get("statDateRange") or "").strip()
    stat_date = str(item.get("statDate") or "").strip()
    if statistics_type == "week" and raw_range:
        label = raw_range
    elif statistics_type == "month":
        label = raw_range or (stat_date[:7] if stat_date else "")
    else:
        label = stat_date or raw_range
    if label and "(PST)" not in label and "（PST）" not in label:
        label = f"{label} (PST)"
    return label


def _build_shop_trends_url(
    *,
    statistics_type: str,
    selected: int,
    ctoken: str,
    cate_id: str,
) -> str:
    params = {
        "action": "OneAction",
        "iName": "vip/home/custom/getShopTrends",
        "statisticsType": statistics_type,
        "selected": str(selected),
        "terminalType": "total",
        "isMyselfUpgraded": "true",
        "cateId": cate_id,
        "statisticType": "os",
        "region": "os",
        "seperateByCate": "false",
        "isVip": "true",
        "ctoken": ctoken,
    }
    query = "&".join(f"{k}={quote(str(v))}" for k, v in params.items())
    return f"{SHOP_TRENDS_API_BASE}?{query}"


def _fetch_shop_trends_batch(
    session: requests.Session,
    *,
    statistics_type: str,
    selected: int,
    ctoken: str,
    cate_id: str,
) -> List[Dict]:
    url = _build_shop_trends_url(
        statistics_type=statistics_type,
        selected=selected,
        ctoken=ctoken,
        cate_id=cate_id,
    )
    headers = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36",
        "Referer": SHOP_TRENDS_REFERER,
        "Accept": "application/json",
    }
    resp = session.get(url, headers=headers, timeout=60, verify=False)
    resp.raise_for_status()
    payload = resp.json()
    if not isinstance(payload, dict):
        raise Exception(f"店铺趋势接口返回异常: statisticsType={statistics_type} selected={selected}")
    code = payload.get("code")
    if code not in (None, 0, 200, "0", "200") and payload.get("data") is None:
        raise Exception(f"店铺趋势接口失败: statisticsType={statistics_type} selected={selected} code={code}")
    data = payload.get("data") or {}
    items = data.get("returnValue") if isinstance(data, dict) else None
    if not isinstance(items, list):
        return []
    return [x for x in items if isinstance(x, dict)]


def _download_store_overview_period(
    task: TaskInfo,
    session: requests.Session,
    *,
    period: str,
    file_path: str,
    end_date: str,
    ctoken: str,
    cate_id: str,
) -> int:
    """拉取单个粒度（日/周/月）并写入对应 Excel，返回新增周期数；-1 表示已是最新无新增。"""
    start_selected = SHOP_TRENDS_START_SELECTED.get(period, 1)
    batch_hint = {"day": "约30天", "week": "约10周", "month": "约6个月"}.get(period, "")
    period_label = {"day": "日", "week": "周", "month": "月"}.get(period, period)

    if task.should_stop():
        return 0
    task.wait_if_paused()
    task.current_step = f"店铺趋势-{period_label}（selected={start_selected}，{batch_hint}）"

    batch = _fetch_shop_trends_batch(
        session,
        statistics_type=period,
        selected=start_selected,
        ctoken=ctoken,
        cate_id=cate_id,
    )
    if not batch:
        raise Exception(f"店铺{period_label}数据接口未返回数据，请检查 Cookie 或类目 cateId")

    period_end = _store_get_latest_date_from_excel(file_path) or end_date
    all_current: Dict[str, Dict] = {}
    all_industry: Dict[str, Dict] = {}
    all_peer: Dict[str, Dict] = {}
    seen_periods: set = set()

    start_keys = _shop_trends_batch_start_keys(batch, period)
    range_tip = ""
    if start_keys:
        range_tip = f" | 批次约 {min(start_keys)}~{max(start_keys)}"

    for item in batch:
        label = _shop_trends_period_label(item, period)
        if not label or label in seen_periods:
            continue
        start_key = _store_parse_period_start(label)
        if start_key and start_key <= period_end:
            continue

        cur, ind, peer = _shop_trends_metrics_from_item(item)
        if not cur:
            continue

        seen_periods.add(label)
        all_current[label] = cur
        all_industry[label] = ind
        all_peer[label] = peer

    logger.info(
        f"店铺趋势-{period_label} | selected={start_selected} | 本批={len(batch)} | 新增={len(all_current)}{range_tip}"
    )

    if not all_current:
        if os.path.exists(file_path):
            return -1
        raise Exception(f"未采集到店铺{period_label}数据，请检查 Cookie 或类目 cateId")

    os.makedirs(os.path.dirname(file_path) or ".", exist_ok=True)
    try:
        existing_current = pd.read_excel(file_path, sheet_name="合并数据")
        existing_industry = pd.read_excel(file_path, sheet_name="行业平均")
        existing_peer = pd.read_excel(file_path, sheet_name="同行优秀")
    except Exception:
        existing_current = pd.DataFrame(columns=["指标名称"])
        existing_industry = pd.DataFrame(columns=["指标名称"])
        existing_peer = pd.DataFrame(columns=["指标名称"])

    merged_current = _store_order_metrics_df(_store_merge_dataframes(all_current, existing_current))
    merged_industry = _store_order_metrics_df(_store_merge_dataframes(all_industry, existing_industry))
    merged_peer = _store_order_metrics_df(_store_merge_dataframes(all_peer, existing_peer))

    with pd.ExcelWriter(file_path, engine="openpyxl") as writer:
        merged_current.to_excel(writer, sheet_name="合并数据", index=False)
        merged_industry.to_excel(writer, sheet_name="行业平均", index=False)
        merged_peer.to_excel(writer, sheet_name="同行优秀", index=False)

    return len(all_current)


def _download_store_overview(task: TaskInfo, cfg, period_type: str):
    """数据概览 - 一次任务下载日/周/月全部店铺运营数据。"""
    task.current_step = "店铺数据 - 准备接口请求..."
    logger.info("店铺数据下载 - getShopTrends 日/周/月全量")

    scfg = cfg.store_overview
    save_path = _normalize_daily_dir(scfg.save_path)
    cate_id = str(getattr(scfg, "cate_id", None) or SHOP_TRENDS_DEFAULT_CATE_ID).strip()
    default_end = str(scfg.default_end_date or "20000101").strip()

    os.makedirs(os.path.dirname(save_path) or ".", exist_ok=True)

    cookies_file = _normalize_download_root(getattr(scfg, "cookie_file", "") or "")
    cookies = _load_cookies_dict_from_file(cookies_file) if cookies_file else {}
    if not cookies:
        cookies_file = _normalize_download_root(getattr(cfg.paths, "cookie_file", "") or "")
        cookies = _load_cookies_dict_from_file(cookies_file)
    if not cookies:
        raise Exception(
            f"未找到有效 Cookie，请先在配置管理中登录并保存 Cookie: {cookies_file or '(未配置)'}"
        )
    ctoken = (cookies.get("ctoken") or "").strip()
    if not ctoken and cookies.get("_m_h5_tk"):
        ctoken = str(cookies.get("_m_h5_tk", "")).split("_")[0]

    session = requests.Session()
    session.cookies.update(cookies)

    stats: List[str] = []
    any_written = False
    day_path = _store_period_save_path(save_path, "day")

    for period in ("day", "week", "month"):
        if task.should_stop():
            break
        file_path = _store_period_save_path(save_path, period)
        try:
            added = _download_store_overview_period(
                task,
                session,
                period=period,
                file_path=file_path,
                end_date=default_end,
                ctoken=ctoken,
                cate_id=cate_id,
            )
        except Exception as e:
            logger.error(f"店铺{period}数据下载失败: {e}")
            raise

        name = {"day": "日", "week": "周", "month": "月"}[period]
        if added > 0:
            any_written = True
            stats.append(f"{name}+{added}")
        elif added == -1:
            stats.append(f"{name}已最新")
        time.sleep(0.25)

    if not any_written and not os.path.exists(day_path):
        raise Exception("未采集到任何店铺运营数据，请检查 Cookie 或类目 cateId")

    try:
        day_current = pd.read_excel(day_path, sheet_name="合并数据")
        day_industry = pd.read_excel(day_path, sheet_name="行业平均")
        day_peer = pd.read_excel(day_path, sheet_name="同行优秀")
        _store_save_weekly_summary_workbook(scfg, day_path, day_current, day_industry, day_peer)
    except Exception as e:
        logger.warning(f"周汇总生成跳过（需先有日数据）: {e}")

    _STORE_OVERVIEW_CACHE.clear()
    summary = "、".join(stats) if stats else "完成"
    task.current_step = f"店铺数据下载完成（{summary}）"
    logger.info(f"店铺数据下载完成 | 日={day_path} | 周={_store_period_save_path(save_path, 'week')} | 月={_store_period_save_path(save_path, 'month')} | {summary}")


TRAFFIC_CHANNEL_API_BASE = "https://mydata.alibaba.com/self/.json"
TRAFFIC_CHANNEL_REFERER = "https://data.alibaba.com/traffic/source"
TRAFFIC_CHANNEL_FIELD_MAP = {
    "channelType": "流量渠道",
    "statDate": "日期",
    "detailUv": "店铺访问人数",
    "tmUv": "店内TM咨询人数",
    "fbUv": "店内询盘人数",
}
# Excel 只存日粒度；周/月接口 statDate 为周期汇总日（如月初），与日数据同日会覆盖/重复累加
TRAFFIC_CHANNEL_FETCH_PLANS = [
    ("day", "日数据", 0, 120),
]
TRAFFIC_CHANNEL_OVERVIEW_API = {
    "week": ("week", 1),
    "month": ("month", 1),
}


def _normalize_traffic_channel_date_key(value) -> str:
    text = str(value).strip()
    if not text:
        return ""
    if re.fullmatch(r"\d{6}", text):
        try:
            return datetime.strptime(text, "%y%m%d").strftime("%y%m%d")
        except Exception:
            return ""
    if re.fullmatch(r"\d{8}", text):
        try:
            return datetime.strptime(text, "%Y%m%d").strftime("%y%m%d")
        except Exception:
            return ""
    try:
        dt = pd.to_datetime(text, errors="coerce")
        if pd.isna(dt):
            return ""
        return dt.strftime("%y%m%d")
    except Exception:
        return ""


def _build_traffic_channel_summary_url(statistics_type: str, selected: int, ctoken: str) -> str:
    params = {
        "action": "OneAction",
        "iName": "vip/channel/summary",
        "isVip": "true",
        "statisticsType": statistics_type,
        "selected": str(selected),
        "terminalType": "total",
        "statisticType": "os",
        "region": "os",
        "hideNoEffectItem": "true",
        "ctoken": ctoken,
    }
    query = "&".join(f"{k}={quote(str(v))}" for k, v in params.items())
    return f"{TRAFFIC_CHANNEL_API_BASE}?{query}"


def _fetch_traffic_channel_summary_page(
    session: requests.Session,
    *,
    statistics_type: str,
    selected: int,
    ctoken: str,
) -> List[Dict]:
    url = _build_traffic_channel_summary_url(statistics_type, selected, ctoken)
    headers = {
        "User-Agent": (
            "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
            "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"
        ),
        "Referer": TRAFFIC_CHANNEL_REFERER,
        "Accept": "application/json, text/plain, */*",
    }
    resp = session.get(url, headers=headers, timeout=(8, 30))
    resp.raise_for_status()
    try:
        json_data = resp.json()
    except json.JSONDecodeError as e:
        preview = (resp.text or "")[:240].replace("\n", " ")
        raise Exception(f"流量渠道接口 JSON 解析失败: type={statistics_type} selected={selected} | {preview}") from e

    data_list = json_data.get("data", []) if isinstance(json_data, dict) else []
    if not isinstance(data_list, list):
        return []
    if not data_list:
        return []

    first_item = data_list[0] if isinstance(data_list[0], dict) else {}
    if not any(k in first_item for k in ["channelType", "statDate", "detailUv"]):
        return []
    return [x for x in data_list if isinstance(x, dict)]


def _traffic_channel_summary_to_rows(items: List[Dict]) -> List[Dict]:
    rows: List[Dict] = []
    for item in items:
        channel = str(item.get("channelType") or "").strip()
        if not channel:
            continue
        uv = float(pd.to_numeric(item.get("detailUv"), errors="coerce") or 0)
        ask = float(pd.to_numeric(item.get("fbUv"), errors="coerce") or 0)
        tm = float(pd.to_numeric(item.get("tmUv"), errors="coerce") or 0)
        rows.append(
            {
                "流量渠道": channel,
                "店铺访问人数": uv,
                "店内询盘人数": ask,
                "店内TM咨询人数": tm,
                "商机转化率": (ask + tm) / uv if uv else 0,
            }
        )
    return rows


def _fetch_traffic_channel_overview_rows(statistics_type: str, selected: int) -> List[Dict]:
    """与阿里后台周/月 Tab 一致：直接读 summary 接口，不用 Excel 日列汇总。"""
    cfg = get_config()
    cookies_file = _normalize_download_root(getattr(cfg.paths, "cookie_file", "") or "")
    cookies = _load_cookies_dict_from_file(cookies_file)
    if not cookies:
        return []

    ctoken = (cookies.get("ctoken") or "").strip()
    if not ctoken and cookies.get("_m_h5_tk"):
        ctoken = str(cookies.get("_m_h5_tk", "")).split("_")[0]

    session = requests.Session()
    session.cookies.update(cookies)
    try:
        items = _fetch_traffic_channel_summary_page(
            session,
            statistics_type=statistics_type,
            selected=selected,
            ctoken=ctoken,
        )
        return _traffic_channel_summary_to_rows(items)
    except Exception as e:
        logger.warning(f"流量渠道概览接口读取失败: type={statistics_type} selected={selected} | {e}")
        return []


def _download_traffic_channel(task: TaskInfo, cfg):
    """
    店铺流量渠道下载（接口直连，无需打开浏览器）
    日/周/月：vip/channel/summary
    """
    task.current_step = "流量渠道下载 - 准备接口请求..."
    logger.info("流量渠道下载 - 接口直连模式")

    dcfg = cfg.data_download
    OUTPUT_FILE = _normalize_download_root(getattr(dcfg, "traffic_channel_output_file", "") or "")
    COOKIES_FILE = _normalize_download_root(getattr(cfg.paths, "cookie_file", "") or "")
    logger.info(f"流量渠道输出文件: {OUTPUT_FILE}")
    logger.info(f"流量渠道Cookie文件: {COOKIES_FILE}")

    FIELD_MAP = TRAFFIC_CHANNEL_FIELD_MAP

    if not OUTPUT_FILE:
        raise Exception("未配置流量渠道输出文件")

    if os.path.dirname(OUTPUT_FILE):
        os.makedirs(os.path.dirname(OUTPUT_FILE), exist_ok=True)

    cookies = _load_cookies_dict_from_file(COOKIES_FILE)
    if not cookies:
        raise Exception(
            f"未找到有效 Cookie，请先在配置管理中登录并保存 Cookie: {COOKIES_FILE or '(未配置)'}"
        )
    ctoken = (cookies.get("ctoken") or "").strip()
    if not ctoken and cookies.get("_m_h5_tk"):
        ctoken = str(cookies.get("_m_h5_tk", "")).split("_")[0]

    session = requests.Session()
    session.cookies.update(cookies)

    def _get_existing_latest_date() -> str:
        if not OUTPUT_FILE or not os.path.exists(OUTPUT_FILE):
            return ""
        try:
            old_book = pd.read_excel(OUTPUT_FILE, sheet_name=None)
        except Exception:
            return ""

        latest_key = ""
        for sheet_df in old_book.values():
            if sheet_df is None or sheet_df.empty:
                continue
            for col in sheet_df.columns:
                col_key = _normalize_traffic_channel_date_key(col)
                if col_key and col_key > latest_key:
                    latest_key = col_key
        return latest_key

    existing_stop_date = _get_existing_latest_date()
    if existing_stop_date:
        logger.info(f"流量渠道历史最新日期: {existing_stop_date}，日数据抓到该日期时停止")

    day_data: List[Dict] = []
    seen_day_channel: set = set()
    stop_when_repeat_date = False
    stop_reason = ""

    def _ingest_items(items: List[Dict], *, statistics_type: str, selected: int) -> int:
        nonlocal stop_when_repeat_date, stop_reason
        new_count = 0
        for item in items:
            d = item.get("statDate")
            c = item.get("channelType")
            if not d or not c:
                continue

            date_key = _normalize_traffic_channel_date_key(d)
            if statistics_type == "day" and existing_stop_date and date_key == existing_stop_date:
                stop_when_repeat_date = True
                stop_reason = date_key
                logger.info(f"流量渠道命中历史最新日期 {date_key}，停止继续拉取日数据")
                continue

            key = (d, c)
            if key in seen_day_channel:
                continue
            seen_day_channel.add(key)
            day_data.append(item)
            new_count += 1

        if new_count > 0:
            logger.info(
                f"流量渠道接口 | {statistics_type} selected={selected} | 本次新增={new_count} | 累计={len(day_data)}"
            )
        return new_count

    for statistics_type, label, start_selected, max_pages in TRAFFIC_CHANNEL_FETCH_PLANS:
        if task.should_stop():
            break
        task.wait_if_paused()
        task.current_step = f"流量渠道接口抓取（{label}）"
        logger.info(f"流量渠道开始抓取: {label} | statisticsType={statistics_type}")

        for selected in range(start_selected, start_selected + max_pages):
            if task.should_stop():
                break
            task.wait_if_paused()
            task.current_step = f"流量渠道接口抓取（{label} selected={selected}）"

            try:
                items = _fetch_traffic_channel_summary_page(
                    session,
                    statistics_type=statistics_type,
                    selected=selected,
                    ctoken=ctoken,
                )
            except Exception as e:
                if selected == start_selected:
                    logger.warning(f"流量渠道 {label} 首页请求失败，跳过: {e}")
                else:
                    logger.warning(f"流量渠道 {label} selected={selected} 请求失败，停止该维度: {e}")
                break

            if not items:
                logger.info(f"流量渠道 {label} selected={selected} 无数据，已到最后一页")
                break

            _ingest_items(items, statistics_type=statistics_type, selected=selected)

            if statistics_type == "day" and stop_when_repeat_date:
                logger.info(f"流量渠道日数据已命中历史日期 {stop_reason}，停止日数据翻页")
                break

            time.sleep(0.25)

    def _process_and_save_traffic_channel():
        if not day_data:
            raise Exception("未采集到任何流量渠道数据，请检查 Cookie 是否有效")

        df_raw = pd.DataFrame(day_data).rename(columns=FIELD_MAP)
        required_cols = ['流量渠道', '日期', '店铺访问人数', '店内TM咨询人数', '店内询盘人数']
        for c in required_cols:
            if c not in df_raw.columns:
                raise Exception(f"流量渠道数据缺少字段: {c}")

        df_raw = df_raw.dropna(subset=['流量渠道', '日期']).copy()
        df_raw['日期'] = pd.to_datetime(df_raw['日期'], errors='coerce')
        df_raw = df_raw.dropna(subset=['日期'])

        for c in ['店铺访问人数', '店内TM咨询人数', '店内询盘人数']:
            df_raw[c] = pd.to_numeric(df_raw[c], errors='coerce').fillna(0)

        df_new = df_raw.groupby(['流量渠道', '日期'], as_index=False)[['店铺访问人数', '店内询盘人数', '店内TM咨询人数']].sum()

        def parse_sheet_to_day_df(sheet_df, channel_name):
            if sheet_df is None or sheet_df.empty:
                return pd.DataFrame(columns=['流量渠道', '日期', '店铺访问人数', '店内询盘人数', '店内TM咨询人数'])
            if '指标' not in sheet_df.columns:
                return pd.DataFrame(columns=['流量渠道', '日期', '店铺访问人数', '店内询盘人数', '店内TM咨询人数'])

            date_cols = [c for c in sheet_df.columns if isinstance(c, str) and re.fullmatch(r'\d{6}', c)]
            if not date_cols:
                return pd.DataFrame(columns=['流量渠道', '日期', '店铺访问人数', '店内询盘人数', '店内TM咨询人数'])

            need_metrics = ['店铺访问人数', '店内询盘人数', '店内TM咨询人数']
            metric_map = {}
            for m in need_metrics:
                row = sheet_df[sheet_df['指标'] == m]
                if row.empty:
                    continue
                metric_map[m] = row.iloc[0]

            if len(metric_map) < 3:
                return pd.DataFrame(columns=['流量渠道', '日期', '店铺访问人数', '店内询盘人数', '店内TM咨询人数'])

            rows = []
            for d in date_cols:
                dt = pd.to_datetime(d, format='%y%m%d', errors='coerce')
                if pd.isna(dt):
                    continue
                rows.append({
                    '流量渠道': channel_name,
                    '日期': dt,
                    '店铺访问人数': pd.to_numeric(metric_map['店铺访问人数'].get(d, 0), errors='coerce'),
                    '店内询盘人数': pd.to_numeric(metric_map['店内询盘人数'].get(d, 0), errors='coerce'),
                    '店内TM咨询人数': pd.to_numeric(metric_map['店内TM咨询人数'].get(d, 0), errors='coerce'),
                })

            if not rows:
                return pd.DataFrame(columns=['流量渠道', '日期', '店铺访问人数', '店内询盘人数', '店内TM咨询人数'])

            out = pd.DataFrame(rows)
            for c in ['店铺访问人数', '店内询盘人数', '店内TM咨询人数']:
                out[c] = pd.to_numeric(out[c], errors='coerce').fillna(0)
            return out

        historical_frames = []
        if os.path.exists(OUTPUT_FILE):
            try:
                old_book = pd.read_excel(OUTPUT_FILE, sheet_name=None)
                for sheet_name, sheet_df in old_book.items():
                    parsed = parse_sheet_to_day_df(sheet_df, sheet_name)
                    if not parsed.empty:
                        historical_frames.append(parsed)
            except Exception as e:
                logger.warning(f"读取历史流量渠道文件失败，将重建: {e}")

        if historical_frames:
            df_old = pd.concat(historical_frames, ignore_index=True)
        else:
            df_old = pd.DataFrame(columns=['流量渠道', '日期', '店铺访问人数', '店内询盘人数', '店内TM咨询人数'])

        if not df_old.empty:
            df_all = pd.concat([df_old, df_new], ignore_index=True)
            df_all = df_all.drop_duplicates(subset=['流量渠道', '日期'], keep='last')
        else:
            df_all = df_new.copy()

        df_all = df_all.groupby(['流量渠道', '日期'], as_index=False)[['店铺访问人数', '店内询盘人数', '店内TM咨询人数']].sum()
        df_all['商机转化率'] = ((df_all['店内询盘人数'] + df_all['店内TM咨询人数']) / df_all['店铺访问人数'].replace(0, pd.NA)).fillna(0)

        metric_order = ['店铺访问人数', '店内询盘人数', '店内TM咨询人数', '商机转化率']

        def make_sheet_name(name, used):
            invalid = ['\\', '/', '*', '?', ':', '[', ']']
            n = str(name)
            for ch in invalid:
                n = n.replace(ch, '_')
            n = n.strip() or '未命名渠道'
            n = n[:31]
            base = n
            idx = 1
            while n in used:
                suffix = f"_{idx}"
                n = (base[:31 - len(suffix)] + suffix) if len(base) + len(suffix) > 31 else (base + suffix)
                idx += 1
            used.add(n)
            return n

        used_sheet_names = set()
        written_sheet_count = 0

        with pd.ExcelWriter(OUTPUT_FILE, engine='openpyxl') as writer:
            for channel in sorted(df_all['流量渠道'].dropna().unique(), reverse=True):
                sub = df_all[df_all['流量渠道'] == channel].copy()
                if sub.empty:
                    continue

                date_cols = sorted(sub['日期'].dt.strftime('%y%m%d').unique(), reverse=True)
                row_dict = {'指标': metric_order}
                for d in date_cols:
                    row_dict[d] = []

                for metric in metric_order:
                    metric_series = sub[['日期', metric]].copy()
                    metric_series['日期'] = metric_series['日期'].dt.strftime('%y%m%d')
                    value_map = dict(zip(metric_series['日期'], metric_series[metric]))
                    for d in date_cols:
                        row_dict[d].append(value_map.get(d, 0))

                out_df = pd.DataFrame(row_dict)

                if len(date_cols) >= 2:
                    latest_col = date_cols[0]
                    second_col = date_cols[1]
                    out_df['异动'] = out_df[latest_col] - out_df[second_col]
                else:
                    out_df['异动'] = 0

                out_df = out_df[['指标', '异动'] + date_cols]
                numeric_cols = ['异动'] + date_cols
                for c in numeric_cols:
                    out_df[c] = pd.to_numeric(out_df[c], errors='coerce').fillna(0)

                sheet_name = make_sheet_name(channel, used_sheet_names)
                out_df.to_excel(writer, sheet_name=sheet_name, index=False)
                written_sheet_count += 1

                ws = writer.sheets[sheet_name]
                rate_rows = out_df.index[out_df['指标'] == '商机转化率'].tolist()
                if rate_rows:
                    excel_row = rate_rows[0] + 2
                    for excel_col in range(2, 3 + len(date_cols)):
                        ws.cell(row=excel_row, column=excel_col).number_format = '0.00%'

            if written_sheet_count == 0:
                pd.DataFrame([{'提示': '无可导出数据'}]).to_excel(writer, sheet_name='结果', index=False)

        logger.info(f"流量渠道报表生成成功: {os.path.abspath(OUTPUT_FILE)}")

    _process_and_save_traffic_channel()
    task.current_step = f"流量渠道下载完成，共 {len(day_data)} 条原始记录"
    logger.info(f"流量渠道下载完成 | 原始记录={len(day_data)} | 文件={OUTPUT_FILE}")


PRODUCT_OPERATE_API_URL = (
    "https://hz-productposting.alibaba.com/product_operate/layer/product_list.do"
)
PRODUCT_OPERATE_REFERER = (
    "https://hz-productposting.alibaba.com/product_operate/product_growth.htm"
)
# 与阿里后台接口一致：潜力优品 / 优品（见产品列表篇.txt）；爆品为旧版兼容，接口不可用则跳过
PRODUCT_OPERATE_STAGES = [
    ("POTENTIAL", "潜力优品"),
    ("PLATFORM_GOOD", "优品"),
    ("HOT", "爆品"),
]
# 阿里接口 size>10 时部分 stage 会返回空列表，与后台页面一致用 10
PRODUCT_OPERATE_PAGE_SIZE = 10
PRODUCT_OPERATE_MAX_PAGES = 500

PRODUCT_OPERATE_FIELD_MAPPING = {
    "productId": "产品ID",
    "searchImpsCnt30d": "近30天搜索曝光数",
    "detailUv30d": "近30天访问人数",
    "byrCnt90d": "近90天支付买家数",
    "fbTmUv90d": "近90天[TM+询盘]人数",
    "fbTmUv90dRate": "近90天[TM+询盘]转化",
    "rfndIssueRate90d": "近 90 天退款纠纷率",
    "taOrdCnt90d": "近 90 天订单数",
    "reviewCnt365d": "近 365 天评价数",
    "dabRate30d": "近 30 天商品问题率(纠纷/差评)",
    "isStarClusterDirectProd": "是否星团直供商品",
    "isAbilityEligibleProd": "是否具备服务能力商品",
    "isTrendProd": "是否趋势品(有流量倾斜)",
    "competitiveLevel": "竞争力等级",
    "bizProdFbCnt90dThreshold": "（商机优品要求）提升近90天[TM+询盘]人数到",
    "bizProdPbCnt90dThreshold": "（商机优品要求）提升近90天商品的支付买家数到",
    "bizProdFbUvRate90dThreshold": "（商机优品要求）提升近90天[TM+询盘]转化到",
    "goodProdDuvCnt30dThreshold": "提升近30天访客人数到",
}

PRODUCT_OPERATE_STAGE_MAP = {
    "POTENTIAL": "潜力优品",
    "PLATFORM_GOOD": "优品",
    "GOOD": "优品",
    "HOT": "爆品",
    "Potential product": "潜力优品",
    "Superior products": "优品",
    "Explosive": "爆品",
    "潜力优品": "潜力优品",
    "优品": "优品",
    "爆品": "爆品",
}

PRODUCT_OPERATE_TYPE_MAP = {
    "Traded goods": "交易品",
    "Business opportunity products": "商机品",
    "交易品": "交易品",
    "商机品": "商机品",
}


def _cookies_list_to_dict(cookie_items: Any) -> Dict[str, str]:
    result: Dict[str, str] = {}
    if isinstance(cookie_items, list):
        for item in cookie_items:
            if isinstance(item, dict) and item.get("name") is not None:
                result[str(item["name"])] = str(item.get("value", ""))
            elif isinstance(item, (tuple, list)) and len(item) >= 2:
                result[str(item[0])] = str(item[1])
    elif isinstance(cookie_items, dict):
        if isinstance(cookie_items.get("cookies"), list):
            return _cookies_list_to_dict(cookie_items.get("cookies"))
        for key, value in cookie_items.items():
            if key not in {"cookies", "origins"}:
                result[str(key)] = str(value)
    return result


def _load_cookies_dict_from_file(cookie_file: str) -> Dict[str, str]:
    if not cookie_file or not os.path.exists(cookie_file):
        return {}

    try:
        with open(cookie_file, "rb") as f:
            payload = pickle.load(f)
        parsed = _cookies_list_to_dict(payload)
        if parsed:
            return parsed
    except Exception:
        pass

    for json_path in [f"{cookie_file}.json", cookie_file]:
        try:
            if not os.path.exists(json_path):
                continue
            with open(json_path, "r", encoding="utf-8") as f:
                payload = json.load(f)
            if isinstance(payload, dict) and isinstance(payload.get("origins"), list):
                merged: Dict[str, str] = {}
                for origin in payload.get("origins") or []:
                    if isinstance(origin, dict) and isinstance(origin.get("cookies"), list):
                        merged.update(_cookies_list_to_dict(origin.get("cookies")))
                if merged:
                    return merged
            parsed = _cookies_list_to_dict(payload)
            if parsed:
                return parsed
        except Exception:
            continue
    return {}


def _parse_product_operate_jsonp(text: str) -> Optional[Dict]:
    if not text:
        return None
    text = text.strip()
    # 阿里常见：/**/jsonp_123456_1({...})
    if text.startswith("/**/"):
        text = text[4:].lstrip()
    if text.startswith("{"):
        try:
            return json.loads(text)
        except json.JSONDecodeError:
            return None
    jsonp_match = re.search(r"^[^(]+\((\{.*\})\)\s*;?\s*$", text, re.DOTALL)
    if jsonp_match:
        try:
            return json.loads(jsonp_match.group(1))
        except json.JSONDecodeError:
            pass
    start = text.find("{")
    end = text.rfind("}")
    if start >= 0 and end > start:
        try:
            return json.loads(text[start : end + 1])
        except json.JSONDecodeError:
            return None
    return None


def _product_operate_row_from_record(prod: Dict, tab_label: str) -> Dict:
    extracted = {
        v: prod.get(k) for k, v in PRODUCT_OPERATE_FIELD_MAPPING.items() if k in prod
    }
    kpi = prod.get("kpi", {}) if isinstance(prod.get("kpi", {}), dict) else {}
    for source_key, target_key in PRODUCT_OPERATE_FIELD_MAPPING.items():
        if source_key in kpi:
            extracted[target_key] = kpi.get(source_key)

    features = prod.get("features", [])
    p_types: List[str] = []
    if isinstance(features, list):
        p_types = [
            PRODUCT_OPERATE_TYPE_MAP.get((item or {}).get("name"), (item or {}).get("name"))
            for item in features
            if isinstance(item, dict) and item.get("code") in ["TRADABLE", "OPPORTUNITY"]
        ]

    raw_stage = prod.get("productStage")
    extracted["产品类型"] = ",".join(p_types) if p_types else None
    extracted["产品级别"] = PRODUCT_OPERATE_STAGE_MAP.get(raw_stage, raw_stage) or tab_label
    if "产品ID" in extracted:
        extracted["产品ID"] = str(extracted["产品ID"])
    return extracted


def _is_product_operate_last_page(payload: Dict, products: List, page: int, page_size: int) -> bool:
    if not products:
        return True

    for key in ("totalPage", "totalPages", "pageCount"):
        total_page = payload.get(key)
        if total_page is not None:
            try:
                if page >= int(total_page):
                    return True
            except (TypeError, ValueError):
                pass

    total = payload.get("total")
    if total is not None and page_size > 0:
        try:
            if page * page_size >= int(total):
                return True
        except (TypeError, ValueError):
            pass

    if len(products) < page_size:
        return True
    return False


def _build_product_operate_list_url(stage: str, page: int, page_size: int, ctoken: str, tb_token: str) -> str:
    params = {
        "subject": "",
        "productId": "",
        "ownerMemberId": "",
        "productType": "",
        "categoryId": "",
        "isCompetitiveAllProd": "",
        "groupId": "",
        "groupLevel": "",
        "notGoodProdReason": "",
        "competitiveLevel": "",
        "adStatus": "",
        "isRepeatProd": "false",
        "isOverseaStock": "false",
        "isFastCustom": "false",
        "is7DayDispatch": "false",
        "isRecommendUpGoodProd": "false",
        "isReadyDownProd": "false",
        "isRecentDownProd": "false",
        "isKeyOperateProd": "false",
        "isTodayNewProd": "false",
        "isRecommendUpPotentialProd": "false",
        "isRecentDownProdP0P1": "false",
        "isReadyDownProdP0P1": "false",
        "page": str(page),
        "size": str(page_size),
        "stage": stage,
        "sort": "DEFAULT_SORT",
        "ctoken": ctoken,
        "_tb_token_": tb_token,
        "callback": f"jsonp_{int(time.time() * 1000)}_{page}",
    }
    query = "&".join(f"{k}={quote(str(v))}" for k, v in params.items())
    return f"{PRODUCT_OPERATE_API_URL}?{query}"


def _save_product_operate_excel(rows: List[Dict], output_file: str) -> None:
    if not rows:
        return
    if not str(output_file).lower().endswith((".xlsx", ".xls")):
        raise Exception(f"产品运营输出文件后缀无效: {output_file}")
    df = pd.DataFrame(rows)
    if "产品ID" in df.columns:
        df["产品ID"] = df["产品ID"].astype(str)
    cols = list(df.columns)
    for col in ["产品ID", "产品类型", "产品级别"]:
        if col in cols:
            cols.remove(col)
    new_cols = ["产品ID", "产品类型", "产品级别"] + cols
    df = df[[c for c in new_cols if c in df.columns]]
    df.to_excel(output_file, index=False)
    logger.info(f"产品运营数据已写入: {output_file} | 条数={len(df)}")


def _fetch_product_operate_page(
    session: requests.Session,
    *,
    stage: str,
    page: int,
    page_size: int,
    ctoken: str,
    tb_token: str,
) -> Dict:
    url = _build_product_operate_list_url(stage, page, page_size, ctoken, tb_token)
    headers = {
        "User-Agent": (
            "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
            "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"
        ),
        "Referer": PRODUCT_OPERATE_REFERER,
        "Accept": "*/*",
    }
    resp = session.get(url, headers=headers, timeout=(8, 30))
    resp.raise_for_status()
    raw_text = resp.text or ""
    data = _parse_product_operate_jsonp(raw_text)
    if not data:
        preview = raw_text[:240].replace("\n", " ")
        raise Exception(
            f"产品运营接口响应解析失败: stage={stage} page={page} | preview={preview}"
        )
    code = data.get("code")
    if not (data.get("success") is True or code == 200 or str(code) == "200"):
        msg = data.get("message") or data.get("msg") or data
        raise Exception(f"产品运营接口返回失败: stage={stage} page={page} | {msg}")
    return data


def _download_product_operate(task: TaskInfo, cfg):
    """
    产品运营工作台数据下载（接口直连，无需打开浏览器）
    接口：product_operate/layer/product_list.do
    """
    task.current_step = "产品运营下载 - 准备接口请求..."
    logger.info("产品运营下载 - 接口直连模式")

    dcfg = cfg.data_download
    output_file = _normalize_download_root(getattr(dcfg, "product_operate_output_file", "") or "")
    cookies_file = _normalize_download_root(getattr(cfg.paths, "cookie_file", "") or "")

    if not output_file:
        raise Exception("未配置产品运营输出文件")
    if not str(output_file).lower().endswith((".xlsx", ".xls")):
        raise Exception(f"产品运营输出文件必须是 Excel 文件（.xlsx/.xls），当前配置: {output_file}")

    output_dir = os.path.dirname(output_file)
    if output_dir:
        os.makedirs(output_dir, exist_ok=True)

    cookies = _load_cookies_dict_from_file(cookies_file)
    if not cookies:
        raise Exception(
            f"未找到有效 Cookie，请先在配置管理中登录并保存 Cookie: {cookies_file or '(未配置)'}"
        )

    ctoken = (cookies.get("ctoken") or "").strip()
    if not ctoken and cookies.get("_m_h5_tk"):
        ctoken = str(cookies.get("_m_h5_tk", "")).split("_")[0]
    tb_token = (cookies.get("_tb_token_") or cookies.get("tb_token") or "").strip()
    if not tb_token:
        raise Exception("Cookie 中缺少 _tb_token_，请重新登录阿里后台后保存 Cookie")

    session = requests.Session()
    session.cookies.update(cookies)

    all_data: List[Dict] = []
    page_size = PRODUCT_OPERATE_PAGE_SIZE

    for stage_idx, (stage_code, tab_label) in enumerate(PRODUCT_OPERATE_STAGES, start=1):
        if task.should_stop():
            break
        task.wait_if_paused()
        task.current_step = f"产品运营接口抓取（{tab_label}）"
        logger.info(f"产品运营开始抓取({stage_idx}/{len(PRODUCT_OPERATE_STAGES)}): {tab_label} | stage={stage_code}")

        page_num = 1
        while page_num <= PRODUCT_OPERATE_MAX_PAGES:
            if task.should_stop():
                break
            task.wait_if_paused()
            task.current_step = f"产品运营接口抓取（{tab_label} 第{page_num}页）"

            try:
                data = _fetch_product_operate_page(
                    session,
                    stage=stage_code,
                    page=page_num,
                    page_size=page_size,
                    ctoken=ctoken,
                    tb_token=tb_token,
                )
            except Exception as e:
                if page_num == 1:
                    logger.warning(f"产品运营阶段跳过（接口不可用）: {tab_label} | {e}")
                    break
                raise
            payload = data.get("data") or {}
            products = payload.get("products") or payload.get("list") or []
            if not isinstance(products, list):
                products = []

            appended = 0
            for prod in products:
                if not isinstance(prod, dict):
                    continue
                all_data.append(_product_operate_row_from_record(prod, tab_label))
                appended += 1

            logger.info(
                f"产品运营接口页完成 | 标签={tab_label} | 第{page_num}页 | 本页={appended} | 累计={len(all_data)}"
            )
            if appended > 0:
                _save_product_operate_excel(all_data, output_file)

            if _is_product_operate_last_page(payload, products, page_num, page_size):
                logger.info(f"产品运营已到最后一页，停止: {tab_label} | 共{page_num}页")
                break

            page_num += 1
            time.sleep(0.35)

    if not all_data:
        raise Exception("产品运营接口未采集到任何数据，请检查 Cookie 是否有效或店铺是否有对应商品")
    _save_product_operate_excel(all_data, output_file)
    task.current_step = f"产品运营下载完成，共 {len(all_data)} 条"
    logger.info(f"产品运营下载完成 | 总条数={len(all_data)} | 文件={output_file}")


def _download_keywords(task: TaskInfo, cfg):
    """
    选词参谋 - 关键词接口下载并自动分析（下载+解析一步完成）
    接口: vip/traffic/keyword/getKeywords
    """
    task.current_step = "关键词 - 准备接口请求..."
    logger.info("关键词下载 - 接口直连并自动分析")

    kcfg = cfg.keyword_download
    download_dir = _normalize_download_root(kcfg.download_folder)
    output_dir = _normalize_download_root(kcfg.output_folder)
    os.makedirs(download_dir, exist_ok=True)
    os.makedirs(output_dir, exist_ok=True)

    cookies_file = _normalize_download_root(getattr(cfg.paths, "cookie_file", "") or "")
    cookies = _load_cookies_dict_from_file(cookies_file)
    if not cookies:
        raise Exception(
            f"未找到有效 Cookie，请先在配置管理中登录并保存 Cookie: {cookies_file or '(未配置)'}"
        )
    ctoken = (cookies.get("ctoken") or "").strip()
    if not ctoken and cookies.get("_m_h5_tk"):
        ctoken = str(cookies.get("_m_h5_tk", "")).split("_")[0]

    session = requests.Session()
    session.cookies.update(cookies)

    stop_date_short = _get_keyword_existing_stop_date(download_dir)
    if stop_date_short:
        logger.info(f"关键词下载历史最新日期: {stop_date_short}，遇到该周将停止增量抓取")

    all_data: Dict[str, pd.DataFrame] = {}
    seen_week_dates: set = set()
    start_selected = 1

    for selected in range(start_selected, start_selected + KEYWORD_API_MAX_SELECTED):
        if task.should_stop():
            break
        task.wait_if_paused()
        task.current_step = f"关键词接口抓取（周 selected={selected}）"

        week_rows: List[Dict] = []
        week_stat_date = ""
        total = 0
        page_no = 1
        max_pages = 1

        while page_no <= max(max_pages, 1):
            if task.should_stop():
                break
            task.wait_if_paused()
            task.current_step = f"关键词接口抓取（周 selected={selected} 第{page_no}页）"

            try:
                page_payload = _fetch_keyword_get_page(
                    session, selected=selected, page_no=page_no, ctoken=ctoken
                )
            except Exception as e:
                if page_no == 1:
                    logger.warning(f"关键词周 selected={selected} 请求失败，停止: {e}")
                else:
                    logger.warning(f"关键词 selected={selected} page={page_no} 失败，结束该周: {e}")
                break

            items = page_payload.get("items") or []
            if not items:
                if page_no == 1:
                    logger.info(f"关键词 selected={selected} 无数据，停止翻周")
                break

            total = int(page_payload.get("total") or total or len(items))
            max_pages = max(1, (total + KEYWORD_API_PAGE_SIZE - 1) // KEYWORD_API_PAGE_SIZE)
            week_rows.extend(items)

            if page_no >= max_pages or len(items) < KEYWORD_API_PAGE_SIZE:
                break
            page_no += 1
            time.sleep(0.2)

        if not week_rows:
            break

        stat_dates = sorted(
            {str(x.get("statDate") or "").strip() for x in week_rows if x.get("statDate")}
        )
        week_stat_date = stat_dates[0] if stat_dates else ""
        date_short = _keyword_stat_date_to_short(week_stat_date)
        if not date_short:
            logger.warning(f"关键词 selected={selected} 缺少 statDate，跳过")
            continue

        if date_short in seen_week_dates:
            logger.info(f"关键词周日期重复 {date_short}，停止继续翻周")
            break
        seen_week_dates.add(date_short)

        if stop_date_short and date_short <= stop_date_short:
            logger.info(f"关键词命中历史日期 {date_short}，停止增量抓取")
            break

        df = pd.DataFrame([_keyword_record_to_row(x) for x in week_rows])
        if "词" in df.columns:
            df = df.drop_duplicates(subset=["词"], keep="last")
        all_data[date_short] = df
        _save_keyword_week_excel(download_dir, week_stat_date, df)
        logger.info(
            f"关键词周数据完成 | selected={selected} | 日期={week_stat_date} | 条数={len(df)} | 累计周数={len(all_data)}"
        )
        time.sleep(0.25)

    if not all_data:
        raise Exception("未采集到任何关键词数据，请检查 Cookie 是否有效")

    result = _build_keyword_summary_and_anomaly(all_data, output_dir, task)
    task.current_step = (
        f"关键词下载并分析完成：{len(all_data)} 周 | "
        f"{os.path.basename(result.get('summary_file', ''))}"
    )
    logger.info(f"关键词流程完成 | 周数={len(all_data)} | summary={result.get('summary_file')}")


def list_data_files(dir_path: Optional[str] = None) -> List[Dict]:
    """列出已下载的数据文件"""
    cfg = get_config()
    save_dir = (dir_path or cfg.paths.download_save_dir or "").strip()
    files = []

    if not save_dir:
        return files

    # 去掉可能的引号
    if (save_dir.startswith("\"") and save_dir.endswith("\"")) or (save_dir.startswith("'") and save_dir.endswith("'")):
        save_dir = save_dir[1:-1].strip()

    # 兼容前端可能传来的 "D:..." 或 "/D:..."
    if re.match(r"^[\\/]+[A-Za-z]:", save_dir):
        save_dir = re.sub(r"^[\\/]+", "", save_dir)

    save_dir = os.path.normpath(save_dir)
    logger.debug(f"文件目录: {save_dir}")
    logger.debug(f"目录存在: {os.path.exists(save_dir)} | 是目录: {os.path.isdir(save_dir)}")

    if not os.path.isdir(save_dir):
        logger.warning("文件目录不存在")
        return files

    for root, _, filenames in os.walk(save_dir):
        for f in filenames:
            filepath = os.path.join(root, f)
            if os.path.isfile(filepath):
                stat = os.stat(filepath)
                rel = os.path.relpath(filepath, save_dir)
                files.append({
                    "name": rel,
                    "path": filepath,
                    "size_bytes": stat.st_size,
                    "modified": stat.st_mtime,
                })

    files.sort(key=lambda x: x.get("modified", 0), reverse=True)
    logger.debug(f"文件数量: {len(files)}")
    if files:
        logger.debug(f"文件样例: {files[0].get('name')}")
    return files


# ===================== 产品360解析辅助 =====================

def _normalize_product360_dir(path: str) -> str:
    if not path:
        return ""
    normalized = os.path.normpath(path)
    # 修正 "\D:" 这类路径
    if re.match(r"^\\[A-Za-z]:", normalized):
        normalized = normalized.lstrip("\\")
    return normalized


def _normalize_download_root(path: str) -> str:
    """统一归一化下载根目录，兼容打包版可能写出的错误前导斜杠路径。"""
    if not path:
        return ""
    normalized = os.path.normpath(path)
    if re.match(r"^\\[A-Za-z]:", normalized):
        normalized = normalized.lstrip("\\")
    return normalized


def _normalize_daily_dir(path: str) -> str:
    if not path:
        return ""
    normalized = os.path.normpath(path)
    if re.match(r"^\\[A-Za-z]:", normalized):
        normalized = normalized.lstrip("\\")
    return normalized


def _get_detail_field_map() -> Dict[str, str]:
    return {
        'statDate': '数据统计日期', 'prodId': '产品ID', 'actlLevelStar': '产品实际星级',
        'prodRateLevel': '平台内部产品分级', 'score': '平台产品综合得分', 'rankSeq': '产品综合搜索排名',
        'kwList': '产品核心搜索关键词', 'isSavingsSplotlight': '惠采购',
        'isPotenyouService': '是否开通平台潜力商家服务', 'sceneCnt': '产品关联场景数量',
        'payOrdByrCntFromd': '支付订单买家数', 'avgProdAbCnt30dRank': '近30天详情访问UV类目排名',
        'compAvgProdAbCnt30dCocRank': '近30天UV竞品对比排名', 'cateStage': '类目生命周期标签',
        'prodTagImps': '平台曝光标签', 'prodTagDuv': '平台访客流量标签',
        'prodTagImpsClk': '平台点击效率标签', 'prodTag': '平台转化效果标签',
        'sumProdShowNum': '产品总曝光量', 'clkImpsRate': '曝光点击率=总点击量/总曝光量',
        'uvDetail': '产品详情页独立访客UV', 'abUvRate': '访问转化率=支付买家数/详情UV',
        'uvFbMc': 'Facebook自然/帖文营销引流UV', 'uvFbAtm': 'Facebook广告投放引流UV',
        'sceneTopDuv': '场景顶部展位访问UV', 'sceneTopDAb': '场景顶部展位访问转化率',
        'sceneTopImpsClk': '场景顶部展位曝光点击率', 'sceneBottomImps': '场景底部展位曝光量',
        'sceneBottomImpsClk': '场景底部展位曝光点击率',
        'sceneBottomDAb': '场景底部展位访问转化率', 'sceneBottomDuv': '场景底部展位访问UV',
        'isNewProd': '是否为店铺维度新品', 'isNewArrival': '是否为NewArrival',
        'isPlatformNewProd': '是否为平台维度新品', 'isWending': '是否开通问鼎推广',
        'isIntlbw': '是否为跨境专供商品', 'channelType': '各流量渠道UV分布',
        'uvCountry': '买家实际访问来源国家'
    }


def _get_detail_headers_order() -> List[str]:
    return [
        '数据统计日期', '产品ID', '产品实际星级', '平台内部产品分级', '平台产品综合得分',
        '产品综合搜索排名', '产品核心搜索关键词', '惠采购',
        '是否开通平台潜力商家服务', '产品关联场景数量', '支付订单买家数',
        '近30天详情访问UV类目排名', '近30天UV竞品对比排名', '类目生命周期标签',
        '平台曝光标签', '平台访客流量标签', '平台点击效率标签', '平台转化效果标签',
        '产品总曝光量', '曝光点击率=总点击量/总曝光量', '产品详情页独立访客UV',
        '访问转化率=支付买家数/详情UV', 'Facebook自然/帖文营销引流UV',
        'Facebook广告投放引流UV', '场景顶部展位访问UV', '场景顶部展位访问转化率',
        '场景顶部展位曝光点击率', '场景底部展位曝光量', '场景底部展位曝光点击率',
        '场景底部展位访问转化率', '场景底部展位访问UV', '是否为店铺维度新品',
        '是否为NewArrival', '是否为平台维度新品', '是否开通问鼎推广', '是否为跨境专供商品',
        '各流量渠道UV分布', '买家实际访问来源国家'
    ]


def get_keyword_latest_anomaly(dir_path: Optional[str] = None) -> Dict[str, List[Dict]]:
    """获取关键词最新异动数据（按最新异动日）"""
    cfg = get_config()
    base_dir = (dir_path or cfg.keyword_download.output_folder or "").strip()
    if not base_dir:
        return {"exposure": [], "click": [], "index": []}

    if re.match(r"^[\\/]+[A-Za-z]:", base_dir):
        base_dir = re.sub(r"^[\\/]+", "", base_dir)
    base_dir = os.path.normpath(base_dir)

    if not os.path.isdir(base_dir):
        return {"exposure": [], "click": [], "index": []}

    files = [
        os.path.join(base_dir, f)
        for f in os.listdir(base_dir)
        if f.startswith("关键词异动分析_") and f.endswith(".xlsx")
    ]
    if not files:
        return {"exposure": [], "click": [], "index": []}

    latest_file = max(files, key=os.path.getmtime)

    def _to_num(v) -> float:
        try:
            if pd.isna(v):
                return 0.0
            return float(str(v).replace(",", "").strip())
        except Exception:
            return 0.0

    def _extract_by_df(df: pd.DataFrame) -> List[Dict]:
        if df is None or df.empty:
            return []
        cols = list(df.columns)
        if len(cols) < 2:
            return []

        latest_col = cols[1]  # 老脚本是日期倒序，最新列在第2列
        result = []
        for _, row in df.iterrows():
            kw = str(row.iloc[0]).strip() if pd.notna(row.iloc[0]) else ""
            if not kw:
                continue
            val = _to_num(row.get(latest_col, 0))
            if abs(val) > 0:
                result.append({"keyword": kw, "value": val})

        result.sort(key=lambda x: abs(x["value"]), reverse=True)
        return result[:3000]

    def _extract(sheet_name: str, sheet_idx: int) -> List[Dict]:
        try:
            df = pd.read_excel(latest_file, sheet_name=sheet_name)
            rows = _extract_by_df(df)
            if rows:
                return rows
        except Exception:
            pass

        # 兜底：按 sheet 下标读取，处理中文 sheet 名在部分环境下识别异常
        try:
            df = pd.read_excel(latest_file, sheet_name=sheet_idx)
            return _extract_by_df(df)
        except Exception:
            return []

    data = {
        "exposure": _extract("搜索曝光次数异动", 0) or _extract("曝光量异动", 0),
        "click": _extract("点击量异动", 1),
        "index": _extract("关键词指数异动", 2),
    }
    logger.info(f"关键词异动返回: exposure={len(data['exposure'])}, click={len(data['click'])}, index={len(data['index'])}")
    return data


def get_keyword_latest_summary(dir_path: Optional[str] = None) -> Dict[str, List[Dict]]:
    """获取关键词数据汇总中的最新曝光量/点击量"""
    cfg = get_config()
    base_dir = (dir_path or cfg.keyword_download.output_folder or "").strip()
    if not base_dir:
        return {"exposure": [], "click": []}

    if re.match(r"^[\\/]+[A-Za-z]:", base_dir):
        base_dir = re.sub(r"^[\\/]+", "", base_dir)
    base_dir = os.path.normpath(base_dir)

    if not os.path.isdir(base_dir):
        return {"exposure": [], "click": []}

    files = [
        os.path.join(base_dir, f)
        for f in os.listdir(base_dir)
        if f.startswith("关键词数据汇总_") and f.endswith(".xlsx")
    ]
    if not files:
        return {"exposure": [], "click": []}

    latest_file = max(files, key=os.path.getmtime)

    def _to_num(v) -> float:
        try:
            if pd.isna(v):
                return 0.0
            return float(str(v).replace(",", "").strip())
        except Exception:
            return 0.0

    def _extract_by_df(df: pd.DataFrame) -> List[Dict]:
        if df is None or df.empty:
            return []
        cols = list(df.columns)
        if len(cols) < 2:
            return []

        latest_col = cols[1]  # 老脚本是日期倒序，最新列在第2列
        result = []
        for _, row in df.iterrows():
            kw = str(row.iloc[0]).strip() if pd.notna(row.iloc[0]) else ""
            if not kw:
                continue
            val = _to_num(row.get(latest_col, 0))
            if val > 0:
                result.append({"keyword": kw, "value": val})

        result.sort(key=lambda x: x["value"], reverse=True)
        return result[:3000]

    def _extract(sheet_name: str, sheet_idx: int) -> List[Dict]:
        try:
            df = pd.read_excel(latest_file, sheet_name=sheet_name)
            rows = _extract_by_df(df)
            if rows:
                return rows
        except Exception:
            pass

        try:
            df = pd.read_excel(latest_file, sheet_name=sheet_idx)
            return _extract_by_df(df)
        except Exception:
            return []

    data = {
        "exposure": _extract("搜索曝光次数", 0) or _extract("曝光量", 0),
        "click": _extract("点击量", 1),
    }
    logger.info(f"关键词汇总返回: exposure={len(data['exposure'])}, click={len(data['click'])}")
    return data



def get_industry_keyword_latest(output_file: Optional[str] = None) -> Dict:
    """读取行业关键词整合结果，并返回源关键词池及最新热度列。"""
    cfg = get_config()
    path = (output_file or cfg.industry_keyword.output_file or "").strip()
    if not path:
        return {"columns": [], "rows": [], "latest_col": "", "source_pools": []}

    if re.match(r"^[\\/]+[A-Za-z]:", path):
        path = re.sub(r"^[\\/]+", "", path)
    path = os.path.normpath(path)

    if os.path.isdir(path) or not os.path.basename(path).lower().endswith((".xlsx", ".xls")):
        path = os.path.join(path, "关键词数据总表_宽表版.xlsx")
    if not os.path.exists(path):
        return {"columns": [], "rows": [], "latest_col": "", "source_pools": []}

    try:
        with pd.ExcelFile(path) as xls:
            target_sheet = "搜索指数" if "搜索指数" in xls.sheet_names else (xls.sheet_names[0] if xls.sheet_names else "")
            if not target_sheet:
                return {"columns": [], "rows": [], "latest_col": "", "source_pools": []}
            df = pd.read_excel(xls, sheet_name=target_sheet)
    except Exception:
        return {"columns": [], "rows": [], "latest_col": "", "source_pools": []}

    if df is None or df.empty:
        return {"columns": [], "rows": [], "latest_col": "", "source_pools": []}

    cols = [str(c) for c in list(df.columns)]
    keyword_col = "关键词" if "关键词" in cols else (cols[0] if cols else "")
    date_cols = [c for c in cols if c not in {"源关键词", "关键词"}]
    latest_col = date_cols[0] if date_cols else ""

    def _to_num(v):
        try:
            if pd.isna(v):
                return 0.0
            s = str(v).strip().replace(",", "").replace("%", "")
            return float(s) if s else 0.0
        except Exception:
            return 0.0

    if latest_col:
        df = df.sort_values(by=latest_col, key=lambda s: s.map(_to_num), ascending=False, kind="mergesort")

    rows = []
    source_pools: List[str] = []
    source_seen = set()
    for _, r in df.iterrows():
        obj = {c: (None if pd.isna(r.get(c)) else r.get(c)) for c in cols}
        if keyword_col and str(obj.get(keyword_col) or "").strip():
            source = str(obj.get("源关键词") or obj.get(keyword_col) or "").strip()
            if "源关键词" not in obj:
                obj["源关键词"] = source
            if source and source.lower() not in source_seen:
                source_seen.add(source.lower())
                source_pools.append(source)
            rows.append(obj)

    return {
        "columns": (["源关键词"] + cols if "源关键词" not in cols else cols),
        "rows": rows[:5000],
        "latest_col": latest_col,
        "source_pools": source_pools,
        "sheet": target_sheet,
        "file": path,
    }


def _extract_ai_response_text(data: Dict) -> str:
    for item in (data or {}).get("output", []):
        if item.get("type") == "message" and item.get("role") == "assistant":
            for c in item.get("content", []):
                if c.get("type") == "output_text":
                    return str(c.get("text") or "")
    return ""


def _normalize_keyword_list(items: List[str]) -> List[str]:
    result: List[str] = []
    seen = set()
    for x in items or []:
        val = str(x or "").strip()
        if not val:
            continue
        k = val.lower()
        if k in seen:
            continue
        seen.add(k)
        result.append(val)
    return result


def _parse_scene_list(text: str) -> List[str]:
    # 场景要求：一行一个场景（不使用逗号拆分，避免把场景名误拆）
    parts = re.split(r"[\n\r]+", str(text or ""))
    return _normalize_keyword_list(parts)


def _append_scene_titles_to_title_excel(excel_path: str, rows: List[Dict]) -> Dict:
    """
    增量写入到标题Excel（paths.title_excel_path）的「产品标题」sheet。
    只追加不覆盖：按(场景, 标题)去重；文件不存在则新建。
    """
    from openpyxl import Workbook, load_workbook

    path = str(excel_path or "").strip()
    if not path:
        raise ValueError("未配置标题Excel路径（paths.title_excel_path）")

    if re.match(r"^[\\/]+[A-Za-z]:", path):
        path = re.sub(r"^[\\/]+", "", path)
    path = os.path.normpath(path)

    parent = os.path.dirname(path)
    if parent:
        os.makedirs(parent, exist_ok=True)

    def _force_trim_title_max_128(title: str) -> str:
        """
        保存前强制规则：
        1) 若超过128字符，先删除最后一个 'for'（含）及其后全部内容；
        2) 若仍超过128字符，则从后往前按单词删除，直到 <=128。
        """
        t = " ".join(str(title or "").strip().split())
        if len(t) <= 128:
            return t

        # 优先删除最后一个 for 及其后内容（大小写不敏感）
        m = list(re.finditer(r"\bfor\b", t, flags=re.IGNORECASE))
        if m:
            t = t[:m[-1].start()].rstrip()
            t = " ".join(t.split())

        # 若仍超长，按单词从后往前删除
        if len(t) > 128:
            parts = t.split()
            while parts and len(" ".join(parts)) > 128:
                parts.pop()
            t = " ".join(parts).strip()
        return t

    wb = None
    created = False
    try:
        if os.path.exists(path):
            wb = load_workbook(path)
        else:
            wb = Workbook()
            created = True

        sheet_name = "产品标题"
        ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.create_sheet(title=sheet_name)

        # header map
        header_map: Dict[str, int] = {}
        if ws.max_row >= 1:
            for idx, cell in enumerate(ws[1], start=1):
                key = str(cell.value).strip() if cell.value is not None else ""
                if key:
                    header_map[key] = idx

        # 兼容旧格式表头：标题场景/标题打乱重组
        scene_col_name = "场景" if "场景" in header_map else ("标题场景" if "标题场景" in header_map else "")
        title_col_name = "标题" if "标题" in header_map else ("标题打乱重组" if "标题打乱重组" in header_map else "")

        if not header_map:
            ws.append(["场景", "标题"])
            header_map = {"场景": 1, "标题": 2}
            scene_col_name = "场景"
            title_col_name = "标题"

        # 若缺列，则在末尾追加列（不改动已有列顺序）
        if not scene_col_name:
            ws.cell(row=1, column=ws.max_column + 1, value="场景")
            scene_col_name = "场景"
            header_map[scene_col_name] = ws.max_column
        if not title_col_name:
            ws.cell(row=1, column=ws.max_column + 1, value="标题")
            title_col_name = "标题"
            header_map[title_col_name] = ws.max_column

        scene_idx = header_map.get(scene_col_name, 1)
        title_idx = header_map.get(title_col_name, 2)

        existing = set()
        # 读取已有数据用于去重
        if ws.max_row >= 2:
            for r in range(2, ws.max_row + 1):
                s = str(ws.cell(row=r, column=scene_idx).value or "").strip()
                t = str(ws.cell(row=r, column=title_idx).value or "").strip()
                if s and t:
                    existing.add((s, t))

        added = 0
        skipped = 0
        for item in rows or []:
            s = str((item or {}).get("scene", "")).strip()
            t = _force_trim_title_max_128(str((item or {}).get("title", "")).strip())
            if not s or not t:
                continue
            key = (s, t)
            if key in existing:
                skipped += 1
                continue

            # 构造整行：保留其他列为空
            row_values = ["" for _ in range(max(header_map.values()))]
            row_values[scene_idx - 1] = s
            row_values[title_idx - 1] = t
            ws.append(row_values)
            existing.add(key)
            added += 1

        # 清理默认Sheet（新建工作簿时）
        if created and "Sheet" in wb.sheetnames and len(wb.sheetnames) > 1:
            try:
                if wb["Sheet"].max_row == 1 and wb["Sheet"].max_column == 1 and wb["Sheet"]["A1"].value is None:
                    del wb["Sheet"]
            except Exception:
                pass

        wb.save(path)
        return {"file": path, "sheet": sheet_name, "added": added, "skipped": skipped}
    except PermissionError as e:
        raise ValueError(f"标题Excel文件被占用，请先关闭Excel后重试：{path}") from e
    finally:
        try:
            if wb:
                wb.close()
        except Exception:
            pass


def generate_industry_keyword_titles(
    mode: str,
    scenes: str,
    titles_per_scene: int,
    keywords: List[str],
    material: Optional[str] = None,
    output_file: Optional[str] = None,
    dropdown_output_file: Optional[str] = None,
    selected_source_pools: Optional[List[str]] = None,
    min_keyword_heat: float = 0,
    task: Optional["TaskInfo"] = None,
) -> Dict:
    """
    调用 AI 生成行业关键词标题（先支持行业热词模式，保留下拉词模式入口）。
    """
    cfg = get_config()
    normalized_mode = str(mode or "").strip()
    if normalized_mode not in {"industry_hot", "dropdown"}:
        raise ValueError("mode 仅支持 industry_hot 或 dropdown")

    scene_list = _parse_scene_list(scenes)
    if not scene_list:
        raise ValueError("请先输入场景（每行一个）")

    count = int(titles_per_scene or 0)
    if count <= 0:
        raise ValueError("每个场景生成数量必须大于0")
    if count > 50:
        raise ValueError("每个场景生成数量过大，请设置为50以内")

    if task:
        task.current_step = "按关键词池和热度筛选关键词"
    selected_pools = _normalize_keyword_list(selected_source_pools or [])
    if not selected_pools:
        raise ValueError("请至少选择一个源关键词池")
    selected_pool_keys = {x.lower() for x in selected_pools}
    try:
        min_heat = max(0.0, float(min_keyword_heat or 0))
    except Exception:
        min_heat = 0.0

    requested_keyword_keys = {x.lower() for x in _normalize_keyword_list(keywords or [])}
    if normalized_mode == "industry_hot":
        table = get_industry_keyword_latest(output_file)
        rows = table.get("rows", []) or []
        latest_col = str(table.get("latest_col") or "")

        def row_keyword(row: Dict) -> str:
            return str((row or {}).get("关键词", "") or "").strip()

        def row_heat(row: Dict) -> float:
            return _parse_numeric_value((row or {}).get(latest_col)) if latest_col else 0.0
    else:
        table = get_industry_keyword_dropdown_latest(dropdown_output_file)
        rows = table.get("rows", []) or []

        def row_keyword(row: Dict) -> str:
            return str((row or {}).get("下拉词", "") or (row or {}).get("US", "") or "").strip()

        def row_heat(row: Dict) -> float:
            return _parse_numeric_value((row or {}).get("关键词热度"))

    filtered_rows = []
    for row in rows:
        source = str((row or {}).get("源关键词", "") or (row or {}).get("原词", "") or "").strip()
        keyword = row_keyword(row)
        heat = row_heat(row)
        if not source or source.lower() not in selected_pool_keys:
            continue
        if heat < min_heat:
            continue
        if requested_keyword_keys and keyword.lower() not in requested_keyword_keys:
            continue
        if keyword:
            filtered_rows.append({"keyword": keyword, "source": source, "heat": heat})

    keyword_list = _normalize_keyword_list([str(row.get("keyword") or "") for row in filtered_rows])
    if not keyword_list:
        raise ValueError(f"所选关键词池中没有热度大于等于 {min_heat:g} 的可用关键词")

    api_key = str(getattr(cfg.data_analysis, "doubao_api_key", "") or "").strip()
    model_name = str(getattr(cfg.data_analysis, "doubao_model_name", "doubao-seed-2-0-pro-260215") or "").strip()
    if not api_key:
        raise ValueError("未配置豆包 API Key，请先在系统配置中填写")

    # 材质：一行一个，注入提示词时做去重合并
    material_items = _normalize_keyword_list(re.split(r"[\n\r]+", str(material or "")))
    material_text = ", ".join(material_items)
    if normalized_mode == "dropdown":
        prompt = (
            "生成阿里国际站英文标题，执行以下铁律，违反一条都禁止输出：\n\n"
            
            "【1. 长度强制锁死】\n"
            "字符数 = 包含空格\n"
            "必须严格控制在 85 - 128 字符之间\n"
            "如果标题字符超越128个，那么需要删除场景部分，删除场景部分还是超过128字符，就继续删除1个属性词\n"
            "生成后必须自己检查长度，超长直接重写\n\n"
            
            "【2. 必须极度精简】\n"
            "禁止任何多余单词\n"
            "禁止长描述\n"
            "禁止叠加修饰词\n"
            "单词越少越好，越短越好\n\n"
            
            "【3. 开头固定】\n"
            f"开头只能用1个：{', '.join(keyword_list)}\n"
            "必须放在最前面\n\n"
            
            "【4. 结构固定】\n"
            "下拉词 + 产品词 + 材质 + 1个属性词 + 场景\n\n"
            
            "【5. 材质】\n"
            f"{material_text or '简单材质词'}\n\n"
            
            "【6. 属性词】\n"
            "只能用1个，禁止多个\n\n"
            
            "【7. 场景】\n"
            f"场景：{', '.join(scene_list)}，每个场景{count}条\n\n"
            
            "【8. 格式】\n"
            "纯英文，无标点，空格分隔\n"

            "================【输出格式】================\n"
            "每行严格格式：\n"
            "场景, 标题\n"
            "禁止任何解释、序号或额外内容。\n"
        )

    else:
        prompt = (
            "你是阿里国际站B2B SEO标题生成专家，请在严格约束下生成英文产品标题。\n\n"

            "================【最高优先级规则】================\n"
            "标题总长度必须在85到128字符之间（包含空格）。\n"
            "如果出现冲突，必须优先减少关键词数量以满足长度要求。\n"
            "超过128字符的标题绝对禁止输出。\n"
            "生成后必须自行检查长度，不合规必须自动修改。\n\n"

            "================【关键词规则】================\n"
            f"只能使用以下关键词库中的词：{', '.join(keyword_list)}\n"
            "每个标题必须使用4到6个关键词（禁止超过6个）。\n"
            "关键词不能重复。\n"
            "优先选择较短关键词，避免接近长度上限。\n\n"

            "================【材质规则】================\n"
            f"材质：{material_text or '未指定则选择合理材质'}\n"
            "材质必须自然出现在标题中。\n\n"

            "================【结构规则】================\n"
            "标题结构必须符合：\n"
            "核心产品词 + 关键属性词 + 材质 + 变体词 + 使用场景\n\n"

            "================【风格规则】================\n"
            "仅输出英文。\n"
            "禁止任何标点符号或特殊字符。\n"
            "只能使用空格分隔。\n"
            "符合B2B外贸SEO风格，表达专业自然。\n"
            "避免关键词堆砌。\n\n"


            "================【使用场景】================\n"
            f"{', '.join(scene_list)}\n\n"

            f"每个场景生成 {count} 条标题。\n\n"

            "================【输出格式】================\n"
            "每行严格格式：\n"
            "场景, 标题\n"
            "禁止任何解释、序号或额外内容。\n"
        )

    payload = {
        "model": model_name or "doubao-seed-2-0-pro-260215",
        "temperature": 0.3,
        "top_p": 0.9,
        "input": [
            {
                "role": "user",
                "content": [{"type": "input_text", "text": prompt}],
            }
        ],
    }

    if task:
        task.current_step = "调用AI生成标题"
    resp = requests.post(
        ARK_API_BASE_URL + ARK_API_ENDPOINT,
        headers={"Authorization": f"Bearer {api_key}", "Content-Type": "application/json"},
        json=payload,
        timeout=300,
    )
    resp.raise_for_status()
    data = resp.json()
    content = _extract_ai_response_text(data).strip()
    if not content:
        raise ValueError("AI未返回有效标题结果")

    if task:
        task.current_step = "解析生成结果"
    parsed_rows: List[Dict] = []
    for ln in content.splitlines():
        line = str(ln or "").strip()
        if not line:
            continue
        if "，" in line:
            scene, title = line.split("，", 1)
        elif "," in line:
            scene, title = line.split(",", 1)
        else:
            continue
        scene = scene.strip()
        title = title.strip()
        if scene and title:
            parsed_rows.append({"scene": scene, "title": title})

    # 需求：不再落地 txt 结果文件，仅返回结果并写入标题Excel
    output_path = ""

    # 增量写入到配置的标题Excel路径（产品标题sheet）
    if task:
        task.current_step = "写入标题Excel"
    excel_write = None
    try:
        excel_write = _append_scene_titles_to_title_excel(getattr(cfg.paths, "title_excel_path", ""), parsed_rows)
    except Exception as e:
        # 写入失败不影响生成内容返回，但要把错误透出给前端提示
        excel_write = {"error": str(e), "file": str(getattr(cfg.paths, "title_excel_path", "") or "").strip(), "sheet": "产品标题", "added": 0, "skipped": 0}

    return {
        "mode": normalized_mode,
        "scene_count": len(scene_list),
        "titles_per_scene": count,
        "keyword_count": len(keyword_list),
        "keywords": keyword_list,
        "selected_source_pools": selected_pools,
        "min_keyword_heat": min_heat,
        "filtered_keyword_rows": filtered_rows,
        "content": content,
        "rows": parsed_rows,
        "output_file": output_path,
        "title_excel_write": excel_write,
    }


def run_industry_keyword_title_task(task: "TaskInfo", req: Dict[str, Any]):
    """
    异步任务入口：生成标题（行业热词/下拉词）。
    结果写入 task.result，并落盘到“标题生成_最新结果.json”。
    """
    from app.core.settings import get_config
    cfg = get_config()
    task.current_step = "初始化"
    if task.should_stop():
        return

    mode = str((req or {}).get("mode") or "").strip()
    scenes = str((req or {}).get("scenes") or "").strip()
    material = str((req or {}).get("material") or "").strip()
    titles_per_scene = int((req or {}).get("titles_per_scene") or 10)
    keywords = (req or {}).get("keywords") or []
    if not isinstance(keywords, list):
        keywords = []
    selected_source_pools = (req or {}).get("selected_source_pools") or []
    if not isinstance(selected_source_pools, list):
        selected_source_pools = []
    min_keyword_heat = (req or {}).get("min_keyword_heat") or 0
    output_file = (req or {}).get("output_file")
    dropdown_output_file = (req or {}).get("dropdown_output_file")

    result = generate_industry_keyword_titles(
        mode=mode,
        scenes=scenes,
        material=material or None,
        titles_per_scene=titles_per_scene,
        keywords=[str(x) for x in keywords],
        output_file=str(output_file).strip() if output_file else None,
        dropdown_output_file=str(dropdown_output_file).strip() if dropdown_output_file else None,
        selected_source_pools=[str(x) for x in selected_source_pools],
        min_keyword_heat=float(min_keyword_heat or 0),
        task=task,
    )
    task.result = result
    task.current_step = "完成"


def get_industry_keyword_title_generate_result() -> Dict:
    """读取最近一次标题生成任务结果（用于前端轮询结束后拉取）。"""
    # 不再落地结果文件；结果仅保存在任务内存中（路由层优先读取task.result）
    return {"generated_at": "", "result": None, "file": ""}


def get_industry_keyword_dropdown_latest(output_file: Optional[str] = None) -> Dict:
    """读取行业关键词下拉词结果表（默认 sheet: 下拉词）。"""
    cfg = get_config()
    path = (output_file or cfg.industry_keyword.dropdown_output_file or "").strip()
    if not path:
        return {"columns": [], "rows": [], "sheet": "", "file": ""}

    if re.match(r"^[\\/]+[A-Za-z]:", path):
        path = re.sub(r"^[\\/]+", "", path)
    path = os.path.normpath(path)

    if os.path.isdir(path) or not os.path.basename(path).lower().endswith((".xlsx", ".xls")):
        path = os.path.join(path, "下拉词结果.xlsx")
    if not os.path.exists(path):
        return {"columns": [], "rows": [], "sheet": "", "file": path}

    try:
        with pd.ExcelFile(path) as xls:
            target_sheet = "下拉词" if "下拉词" in xls.sheet_names else (xls.sheet_names[0] if xls.sheet_names else "")
            if not target_sheet:
                return {"columns": [], "rows": [], "sheet": "", "file": path}
            df = pd.read_excel(xls, sheet_name=target_sheet)
    except Exception:
        return {"columns": [], "rows": [], "sheet": "", "file": path}

    if df is None or df.empty:
        return {"columns": [], "rows": [], "sheet": target_sheet, "file": path}

    cols = [str(c) for c in list(df.columns)]
    rows = []
    for _, r in df.iterrows():
        obj = {}
        for c in cols:
            v = r.get(c)
            if pd.isna(v):
                obj[c] = ""
            elif isinstance(v, float):
                obj[c] = float(v)
            else:
                obj[c] = v
        # 过滤空行
        if any(str(obj.get(c, "")).strip() for c in cols):
            rows.append(obj)

    source_pools = _normalize_keyword_list([str((r or {}).get("源关键词", "") or (r or {}).get("原词", "")) for r in rows])
    return {"columns": cols, "rows": rows[:5000], "source_pools": source_pools, "sheet": target_sheet, "file": path}


def _read_excel_all_sheets(path: str) -> tuple[List[str], Dict[str, pd.DataFrame]]:
    """在单个 ExcelFile 上下文中读完所有 sheet，避免 Windows 下重复打开导致文件锁。"""
    with pd.ExcelFile(path) as xls:
        sheet_names = list(xls.sheet_names or [])
        frames: Dict[str, pd.DataFrame] = {}
        for sheet in sheet_names:
            try:
                frames[sheet] = pd.read_excel(xls, sheet_name=sheet)
            except Exception:
                frames[sheet] = pd.DataFrame()
    return sheet_names, frames


def _atomic_replace_excel(
    path: str,
    sheet_names: List[str],
    sheet_frames: Dict[str, pd.DataFrame],
    *,
    empty_sheet_name: str,
    empty_columns: List[str],
) -> None:
    """写入临时 Excel 后原子替换；Windows 下若被并发读取占用则短暂重试。"""
    dest_dir = os.path.dirname(path) or "."
    fd, tmp_path = tempfile.mkstemp(suffix=".xlsx", prefix="excel_atomic_", dir=dest_dir)
    os.close(fd)
    try:
        with pd.ExcelWriter(tmp_path, engine="openpyxl") as writer:
            wrote_any = False
            for sheet in sheet_names:
                cur = sheet_frames.get(sheet, pd.DataFrame())
                cur.to_excel(writer, sheet_name=sheet, index=False)
                wrote_any = True
            if not wrote_any:
                pd.DataFrame(columns=empty_columns).to_excel(writer, sheet_name=empty_sheet_name, index=False)

        gc.collect()
        last_err: Optional[Exception] = None
        for attempt in range(10):
            try:
                os.replace(tmp_path, path)
                return
            except PermissionError as e:
                last_err = e
                if attempt < 9:
                    time.sleep(0.25 * (attempt + 1))
                    gc.collect()
        raise ValueError(
            "文件暂无法写入（可能被 Excel/WPS、资源管理器预览或页面自动刷新占用），"
            f"请关闭相关程序或稍后重试：{path}"
        ) from last_err
    finally:
        try:
            if os.path.exists(tmp_path):
                os.remove(tmp_path)
        except Exception:
            pass


def delete_industry_keyword_rows(output_file: Optional[str], keywords: List[str]) -> Dict:
    """删除行业关键词整合表（搜索指数）中指定关键词行。"""
    cfg = get_config()
    path = (output_file or cfg.industry_keyword.output_file or "").strip()
    if not path:
        return {"deleted": 0, "total": 0, "file": ""}

    if re.match(r"^[\\/]+[A-Za-z]:", path):
        path = re.sub(r"^[\\/]+", "", path)
    path = os.path.normpath(path)
    if os.path.isdir(path) or not os.path.basename(path).lower().endswith((".xlsx", ".xls")):
        path = os.path.join(path, "关键词数据总表_宽表版.xlsx")
    if not os.path.exists(path):
        return {"deleted": 0, "total": 0, "file": path}

    targets = {str(k).strip() for k in (keywords or []) if str(k).strip()}
    if not targets:
        return {"deleted": 0, "total": 0, "file": path}

    sheet_names, sheet_frames = _read_excel_all_sheets(path)
    if not sheet_names:
        return {"deleted": 0, "total": 0, "file": path}

    target_sheet = "搜索指数" if "搜索指数" in sheet_names else sheet_names[0]
    deleted = 0
    total = 0

    for sheet in sheet_names:
        df = sheet_frames.get(sheet, pd.DataFrame())
        if df is None:
            df = pd.DataFrame()
        if sheet == target_sheet and not df.empty:
            cols = [str(c) for c in list(df.columns)]
            kw_col = "关键词" if "关键词" in cols else (cols[0] if cols else "")
            if kw_col:
                before = len(df)
                df = df[~df[kw_col].map(lambda x: str(x).strip() in targets)]
                after = len(df)
                deleted += (before - after)
                total += before
        sheet_frames[sheet] = df

    _atomic_replace_excel(
        path,
        sheet_names,
        sheet_frames,
        empty_sheet_name="搜索指数",
        empty_columns=["关键词"],
    )
    return {"deleted": deleted, "total": total, "file": path, "sheet": target_sheet}


def delete_industry_keyword_dropdown_rows(output_file: Optional[str], rows: List[Dict]) -> Dict:
    """删除下拉词结果表中指定行（按 原词+下拉词 精确匹配）。"""
    cfg = get_config()
    path = (output_file or cfg.industry_keyword.dropdown_output_file or "").strip()
    if not path:
        return {"deleted": 0, "total": 0, "file": ""}

    if re.match(r"^[\\/]+[A-Za-z]:", path):
        path = re.sub(r"^[\\/]+", "", path)
    path = os.path.normpath(path)
    if os.path.isdir(path) or not os.path.basename(path).lower().endswith((".xlsx", ".xls")):
        path = os.path.join(path, "下拉词结果.xlsx")
    if not os.path.exists(path):
        return {"deleted": 0, "total": 0, "file": path}

    targets = {
        (
            str((r or {}).get("源关键词", "")).strip(),
            str((r or {}).get("原词", "")).strip(),
            str((r or {}).get("下拉词", "")).strip(),
        )
        for r in (rows or [])
        if str((r or {}).get("原词", "")).strip() or str((r or {}).get("下拉词", "")).strip()
    }
    if not targets:
        return {"deleted": 0, "total": 0, "file": path}

    sheet_names, sheet_frames = _read_excel_all_sheets(path)
    target_sheet = "下拉词" if "下拉词" in sheet_names else (sheet_names[0] if sheet_names else "")
    if not target_sheet:
        return {"deleted": 0, "total": 0, "file": path}

    df = sheet_frames.get(target_sheet, pd.DataFrame())
    if df is None or df.empty:
        return {"deleted": 0, "total": 0, "file": path}

    cols = [str(c) for c in list(df.columns)]
    source_col = "源关键词" if "源关键词" in cols else ""
    origin_col = "原词" if "原词" in cols else (cols[0] if cols else "")
    suggestion_col = "下拉词" if "下拉词" in cols else ("US" if "US" in cols else (cols[-1] if len(cols) > 1 else ""))
    if not origin_col or not suggestion_col:
        return {"deleted": 0, "total": len(df), "file": path}

    before = len(df)
    mask = df.apply(
        lambda r: (
            str(r.get(source_col, "")).strip() if source_col else "",
            str(r.get(origin_col, "")).strip(),
            str(r.get(suggestion_col, "")).strip(),
        ) in targets,
        axis=1,
    )
    df = df[~mask]
    deleted = before - len(df)
    sheet_frames[target_sheet] = df

    _atomic_replace_excel(
        path,
        sheet_names,
        sheet_frames,
        empty_sheet_name="下拉词",
        empty_columns=["源关键词", "原词", "关键词热度", "US"],
    )
    return {"deleted": deleted, "total": before, "file": path, "sheet": target_sheet}


def get_product360_table(output_dir: Optional[str] = None, sheet_name: Optional[str] = None) -> Dict:
    """严格读取产品360配置的 Excel结果 路径下固定文件：产品数据总报告.xlsx。"""
    cfg = get_config()
    excel_result_dir = (output_dir or getattr(cfg.data_download, "product360_excel_result_dir", "") or "").strip()
    if not excel_result_dir:
        return {"sheet": "", "sheets": [], "columns": [], "rows": []}

    if re.match(r"^[\\/]+[A-Za-z]:", excel_result_dir):
        excel_result_dir = re.sub(r"^[\\/]+", "", excel_result_dir)
    excel_result_dir = os.path.normpath(excel_result_dir)

    if os.path.isfile(excel_result_dir):
        target_file = excel_result_dir
    else:
        if not os.path.isdir(excel_result_dir):
            return {"sheet": "", "sheets": [], "columns": [], "rows": []}
        target_file = os.path.join(excel_result_dir, "产品数据总报告.xlsx")

    if not os.path.exists(target_file):
        return {"sheet": "", "sheets": [], "columns": [], "rows": []}

    try:
        xl = pd.ExcelFile(target_file)
        sheets = xl.sheet_names or []
        default_sheet = "产品详细信息" if "产品详细信息" in sheets else (sheets[0] if sheets else "")
        target = sheet_name or default_sheet
        if not target:
            return {"sheet": "", "sheets": sheets, "columns": [], "rows": []}

        df = pd.read_excel(target_file, sheet_name=target)
        if df is None or df.empty:
            return {"sheet": target, "sheets": sheets, "columns": [], "rows": []}

        # 指定sheet字段筛选
        if target == "产品详细信息":
            wanted = [
                "产品ID", "平台内部产品分级", "产品综合搜索排名", "平台曝光标签", "平台访客流量标签",
                "平台点击效率标签", "平台转化效果标签", "产品总曝光量", "各流量渠道UV分布", "买家实际访问来源国家",
            ]
            keep = [c for c in wanted if c in df.columns]
            if keep:
                df = df[keep]
        elif target == "访客地域":
            # 必须保留产品ID，前端据此按当前选中产品过滤
            wanted = ["产品ID", "国家(中文)", "国家代码", "访客数(UV)", "访客量"]
            keep = [c for c in wanted if c in df.columns]
            if keep:
                df = df[keep]
        elif target == "流量来源":
            wanted = ["产品ID", "日期", "流量渠道类型", "店铺访问人数", "询盘人数", "TM咨询人数"]
            keep = [c for c in wanted if c in df.columns]
            if keep:
                df = df[keep]

        cols = [str(c) for c in df.columns.tolist()]
        rows = []
        for _, r in df.iterrows():
            item = {}
            for c in cols:
                v = r.get(c)
                if pd.isna(v):
                    item[c] = ""
                elif isinstance(v, float):
                    item[c] = float(v)
                else:
                    item[c] = v
            rows.append(item)

        return {"sheet": target, "sheets": sheets, "columns": cols, "rows": rows, "file": target_file}
    except Exception:
        return {"sheet": "", "sheets": [], "columns": [], "rows": []}


def get_product360_traffic_channels(output_dir: Optional[str] = None, product_ids: Optional[List[str]] = None) -> Dict:
    """
    从产品360 Excel结果（产品数据总报告.xlsx）的「流量来源」sheet 中，
    按给定产品ID筛选，并对每个产品取最新日期的数据，返回“产品ID + 各渠道(店铺访问人数)”的透视表。
    """
    cfg = get_config()
    excel_result_dir = (output_dir or getattr(cfg.data_download, "product360_excel_result_dir", "") or "").strip()
    if not excel_result_dir:
        return {"columns": ["产品ID"], "rows": [], "file": "", "error": "未配置产品360 Excel结果路径"}

    if re.match(r"^[\\/]+[A-Za-z]:", excel_result_dir):
        excel_result_dir = re.sub(r"^[\\/]+", "", excel_result_dir)
    excel_result_dir = os.path.normpath(excel_result_dir)

    if os.path.isfile(excel_result_dir):
        target_file = excel_result_dir
    else:
        if not os.path.isdir(excel_result_dir):
            return {"columns": ["产品ID"], "rows": [], "file": "", "error": "Excel结果路径不存在"}
        target_file = os.path.join(excel_result_dir, "产品数据总报告.xlsx")

    if not os.path.exists(target_file):
        return {"columns": ["产品ID"], "rows": [], "file": target_file, "error": "产品数据总报告.xlsx 不存在"}

    def _norm_pid(v: object) -> str:
        s = str(v or "").strip()
        if s.endswith(".0"):
            s = re.sub(r"\.0+$", "", s)
        return s

    wanted_ids = [_norm_pid(x) for x in (product_ids or []) if _norm_pid(x)]
    wanted_set = set(wanted_ids)
    if not wanted_set:
        return {"columns": ["产品ID"], "rows": [], "file": target_file, "error": "未提供产品ID"}

    try:
        xl = pd.ExcelFile(target_file)
        sheets = xl.sheet_names or []
        sheet = "流量来源" if "流量来源" in sheets else ("流量来源 " if "流量来源 " in sheets else "")
        if not sheet:
            # 模糊匹配一次
            for s in sheets:
                if "流量" in str(s) and "来源" in str(s):
                    sheet = s
                    break
        if not sheet:
            return {"columns": ["产品ID"], "rows": [], "file": target_file, "error": "未找到“流量来源”sheet"}

        df = pd.read_excel(
            target_file,
            sheet_name=sheet,
            usecols=lambda c: str(c).strip() in ("产品ID", "日期", "流量渠道", "流量渠道类型", "店铺访问人数"),
            dtype={"产品ID": str},
        )
        if df is None or df.empty:
            return {"columns": ["产品ID"], "rows": [], "file": target_file, "sheet": sheet, "error": "流量来源 sheet 为空"}

        # 统一字段名
        if "流量渠道" not in df.columns and "流量渠道类型" in df.columns:
            df = df.rename(columns={"流量渠道类型": "流量渠道"})

        df["产品ID"] = df["产品ID"].map(_norm_pid)
        df = df[df["产品ID"].isin(wanted_set)]
        if df.empty:
            return {"columns": ["产品ID"], "rows": [], "file": target_file, "sheet": sheet, "error": "未匹配到任何产品ID"}

        # 日期解析
        if "日期" in df.columns:
            df["日期"] = pd.to_datetime(df["日期"], errors="coerce")
        else:
            df["日期"] = pd.NaT

        df["店铺访问人数"] = pd.to_numeric(df.get("店铺访问人数"), errors="coerce").fillna(0)
        df["流量渠道"] = df.get("流量渠道", "").astype(str).str.strip()

        # 每个产品取最新日期（忽略 NaT）
        latest_date = df.dropna(subset=["日期"]).groupby("产品ID")["日期"].max()
        # 没有日期的产品也要保留：用 NaT 作为 key
        df = df.merge(latest_date.rename("最新日期"), on="产品ID", how="left")
        df_latest = df[(df["日期"].isna() & df["最新日期"].isna()) | (df["日期"] == df["最新日期"])]

        # pivot: 产品ID x 流量渠道 -> 店铺访问人数
        pivot = (
            df_latest.pivot_table(index="产品ID", columns="流量渠道", values="店铺访问人数", aggfunc="sum", fill_value=0)
            .reset_index()
        )

        # 渠道列顺序：按出现顺序（在 df_latest 中）
        channel_order = []
        seen = set()
        for ch in df_latest["流量渠道"].tolist():
            chs = str(ch or "").strip()
            if not chs or chs in seen:
                continue
            seen.add(chs)
            channel_order.append(chs)

        cols = ["产品ID"] + [c for c in channel_order if c in pivot.columns]
        # 防止有渠道列未进入 order（极少数），补到末尾
        for c in pivot.columns.tolist():
            if c != "产品ID" and c not in cols:
                cols.append(str(c))
        pivot = pivot[cols]

        rows = pivot.to_dict(orient="records")
        return {"columns": cols, "rows": rows, "file": target_file, "sheet": sheet, "error": ""}
    except Exception as e:
        return {"columns": ["产品ID"], "rows": [], "file": target_file, "error": str(e)}


def get_product_operate_table(file_path: Optional[str] = None) -> Dict:
    """读取产品运营导出 Excel，返回列与行。"""
    cfg = get_config()
    raw_path = (file_path or "").strip()
    cfg_path = (getattr(cfg.data_download, "product_operate_output_file", "") or "").strip()

    def _clean_path(p: str) -> str:
        p = str(p or "").strip().strip('"').strip("'")
        # 兼容从 IM/文档复制路径时出现的全角符号与零宽字符
        p = p.replace("\u200b", "").replace("\ufeff", "")
        p = p.replace("：", ":").replace("／", "/").replace("＼", "\\")
        if re.match(r"^[\\/]+[A-Za-z]:", p):
            p = re.sub(r"^[\\/]+", "", p)
        p = os.path.normpath(p) if p else ""
        return p

    def _pick_existing_excel_path(primary: str, fallback: str) -> str:
        candidates: List[str] = []
        p1 = _clean_path(primary)
        p2 = _clean_path(fallback)
        if p1:
            candidates.append(p1)
        if p2 and p2 not in candidates:
            candidates.append(p2)

        # 相对路径按 project_files_root 补全（安装包常见）
        root = str(getattr(cfg.paths, "project_files_root", "") or "").strip()
        for cp in list(candidates):
            if cp and not os.path.isabs(cp) and root:
                abs_cp = os.path.normpath(os.path.join(root, cp))
                if abs_cp not in candidates:
                    candidates.append(abs_cp)

        for cp in candidates:
            if cp and os.path.exists(cp):
                return cp

        # 若目标文件不存在，尝试同目录下最新的 xlsx/xls
        for cp in candidates:
            if not cp:
                continue
            parent = os.path.dirname(cp)
            if not parent or not os.path.isdir(parent):
                continue
            try:
                excel_files = [
                    os.path.join(parent, fn)
                    for fn in os.listdir(parent)
                    if str(fn).lower().endswith((".xlsx", ".xls")) and os.path.isfile(os.path.join(parent, fn))
                ]
                if excel_files:
                    excel_files.sort(key=lambda f: os.path.getmtime(f), reverse=True)
                    return excel_files[0]
            except Exception:
                continue
        return candidates[0] if candidates else ""

    path = _pick_existing_excel_path(raw_path, cfg_path)
    if not path:
        return {"columns": [], "rows": [], "file": "", "error": "未配置产品运营输出文件路径"}

    if not os.path.exists(path):
        return {
            "columns": [],
            "rows": [],
            "file": path,
            "error": "文件不存在，可能是安装包环境下路径未同步；请在右侧配置里重选一次输出文件并保存后刷新",
        }

    last_error = ""
    for _ in range(3):
        try:
            # 兼容：有些文件首个 sheet 为空，但其他 sheet 有数据
            book = pd.read_excel(path, sheet_name=None)
            if not isinstance(book, dict) or not book:
                return {"columns": [], "rows": [], "file": path, "error": "Excel 无可读取 sheet"}

            target_sheet = ""
            df = None
            max_rows = -1
            for sh, sdf in book.items():
                if sdf is None:
                    continue
                row_count = int(getattr(sdf, "shape", [0, 0])[0] or 0)
                if row_count > max_rows:
                    max_rows = row_count
                    target_sheet = str(sh)
                    df = sdf

            if df is None:
                return {"columns": [], "rows": [], "file": path, "error": "Excel sheet 读取失败"}
            if df.empty:
                logger.info(f"产品运营表读取成功但为空: file={path} sheet={target_sheet}")
                return {"columns": [], "rows": [], "file": path, "sheet": target_sheet, "error": "文件存在但暂无数据行"}

            cols = [str(c) for c in df.columns.tolist()]
            rows = []
            for _, r in df.iterrows():
                item = {}
                for c in cols:
                    v = r.get(c)
                    if pd.isna(v):
                        item[c] = ""
                    elif isinstance(v, float):
                        item[c] = float(v)
                    else:
                        item[c] = v
                rows.append(item)
            logger.info(f"产品运营表读取成功: file={path} sheet={target_sheet} rows={len(rows)} cols={len(cols)}")
            return {"columns": cols, "rows": rows, "file": path, "sheet": target_sheet, "error": ""}
        except Exception as e:
            last_error = str(e)
            time.sleep(0.15)

    logger.warning(f"读取产品运营表格失败: {path} | {last_error}")
    return {"columns": [], "rows": [], "file": path, "error": f"读取失败: {last_error}"}


def get_traffic_channel_overview(file_path: Optional[str] = None, sheet_name: Optional[str] = None) -> Dict:
    """读取流量渠道分析结果，返回当天/本周/本月汇总和分析表。"""
    cfg = get_config()
    path = (file_path or getattr(cfg.data_download, "traffic_channel_output_file", "") or "").strip()
    if not path:
        return {"sheets": [], "sheet": "", "today": [], "week": [], "month": [], "analysis": {"columns": [], "rows": []}}

    if re.match(r"^[\\/]+[A-Za-z]:", path):
        path = re.sub(r"^[\\/]+", "", path)
    path = os.path.normpath(path)

    if not os.path.exists(path):
        return {"sheets": [], "sheet": "", "today": [], "week": [], "month": [], "analysis": {"columns": [], "rows": []}}

    try:
        xl = pd.ExcelFile(path)
        sheets = xl.sheet_names or []
        target = sheet_name or (sheets[0] if sheets else "")
        if not target:
            return {"sheets": sheets, "sheet": "", "today": [], "week": [], "month": [], "analysis": {"columns": [], "rows": []}}

        def _pick_metric(sheet_df: pd.DataFrame, metric: str):
            if sheet_df is None or sheet_df.empty or "指标" not in sheet_df.columns:
                return None
            row = sheet_df[sheet_df["指标"] == metric]
            return row.iloc[0] if not row.empty else None

        def _date_cols(sheet_df: pd.DataFrame):
            cols = []
            for c in sheet_df.columns:
                s = str(c)
                if re.fullmatch(r"\d{6}", s):
                    cols.append(s)
            return sorted(cols, reverse=True)

        today_rows = []
        week_rows = []
        month_rows = []

        for sh in sheets:
            sdf = pd.read_excel(path, sheet_name=sh)
            dcols = _date_cols(sdf)
            if not dcols:
                continue

            uv_row = _pick_metric(sdf, "店铺访问人数")
            ask_row = _pick_metric(sdf, "店内询盘人数")
            tm_row = _pick_metric(sdf, "店内TM咨询人数")
            if uv_row is None or ask_row is None or tm_row is None:
                continue

            dated_cols = []
            for c in dcols:
                dt = pd.to_datetime(c, format="%y%m%d", errors="coerce")
                if pd.isna(dt):
                    continue
                dated_cols.append((c, dt.date()))
            if not dated_cols:
                continue

            latest_date = max(d for _, d in dated_cols)

            def _sum_for(row, cols):
                vals = [pd.to_numeric(row.get(c, 0), errors="coerce") for c in cols]
                vals = [0 if pd.isna(v) else float(v) for v in vals]
                return sum(vals)

            today_cols = [c for c, d in dated_cols if d == latest_date]

            t_uv = _sum_for(uv_row, today_cols)
            t_ask = _sum_for(ask_row, today_cols)
            t_tm = _sum_for(tm_row, today_cols)

            today_rows.append(
                {
                    "流量渠道": sh,
                    "店铺访问人数": t_uv,
                    "店内询盘人数": t_ask,
                    "店内TM咨询人数": t_tm,
                    "商机转化率": ((t_ask + t_tm) / t_uv if t_uv else 0),
                }
            )

        # 上周 / 上月：与阿里接口 statisticsType=week|month & selected=1 对齐（勿用日列相加）
        week_type, week_sel = TRAFFIC_CHANNEL_OVERVIEW_API["week"]
        month_type, month_sel = TRAFFIC_CHANNEL_OVERVIEW_API["month"]
        week_rows = _fetch_traffic_channel_overview_rows(week_type, week_sel)
        month_rows = _fetch_traffic_channel_overview_rows(month_type, month_sel)

        if not week_rows:
            logger.warning("流量渠道上周数据接口为空，回退为 Excel 日列汇总")
            week_rows = []
            for sh in sheets:
                sdf = pd.read_excel(path, sheet_name=sh)
                dcols = _date_cols(sdf)
                if not dcols:
                    continue
                uv_row = _pick_metric(sdf, "店铺访问人数")
                ask_row = _pick_metric(sdf, "店内询盘人数")
                tm_row = _pick_metric(sdf, "店内TM咨询人数")
                if uv_row is None or ask_row is None or tm_row is None:
                    continue
                dated_cols = []
                for c in dcols:
                    dt = pd.to_datetime(c, format="%y%m%d", errors="coerce")
                    if pd.isna(dt):
                        continue
                    dated_cols.append((c, dt.date()))
                if not dated_cols:
                    continue
                latest_date = max(d for _, d in dated_cols)
                cur_week_start = latest_date - timedelta(days=(latest_date.weekday() + 1) % 7)
                last_week_start = cur_week_start - timedelta(days=7)
                last_week_end = cur_week_start - timedelta(days=1)
                week_cols = [c for c, d in dated_cols if last_week_start <= d <= last_week_end]
                w_uv = _sum_for(uv_row, week_cols)
                w_ask = _sum_for(ask_row, week_cols)
                w_tm = _sum_for(tm_row, week_cols)
                week_rows.append(
                    {
                        "流量渠道": sh,
                        "店铺访问人数": w_uv,
                        "店内询盘人数": w_ask,
                        "店内TM咨询人数": w_tm,
                        "商机转化率": ((w_ask + w_tm) / w_uv if w_uv else 0),
                    }
                )

        if not month_rows:
            logger.warning("流量渠道上月数据接口为空，回退为 Excel 日列汇总")
            month_rows = []
            for sh in sheets:
                sdf = pd.read_excel(path, sheet_name=sh)
                dcols = _date_cols(sdf)
                if not dcols:
                    continue
                uv_row = _pick_metric(sdf, "店铺访问人数")
                ask_row = _pick_metric(sdf, "店内询盘人数")
                tm_row = _pick_metric(sdf, "店内TM咨询人数")
                if uv_row is None or ask_row is None or tm_row is None:
                    continue
                dated_cols = []
                for c in dcols:
                    dt = pd.to_datetime(c, format="%y%m%d", errors="coerce")
                    if pd.isna(dt):
                        continue
                    dated_cols.append((c, dt.date()))
                if not dated_cols:
                    continue
                latest_date = max(d for _, d in dated_cols)
                if latest_date.month == 1:
                    last_month_year = latest_date.year - 1
                    last_month = 12
                else:
                    last_month_year = latest_date.year
                    last_month = latest_date.month - 1
                month_cols = [
                    c for c, d in dated_cols if d.year == last_month_year and d.month == last_month
                ]
                m_uv = _sum_for(uv_row, month_cols)
                m_ask = _sum_for(ask_row, month_cols)
                m_tm = _sum_for(tm_row, month_cols)
                month_rows.append(
                    {
                        "流量渠道": sh,
                        "店铺访问人数": m_uv,
                        "店内询盘人数": m_ask,
                        "店内TM咨询人数": m_tm,
                        "商机转化率": ((m_ask + m_tm) / m_uv if m_uv else 0),
                    }
                )

        analysis_df = pd.read_excel(path, sheet_name=target)
        if analysis_df is None or analysis_df.empty:
            analysis = {"columns": [], "rows": []}
        else:
            analysis_df = analysis_df.replace([pd.NA], None)
            analysis = {
                "columns": [str(c) for c in analysis_df.columns],
                "rows": analysis_df.to_dict(orient="records"),
            }

        return {
            "sheets": sheets,
            "sheet": target,
            "today": today_rows,
            "week": week_rows,
            "month": month_rows,
            "analysis": analysis,
        }
    except Exception:
        return {"sheets": [], "sheet": "", "today": [], "week": [], "month": [], "analysis": {"columns": [], "rows": []}}


def get_store_summary_table(file_path: Optional[str] = None, sheet_name: Optional[str] = None) -> Dict:
    """读取店铺周汇总Excel表格，默认返回“总结”sheet。"""
    cfg = get_config()
    path = (file_path or cfg.store_overview.summary_output_path or "").strip()
    if not path:
        return {"sheet": "", "sheets": [], "columns": [], "rows": []}

    if re.match(r"^[\\/]+[A-Za-z]:", path):
        path = re.sub(r"^[\\/]+", "", path)
    path = os.path.normpath(path)

    if not os.path.exists(path):
        return {"sheet": "", "sheets": [], "columns": [], "rows": []}

    try:
        xl = pd.ExcelFile(path)
        sheets = xl.sheet_names or []
        target = sheet_name or ("总结" if "总结" in sheets else (sheets[0] if sheets else ""))
        if not target:
            return {"sheet": "", "sheets": sheets, "columns": [], "rows": []}

        # “总结”sheet 是两行表头（第一行分组合并，第二行明细），前端按第二行展示可避免 Unnamed 列
        if str(target).strip() == "总结":
            df = pd.read_excel(path, sheet_name=target, header=1)
        else:
            df = pd.read_excel(path, sheet_name=target)

        if df is None or df.empty:
            return {"sheet": target, "sheets": sheets, "columns": [], "rows": []}

        cols = [str(c) for c in df.columns.tolist()]
        if cols and (cols[0].lower() == "unnamed: 0" or cols[0].strip() == ""):
            cols[0] = "指标名称"
            df.columns = cols
        rows = []
        for _, r in df.iterrows():
            item = {}
            for c in cols:
                v = r.get(c)
                if pd.isna(v):
                    item[c] = ""
                elif isinstance(v, float):
                    item[c] = float(v)
                else:
                    item[c] = v
            rows.append(item)

        return {"sheet": target, "sheets": sheets, "columns": cols, "rows": rows}
    except Exception:
        return {"sheet": "", "sheets": [], "columns": [], "rows": []}


def _store_image_download_image(url: str, save_path: str):
    if os.path.exists(save_path):
        return False
    try:
        headers = {
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36",
            "Referer": "https://szdabojin.en.alibaba.com/",
        }
        response = requests.get(url, headers=headers, timeout=20)
        if response.status_code == 200:
            with open(save_path, "wb") as f:
                f.write(response.content)
            return True
    except Exception as e:
        logger.warning(f"图片下载失败: {url} | {e}")
    return False


def _download_store_images(task: TaskInfo, cfg):
    """店铺图片采集：完全对齐老脚本提取图片和标题.py 的逻辑"""
    task.current_step = "店铺图片采集 - 启动浏览器..."

    dcfg = cfg.data_download
    base_url = (getattr(dcfg, "store_image_target_url", "") or "https://szdabojin.en.alibaba.com/productlist.html").strip()
    save_dir = _normalize_daily_dir(getattr(dcfg, "store_image_save_dir", "") or "")
    logger.info(f"店铺图片采集保存目录: {save_dir}")
    logger.info(f"店铺图片采集当前工作目录: {os.getcwd()}")
    logger.info(f"店铺图片采集目标页: {base_url}")
    max_total_pages = int(getattr(dcfg, "store_image_max_pages", 100) or 100)
    max_total_pages = max(1, min(max_total_pages, 500))
    excel_file = os.path.join(save_dir, "产品标题.xlsx")

    if not save_dir:
        raise Exception("未配置店铺图片保存目录")
    os.makedirs(save_dir, exist_ok=True)

    def download_image(url: str, save_path: str) -> bool:
        if os.path.exists(save_path):
            return False
        try:
            headers = {
                "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36",
                "Referer": "https://szdabojin.en.alibaba.com/",
            }
            response = requests.get(url, headers=headers, timeout=20)
            if response.status_code == 200:
                with open(save_path, "wb") as f:
                    f.write(response.content)
                return True
        except Exception:
            return False
        return False

    def load_existing_titles() -> Dict[str, str]:
        title_map: Dict[str, str] = {}
        if os.path.exists(excel_file):
            try:
                df = pd.read_excel(excel_file, engine="openpyxl", dtype=str)
                for _, row in df.iterrows():
                    pid = str(row.get("产品ID", "")).strip()
                    title = str(row.get("产品标题", "")).strip()
                    if pid and title and pid != "nan" and title != "nan":
                        title_map[pid] = title
            except Exception:
                pass
        return title_map

    def append_title_to_excel(new_pid: str, new_title: str, title_map: Dict[str, str]):
        if not new_pid or not new_title:
            return
        if new_pid in title_map:
            return
        new_row = pd.DataFrame([{"产品ID": new_pid, "产品标题": new_title}])
        try:
            if not os.path.exists(excel_file):
                new_row.to_excel(excel_file, index=False, engine="openpyxl")
            else:
                try:
                    df = pd.read_excel(excel_file, engine="openpyxl", dtype=str)
                except Exception:
                    df = pd.DataFrame(columns=["产品ID", "产品标题"])
                df = pd.concat([df, new_row], ignore_index=True)
                df.to_excel(excel_file, index=False, engine="openpyxl")
            title_map[new_pid] = new_title
            logger.info(f"店铺图片采集标题已写入: {excel_file} | {new_pid} -> {new_title}")
        except Exception as e:
            logger.error(f"店铺图片采集标题写入失败: {excel_file} | {new_pid} | {e}")

    def auto_wait_verification(page):
        print("[check] 检查验证状态...")
        for _ in range(300):
            time.sleep(1)
            try:
                if page.query_selector(".icbu-product-card, .product-item"):
                    print("[ok] 已进入产品列表")
                    return True
                body = page.text_content("body") or ""
                if any(k in body for k in ["滑块", "验证", "slide", "安全验证"]):
                    print("[wait] 等待验证...")
            except Exception:
                continue
        print("[fail] 验证超时")
        return False

    def clean_img_url(src: str) -> str:
        if not src:
            return ""
        clean = str(src).split(".jpg_")[0] + ".jpg"
        if clean.startswith("//"):
            clean = f"https:{clean}"
        return clean

    from playwright.sync_api import sync_playwright

    title_map = load_existing_titles()
    print(f"[info] 已加载已有标题：{len(title_map)} 条")

    with sync_playwright() as p:
        browser = _launch_playwright_browser(
            p,
            headless=False,
            slow_mo=50,
            args=["--start-maximized", "--disable-blink-features=AutomationControlled"],
        )
        logger.info("店铺图片采集浏览器已启动")
        ctx = browser.new_context(
            user_agent="Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36",
            viewport=None,
        )
        ctx.add_init_script("Object.defineProperty(navigator,'webdriver',{get:()=>false})")
        page = ctx.new_page()

        task.total = max_total_pages
        task.progress = 0
        print(f"[info] 开始抓取，共 {max_total_pages} 页（手动设置）")

        all_processed_ids = set()
        for page_num in range(1, max_total_pages + 1):
            if task.should_stop():
                break
            task.wait_if_paused()
            task.current_step = f"店铺图片采集 - 第 {page_num}/{max_total_pages} 页"
            task.progress = page_num - 1

            url = f"{base_url.rsplit('.html', 1)[0]}-{page_num}.html" if page_num > 1 else base_url
            print(f"[page] ===== 第 {page_num} 页 =====")

            try:
                page.goto(url, wait_until="domcontentloaded", timeout=60000)
                logger.info(f"店铺图片采集已打开页面: {page.url}")
                verified = auto_wait_verification(page)
                logger.info(f"店铺图片采集验证结果: {verified}")
                page.mouse.wheel(0, 1000)
                time.sleep(2)

                data = page.evaluate('''()=>{
                    return Array.from(document.querySelectorAll('.icbu-product-card,.product-item')).map(i=>{
                        return {
                            id:i.getAttribute('data-id'),
                            src:i.querySelector('img')?.src||'',
                            title:i.querySelector('.title-con')?.innerText.trim()||''
                        }
                    }).filter(x=>x.id&&x.src)
                }''')

                if not data:
                    print("[warn] 无产品")
                    continue

                tasks = []
                skip = 0
                add_title = 0

                for item in data:
                    pid = str(item.get("id") or "").strip()
                    src = str(item.get("src") or "").strip()
                    title = str(item.get("title") or "").strip()
                    img_path = os.path.join(save_dir, f"{pid}.jpg")

                    img_exist = os.path.exists(img_path)
                    title_exist = pid in title_map

                    if img_exist and title_exist:
                        skip += 1
                        continue
                    if img_exist and not title_exist:
                        if title:
                            append_title_to_excel(pid, title, title_map)
                            add_title += 1
                    if not img_exist and title_exist:
                        clean = clean_img_url(src)
                        if clean:
                            tasks.append((clean, img_path))
                            all_processed_ids.add(pid)
                    if not img_exist and not title_exist:
                        clean = clean_img_url(src)
                        if clean:
                            tasks.append((clean, img_path))
                            all_processed_ids.add(pid)
                        if title:
                            append_title_to_excel(pid, title, title_map)
                            add_title += 1

                downloaded = 0
                for clean, img_path in tasks:
                    if download_image(clean, img_path):
                        downloaded += 1

                if tasks:
                    print(f"[download] 下载图片：{downloaded} 张")
                if add_title > 0:
                    print(f"[title] 增量新增标题：{add_title} 条")
                if skip > 0:
                    print(f"[skip] 已完整（图+标题）：{skip} 个 -> 跳过")
                logger.info(f"第{page_num}页处理完成：下载 {downloaded}/{len(tasks)}，新增标题 {add_title}，跳过 {skip}")

            except Exception as e:
                print(f"[error] 页面异常：{e}")

        task.progress = task.total
        task.current_step = f"店铺图片采集完成，共采集图片 {len(all_processed_ids)} 张，标题 {len(title_map)} 条"
        try:
            browser.close()
        except Exception:
            pass


def get_store_image_list(save_dir: Optional[str] = None, keyword: Optional[str] = None) -> Dict:
    cfg = get_config()
    path = (save_dir or getattr(cfg.data_download, "store_image_save_dir", "") or "").strip()
    if not path:
        return {"save_dir": "", "total": 0, "items": [], "titles": [], "title_total": 0}

    if re.match(r"^[\\/]+[A-Za-z]:", path):
        path = re.sub(r"^[\\/]+", "", path)
    path = os.path.normpath(path)

    if not os.path.isdir(path):
        return {"save_dir": path, "total": 0, "items": [], "titles": [], "title_total": 0}

    kw = (keyword or "").strip().lower()
    items = []
    for name in os.listdir(path):
        lower = name.lower()
        if not lower.endswith((".jpg", ".jpeg", ".png", ".webp")):
            continue
        if kw and kw not in lower:
            continue
        full = os.path.join(path, name)
        try:
            stat = os.stat(full)
            items.append({
                "id": os.path.splitext(name)[0],
                "name": name,
                "path": full,
                "size": int(stat.st_size),
                "mtime": int(stat.st_mtime),
            })
        except Exception:
            continue

    title_file = os.path.join(path, "产品标题.xlsx")
    titles = []
    if os.path.exists(title_file):
        try:
            df = pd.read_excel(title_file, engine="openpyxl", dtype=str)
            for _, r in df.iterrows():
                pid = str(r.get("产品ID", "")).strip()
                title = str(r.get("产品标题", "")).strip()
                if not pid or pid.lower() == "nan":
                    continue
                if not title or title.lower() == "nan":
                    continue
                if kw and (kw not in pid.lower() and kw not in title.lower()):
                    continue
                titles.append({"id": pid, "title": title})
        except Exception:
            pass

    items.sort(key=lambda x: x.get("mtime", 0), reverse=True)
    titles.sort(key=lambda x: x.get("id", ""), reverse=False)
    return {"save_dir": path, "total": len(items), "items": items, "titles": titles, "title_total": len(titles)}


def get_store_overview_latest(save_path: Optional[str] = None, include_details: bool = True) -> Dict:
    """读取店铺运营Excel并返回最新指标+周期趋势"""
    cfg = get_config()
    path = (save_path or cfg.store_overview.save_path or "").strip()
    if not path:
        return {"indicators": [], "periods": []}

    if re.match(r"^[\\/]+[A-Za-z]:", path):
        path = re.sub(r"^[\\/]+", "", path)
    path = os.path.normpath(path)

    if not os.path.exists(path):
        return {"indicators": [], "periods": []}

    try:
        mtime = os.path.getmtime(path)
    except Exception:
        mtime = 0
    cache_key = f"{path}|{mtime}"
    if cache_key in _STORE_OVERVIEW_CACHE:
        return _STORE_OVERVIEW_CACHE[cache_key]

    def _to_num(v):
        try:
            if pd.isna(v):
                return 0.0
            return float(str(v).replace(",", "").strip())
        except Exception:
            return 0.0

    def _category(name: str) -> str:
        if "曝光" in name:
            return "曝光"
        if "点击" in name:
            return "点击"
        if "商机" in name:
            return "商机"
        return "其他"

    try:
        df_cur = pd.read_excel(path, sheet_name="合并数据")
        df_ind = pd.read_excel(path, sheet_name="行业平均")
        df_peer = pd.read_excel(path, sheet_name="同行优秀")
    except Exception:
        return {"indicators": [], "periods": []}

    if df_cur is None or df_cur.empty or len(df_cur.columns) < 2:
        return {"indicators": [], "periods": []}

    metric_col = df_cur.columns[0]
    latest_col = df_cur.columns[1]

    cur_map = {str(r[metric_col]).strip(): _to_num(r[latest_col]) for _, r in df_cur.iterrows() if pd.notna(r[metric_col])}

    ind_map = {}
    if df_ind is not None and not df_ind.empty and len(df_ind.columns) >= 2:
        mc, lc = df_ind.columns[0], df_ind.columns[1]
        ind_map = {str(r[mc]).strip(): _to_num(r[lc]) for _, r in df_ind.iterrows() if pd.notna(r[mc])}

    peer_map = {}
    if df_peer is not None and not df_peer.empty and len(df_peer.columns) >= 2:
        mc, lc = df_peer.columns[0], df_peer.columns[1]
        peer_map = {str(r[mc]).strip(): _to_num(r[lc]) for _, r in df_peer.iterrows() if pd.notna(r[mc])}

    indicators = []
    for name, value in cur_map.items():
        indicators.append({
            "name": name,
            "current": value,
            "industryAvg": ind_map.get(name, 0.0),
            "peerExcellent": peer_map.get(name, 0.0),
            "category": _category(name),
        })

    periods = []
    for c in df_cur.columns[1:]:
        label = str(c).strip()
        row_exp = df_cur[df_cur[metric_col] == "全店曝光次数"]
        row_clk = df_cur[df_cur[metric_col] == "全店点击次数"]
        row_bus = df_cur[df_cur[metric_col] == "全店品的商机量"]
        periods.append({
            "period": label,
            "exposure": _to_num(row_exp.iloc[0][c]) if not row_exp.empty else 0.0,
            "clicks": _to_num(row_clk.iloc[0][c]) if not row_clk.empty else 0.0,
            "opportunities": _to_num(row_bus.iloc[0][c]) if not row_bus.empty else 0.0,
        })

    result = {"indicators": indicators, "periods": periods[:80]}
    _STORE_OVERVIEW_CACHE.clear()
    _STORE_OVERVIEW_CACHE[cache_key] = result
    return result


def _get_country_code_map() -> Dict[str, str]:
    return {
        "PR": "波多黎各", "PS": "巴勒斯坦", "PT": "葡萄牙", "PW": "帕劳", "PY": "巴拉圭", "QA": "卡塔尔",
        "AD": "安道尔", "AE": "阿联酋", "AF": "阿富汗", "AG": "安提瓜和巴布达", "AI": "安圭拉", "AL": "阿尔巴尼亚",
        "AM": "亚美尼亚", "AN": "荷属安的列斯", "AO": "安哥拉", "AP": "亚洲/太平洋地区", "AQ": "南极洲", "AR": "阿根廷",
        "AS": "美属萨摩亚", "RE": "留尼旺岛", "AT": "奥地利", "AU": "澳大利亚", "AW": "阿鲁巴", "AZ": "阿塞拜疆",
        "RO": "罗马尼亚", "BA": "波黑", "BB": "巴巴多斯", "RS": "塞尔维亚", "BD": "孟加拉国", "RU": "俄罗斯",
        "BE": "比利时", "BF": "布基那法索", "RW": "卢旺达", "BG": "保加利亚", "BH": "巴林", "BI": "布隆迪",
        "BJ": "贝宁", "BM": "百慕大", "BN": "汶莱", "BO": "玻利维亚", "SA": "沙特", "SB": "所罗门群岛",
        "SC": "塞舌尔", "BR": "巴西", "SD": "苏丹", "ASC": "阿森松岛", "BS": "巴哈马", "SE": "瑞典",
        "BT": "不丹", "SG": "新加坡", "BV": "布维岛", "SH": "圣赫勒拿", "BW": "博茨瓦纳", "SI": "斯洛文尼亚",
        "SJ": "斯瓦尔巴群岛和扬马延", "BY": "白俄罗斯", "SK": "斯洛伐克", "BZ": "伯利兹", "SL": "塞拉利昂",
        "SM": "圣马力诺", "SN": "塞内加尔", "SO": "索马里", "CA": "加拿大", "SR": "苏里南", "CC": "科科斯（基林）群岛",
        "ST": "圣多美和普林西", "CD": "刚果民主共和国", "SV": "萨尔瓦多", "CF": "中非共和国", "CG": "刚果",
        "CH": "瑞士", "CI": "科特迪瓦", "SY": "叙利亚", "SZ": "斯威士兰", "CK": "库克群岛", "CL": "智利",
        "CM": "喀麦隆", "CN": "中国", "CO": "哥伦比亚", "TC": "特克斯和凯科斯群岛", "CR": "哥斯达黎加",
        "TD": "乍得", "OTHER": "其它", "CU": "古巴", "TF": "法国南部领土", "CV": "佛得角", "TG": "多哥",
        "TH": "泰国", "CW": "库拉索", "CX": "圣诞岛", "TJ": "塔吉克斯坦", "CY": "塞浦路斯", "TK": "托克劳",
        "CZ": "捷克共和国", "TL": "东帝汶", "TM": "土库曼斯坦", "TN": "突尼斯", "TO": "汤加", "TP": "东帝汶",
        "SGS": "南乔治亚岛和南桑威奇群岛", "TR": "土耳其", "JEY": "泽西岛", "TT": "特里尼达和多巴哥", "DE": "德国",
        "TV": "图瓦卢", "TW": "中国台湾", "TZ": "坦桑尼亚", "DJ": "吉布提", "DK": "丹麦", "DM": "多米尼加",
        "DO": "多米尼加共和国", "UA": "乌克兰", "UG": "乌干达", "UK": "英国", "DZ": "阿尔及利亚",
        "UM": "美国本土外小岛屿", "US": "美国", "EC": "厄瓜多尔", "EE": "爱沙尼亚", "EG": "埃及", "EH": "西撒哈拉",
        "UY": "乌拉圭", "UZ": "乌兹别克斯坦", "VA": "梵蒂冈", "VC": "圣文森特和格林纳丁斯", "ER": "厄立特里亚",
        "ES": "西班牙", "VE": "委内瑞拉", "ET": "埃塞俄比亚", "EU": "欧洲", "VG": "维尔京群岛（英国）",
        "VI": "维尔京群岛（美国）", "VN": "越南", "VU": "瓦努阿图", "FI": "芬兰", "FJ": "斐济", "MNE": "黑山",
        "FK": "福克兰群岛（马尔维纳斯）", "FM": "密克罗尼西亚", "GBA": "奥尔德尼", "FO": "法罗群岛", "FR": "法国",
        "WF": "瓦利斯和富图纳群岛", "FX": "法属美特罗波利坦", "GA": "加蓬", "TLS": "东帝汶", "GB": "英国",
        "WS": "萨摩亚", "GD": "格林纳达", "GE": "格鲁吉亚", "GF": "法属圭亚那", "GG": "格恩西", "GH": "加纳",
        "GI": "直布罗陀", "GL": "格陵兰", "GM": "冈比亚", "GN": "几内亚", "GP": "瓜德罗普岛", "GQ": "赤道几内亚",
        "GR": "希腊", "GS": "南乔治亚岛和南桑威奇群岛", "GT": "危地马拉", "GU": "关岛", "GW": "几内亚比绍",
        "GY": "圭亚那", "HK": "中国香港", "HM": "赫德和麦克唐纳群岛", "HN": "洪都拉斯", "HR": "克罗地亚",
        "none": "无", "HT": "海地", "YE": "也门", "HU": "匈牙利", "YT": "马约特", "ID": "印尼", "YU": "南斯拉夫",
        "IE": "爱尔兰", "IL": "以色列", "IN": "印度", "IO": "英属印度洋领地", "ZA": "南非", "IQ": "伊拉克",
        "IR": "伊朗", "IS": "冰岛", "IT": "意大利", "ZM": "赞比亚", "MAF": "圣马丁岛", "ZR": "刚果民主共和国",
        "ZW": "津巴布韦", "EAZ": "桑给巴尔", "BLM": "圣巴泰勒米岛", "JM": "牙买加", "JO": "约旦", "JP": "日本",
        "KE": "肯尼亚", "KG": "吉尔吉斯斯坦", "KH": "柬埔寨", "KI": "基里巴斯", "KM": "科摩罗", "KN": "圣基茨和尼维斯",
        "KP": "朝鲜", "KR": "韩国", "KS": "科索沃", "KW": "科威特", "KY": "开曼群岛", "KZ": "哈萨克斯坦",
        "LA": "老挝", "ALA": "奥兰群岛", "LB": "黎巴嫩", "LC": "圣卢西亚", "GGY": "格恩西", "LI": "列支敦士登",
        "LK": "斯里兰卡", "LR": "利比里亚", "LS": "莱索托", "LT": "立陶宛", "LU": "卢森堡", "LV": "拉脱维亚",
        "LY": "利比亚", "MA": "摩洛哥", "MC": "摩纳哥", "MD": "摩尔多瓦", "ME": "黑山", "MF": "圣马丁岛",
        "MG": "马达加斯加", "MH": "马绍尔群岛", "MK": "马其顿共和国", "ML": "马里", "MM": "缅甸", "MN": "蒙古",
        "MO": "中国澳门", "MP": "北马里亚纳群岛", "MQ": "马提尼克岛", "MR": "毛里塔尼亚", "MS": "蒙特塞拉特",
        "MT": "马耳他", "MU": "毛里求斯", "MV": "马尔代夫", "MW": "马拉维", "MX": "墨西哥", "MY": "马来西亚",
        "MZ": "莫桑比克", "NA": "纳米比亚", "NC": "新喀里多尼亚", "NE": "尼日尔", "NF": "诺福克岛", "NG": "尼日利亚",
        "NI": "尼加拉瓜", "NL": "荷兰", "NO": "挪威", "NP": "尼泊尔", "SRB": "塞尔维亚", "NR": "瑙鲁", "NU": "纽埃",
        "NZ": "新西兰", "OM": "阿曼", "PA": "巴拿马", "PE": "秘鲁", "PF": "法属波利尼西亚", "PG": "巴布亚新几内亚",
        "PH": "菲律宾", "PK": "巴基斯坦", "PL": "波兰", "PM": "圣皮埃尔和密克隆岛", "PN": "皮特凯恩"
    }
